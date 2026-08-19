# -*- coding: utf-8 -*-
"""Extrai unidades AEB das planilhas Compras Assistência e Educação."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Optional

from openpyxl import load_workbook

from compras_unidades_nome_utils import (
    chave_unidade,
    limpar_cnpj,
    nome_fantasia_de_aba,
    usa_cnpj_matriz,
)

_ABAS_IGNORAR = frozenset({"CALENDARIO", "CALENDARIO ", "CALENDÁRIO", "CALENDÁRIO "})
_RE_CNPJ = re.compile(r"\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}")
_RE_CEP = re.compile(r"CEP[\s\-:]*(\d{5})[\-]?(\d{3})", re.I)
_RE_ENDERECO = re.compile(
    r"Endere[cç]o:\s*(.+?)(?:\.\s*(?:CEP|S[aã]o Paulo)|\s*CEP[\s\-:]|\s*S[aã]o Paulo/SP|$)",
    re.I,
)
_RE_CIDADE_UF = re.compile(r"([A-Za-zÀ-ú\s\.]+)/SP", re.I)


@dataclass
class UnidadePlanilha:
    aba: str
    nome_fantasia: str
    cnpj: Optional[str] = None
    usa_cnpj_matriz: bool = False
    cep: Optional[str] = None
    logradouro: Optional[str] = None
    numero: Optional[str] = None
    bairro: Optional[str] = None
    cidade: Optional[str] = None
    uf: Optional[str] = "SP"
    fonte: str = ""
    observacoes_parse: list[str] = field(default_factory=list)


def _cel(valor: object) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _titulo_aba(ws) -> str:
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=6, values_only=True):
        for idx, cell in enumerate(row):
            if idx == 0:
                continue
            texto = _cel(cell)
            if texto and len(texto) > 2 and "DATA DO PEDIDO" not in texto.upper():
                if _RE_CNPJ.search(texto):
                    continue
                return texto
    return ""


def _blob_inicial(ws, max_row: int = 25) -> str:
    partes: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=8, values_only=True):
        partes.extend(_cel(c) for c in row if c not in (None, ""))
    return " ".join(partes).replace("\n", " ")


def _extrair_cnpj(blob: str) -> Optional[str]:
    for bruto in _RE_CNPJ.findall(blob):
        limpo = limpar_cnpj(bruto)
        if limpo:
            return limpo
    return None


def _parse_endereco(blob: str) -> dict[str, Optional[str]]:
    out: dict[str, Optional[str]] = {
        "cep": None,
        "logradouro": None,
        "numero": None,
        "bairro": None,
        "cidade": "São Paulo",
        "uf": "SP",
    }
    cep_m = _RE_CEP.search(blob)
    if cep_m:
        out["cep"] = f"{cep_m.group(1)}{cep_m.group(2)}"

    end_m = _RE_ENDERECO.search(blob)
    texto_end = (end_m.group(1) if end_m else "").strip(" .|")
    if not texto_end:
        # fallback: trecho após CNPJ com padrão "Rua ..."
        pos = blob.upper().find("RUA ")
        if pos < 0:
            pos = blob.upper().find("AV")
        if pos >= 0:
            texto_end = blob[pos:].split("CEP")[0].strip(" .")
    if not texto_end:
        return out

    texto_end = re.sub(r"\|.*?(?=Rua|Av|Avenida|Estrada)", "", texto_end, flags=re.I).strip()
    cidade_m = _RE_CIDADE_UF.search(texto_end)
    if cidade_m:
        out["cidade"] = cidade_m.group(1).strip(" .")

    # "Rua X, 123 - Bairro" ou "Av Y, 94 – Bairro"
    match = re.match(
        r"^(?P<log>.+?),\s*(?P<num>[\dA-Za-z/S\-]+)\s*[–\-]\s*(?P<bairro>.+?)(?:\.\s*S[aã]o Paulo|$)",
        texto_end,
        re.I,
    )
    if match:
        out["logradouro"] = match.group("log").strip()
        out["numero"] = match.group("num").strip()
        out["bairro"] = match.group("bairro").strip(" .")
    else:
        out["logradouro"] = texto_end[:180].strip(" .")

    return out


def _parse_aba(ws, aba: str, fonte: str) -> Optional[UnidadePlanilha]:
    blob = _blob_inicial(ws)
    cnpj = _extrair_cnpj(blob)
    titulo = _titulo_aba(ws)
    nome = nome_fantasia_de_aba(aba, titulo)
    end = _parse_endereco(blob)
    return UnidadePlanilha(
        aba=aba.strip(),
        nome_fantasia=nome,
        cnpj=cnpj,
        usa_cnpj_matriz=usa_cnpj_matriz(cnpj),
        cep=end["cep"],
        logradouro=end["logradouro"],
        numero=end["numero"],
        bairro=end["bairro"],
        cidade=end["cidade"],
        uf=end["uf"],
        fonte=fonte,
    )


def extrair_unidades_xlsx(conteudo: bytes, nome_arquivo: str) -> list[UnidadePlanilha]:
    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    saida: list[UnidadePlanilha] = []
    for aba in wb.sheetnames:
        if aba.strip().upper() in {x.strip().upper() for x in _ABAS_IGNORAR}:
            continue
        if aba.strip().upper().startswith("CALEND"):
            continue
        parsed = _parse_aba(wb[aba], aba, nome_arquivo)
        if parsed and parsed.nome_fantasia:
            saida.append(parsed)
    wb.close()
    return saida


def mesclar_unidades(registros: list[UnidadePlanilha]) -> list[UnidadePlanilha]:
    """Une Assistência + Educação pela chave canônica; preferir registro com endereço."""
    mapa: dict[str, UnidadePlanilha] = {}
    for reg in registros:
        chave = chave_unidade(reg.nome_fantasia)
        if not chave:
            continue
        atual = mapa.get(chave)
        if not atual:
            mapa[chave] = reg
            continue
        def score(u: UnidadePlanilha) -> int:
            s = 0
            if u.logradouro:
                s += 2
            if u.cep:
                s += 1
            if u.cnpj and not u.usa_cnpj_matriz:
                s += 1
            return s
        if score(reg) > score(atual):
            mapa[chave] = reg
        elif score(reg) == score(atual):
            # complementa campos vazios
            for campo in ("cnpj", "cep", "logradouro", "numero", "bairro", "cidade", "uf"):
                if not getattr(atual, campo) and getattr(reg, campo):
                    setattr(atual, campo, getattr(reg, campo))
            atual.fonte = f"{atual.fonte}; {reg.fonte}"
    return sorted(mapa.values(), key=lambda u: u.nome_fantasia.upper())


def extrair_unidades_compras_aeb(
    caminho_assistencia: Optional[object] = None,
    caminho_educacao: Optional[object] = None,
) -> list[UnidadePlanilha]:
    from pathlib import Path

    root = Path.home() / "Downloads" / "Arquivos Módulo Compras"
    arquivos: list[tuple[Path, str]] = []
    assist = Path(caminho_assistencia) if caminho_assistencia else None
    educ = Path(caminho_educacao) if caminho_educacao else None
    if not assist:
        for cand in root.rglob("*ASSISTENCIA*2026*.xlsx"):
            assist = cand
            break
    if not educ:
        for cand in root.rglob("*EDUCA*2026*.xlsx"):
            educ = cand
            break
    if assist and assist.exists():
        arquivos.append((assist, assist.name))
    if educ and educ.exists():
        arquivos.append((educ, educ.name))

    bruto: list[UnidadePlanilha] = []
    for path, nome in arquivos:
        bruto.extend(extrair_unidades_xlsx(path.read_bytes(), nome))
    return mesclar_unidades(bruto)
