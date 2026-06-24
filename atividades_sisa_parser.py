from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from fastapi import HTTPException

COL_DIMENSAO = 3
COL_DESCRICAO_ATIVIDADE = 7
COL_DESCRICAO_TEMA = 10
COL_DATA_ATIVIDADE = 15
COL_HORARIO = 18
COL_PARTICIPACOES = 19
LINHA_CABECALHO_MIN = 8


def _texto_planilha(valor: Any) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def normalizar_texto_sisa(valor: str | None) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
    texto = re.sub(r"\s+", " ", texto.upper().strip())
    return texto


def normalizar_horario_sisa(valor: str | None) -> str:
    texto = normalizar_texto_sisa(valor)
    if not texto:
        return ""
    texto = texto.replace(" AS ", " A ").replace(" A ", " - ")
    texto = re.sub(r"\s*-\s*", "-", texto)
    horarios = re.findall(r"\d{1,2}:\d{2}", texto)
    if len(horarios) >= 2:
        return f"{horarios[0]}-{horarios[1]}"
    if len(horarios) == 1:
        return horarios[0]
    return texto


def _data_br_para_date(valor: str) -> date | None:
    match = re.search(r"(\d{2})/(\d{2})/(\d{4})", str(valor or ""))
    if not match:
        return None
    dia, mes, ano = map(int, match.groups())
    try:
        return date(ano, mes, dia)
    except ValueError:
        return None


def _extrair_periodo_referencia_sisa(texto: str) -> tuple[date | None, date | None]:
    datas = [
        _data_br_para_date(item)
        for item in re.findall(r"\d{2}/\d{2}/\d{4}", str(texto or ""))
    ]
    datas = [item for item in datas if item]
    if len(datas) >= 2:
        return datas[0], datas[1]
    if datas:
        return datas[0], datas[0]
    return None, None


def _parse_data_celula(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, (int, float)) and valor:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(valor))).date()
    return _data_br_para_date(str(valor or ""))


def _parse_participacoes(valor: Any) -> int:
    texto = _texto_planilha(valor)
    if not texto or texto == "---":
        return 0
    try:
        return max(int(float(texto.replace(",", "."))), 0)
    except (TypeError, ValueError):
        return 0


def _linha_parece_total(sheet, indice: int) -> bool:
    texto = " ".join(
        _texto_planilha(sheet.cell_value(indice, coluna)) or ""
        for coluna in range(sheet.ncols)
    ).upper()
    return (
        "TOTAL DE ATIVIDADES" in texto
        or "TOTAL DE PARTICIPACOES" in texto
        or "TOTAL DE PARTICIPAÇÕES" in texto
        or "TOTAL DE CIDADAOS" in texto
        or "TOTAL DE CIDADÃOS" in texto
    )


def _linha_resumo_para_registro(sheet, indice: int) -> dict | None:
    descricao_atividade = _texto_planilha(sheet.cell_value(indice, COL_DESCRICAO_ATIVIDADE))
    descricao_tema = _texto_planilha(sheet.cell_value(indice, COL_DESCRICAO_TEMA))
    data_sessao = _parse_data_celula(sheet.cell_value(indice, COL_DATA_ATIVIDADE))
    horario = _texto_planilha(sheet.cell_value(indice, COL_HORARIO))

    if not descricao_atividade and not descricao_tema:
        return None
    if not data_sessao:
        return None

    dimensao = _texto_planilha(sheet.cell_value(indice, COL_DIMENSAO)) or ""
    participacoes = _parse_participacoes(sheet.cell_value(indice, COL_PARTICIPACOES))
    horario_norm = normalizar_horario_sisa(horario)
    tipo_norm = normalizar_texto_sisa(descricao_atividade)
    tema_norm = normalizar_texto_sisa(descricao_tema)
    chave = f"{data_sessao.isoformat()}|{horario_norm}|{tipo_norm}|{tema_norm}"

    return {
        "chave": chave,
        "dimensao": dimensao,
        "descricao_atividade": descricao_atividade or "",
        "descricao_tema": descricao_tema or "",
        "descricao_atividade_norm": tipo_norm,
        "descricao_tema_norm": tema_norm,
        "data_sessao": data_sessao,
        "horario": horario or "",
        "horario_norm": horario_norm,
        "participacoes_sisa": participacoes,
    }


class _SheetAdapter:
    def __init__(self, rows: list[tuple]):
        self.rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(row) for row in rows), default=0)

    def cell_value(self, row_index: int, col_index: int):
        try:
            return self.rows[row_index][col_index]
        except IndexError:
            return None


def _ler_planilha_xls(conteudo: bytes) -> _SheetAdapter:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Leitor de planilha .xls não instalado. Rode pip install -r requirements.txt.",
        ) from exc

    workbook = xlrd.open_workbook(file_contents=conteudo)
    sheet = workbook.sheet_by_index(0)
    rows = [tuple(sheet.row_values(indice)) for indice in range(sheet.nrows)]
    return _SheetAdapter(rows)


def _ler_planilha_xlsx(conteudo: bytes) -> _SheetAdapter:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Leitor de planilha .xlsx não instalado. Rode pip install -r requirements.txt.",
        ) from exc

    workbook = load_workbook(BytesIO(conteudo), data_only=True)
    sheet = workbook.active
    rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    return _SheetAdapter(rows)


def ler_relatorio_resumo_atividades_sisa(conteudo: bytes, nome_arquivo: str) -> dict:
    extensao = (nome_arquivo or "").lower().rsplit(".", 1)[-1]
    if extensao == "xls":
        sheet = _ler_planilha_xls(conteudo)
    elif extensao == "xlsx":
        sheet = _ler_planilha_xlsx(conteudo)
    else:
        raise HTTPException(
            status_code=400,
            detail="Envie o Relatório Resumo de Atividades exportado do SISA (.xls ou .xlsx).",
        )

    data_inicio_referencia = None
    data_fim_referencia = None
    servico = None
    projeto = None
    totais: dict[str, int] = {}

    for indice in range(min(12, sheet.nrows)):
        texto_linha = " ".join(
            _texto_planilha(sheet.cell_value(indice, coluna)) or ""
            for coluna in range(sheet.ncols)
        )
        inicio, fim = _extrair_periodo_referencia_sisa(texto_linha)
        if inicio and fim and not data_fim_referencia:
            data_inicio_referencia = inicio
            data_fim_referencia = fim
        if "Tipo de Servi" in texto_linha or "SIAT" in texto_linha:
            servico = texto_linha.strip()
            if "SIAT" in texto_linha:
                projeto = texto_linha.split("/")[0].strip() if "/" in texto_linha else texto_linha.strip()

    linhas = []
    for indice in range(LINHA_CABECALHO_MIN, sheet.nrows):
        if _linha_parece_total(sheet, indice):
            texto = " ".join(
                _texto_planilha(sheet.cell_value(indice, coluna)) or ""
                for coluna in range(sheet.ncols)
            )
            numeros = re.findall(r"\d+", texto)
            if "Total de Atividades" in texto and numeros:
                totais["total_atividades"] = int(numeros[-1])
            if "Total de Particip" in texto and numeros:
                totais["total_participacoes"] = int(numeros[-1])
            if "Total de Cidad" in texto and numeros:
                totais["total_cidadaos"] = int(numeros[-1])
            continue

        registro = _linha_resumo_para_registro(sheet, indice)
        if registro:
            linhas.append(registro)

    if not data_inicio_referencia or not data_fim_referencia:
        raise HTTPException(
            status_code=400,
            detail="Não foi possível identificar o período de referência da planilha SISA.",
        )

    if not linhas:
        raise HTTPException(
            status_code=400,
            detail="Nenhuma linha de atividade foi encontrada na planilha SISA.",
        )

    return {
        "data_inicio_referencia": data_inicio_referencia,
        "data_fim_referencia": data_fim_referencia,
        "servico": servico,
        "projeto": projeto,
        "totais_sisa": totais,
        "linhas": linhas,
    }
