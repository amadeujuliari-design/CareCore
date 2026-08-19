# -*- coding: utf-8 -*-
"""Leitura do inventário de bens AEB (xlsx por unidade)."""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from compras_patrimonio_utils import (
    etiqueta_texto,
    normalizar_origem,
    normalizar_propriedade,
    normalizar_situacao,
    parse_data_aquisicao,
    reais_para_centavos,
)
from compras_regras import ESCOPO_PROJETO, ESCOPO_SEDE, PATRIMONIO_ORIGEM_INVENTARIO
from compras_unidades_nome_utils import chave_unidade, rotulo_nfp_canonico, _norm_chave

MAPA_INVENTARIO_UNIDADE = {
    "SEDE": "SEDE AEB",
    "AEB SEDE": "SEDE AEB",
    "CEI V.N.CACHOEIRINHA": "CEI VILA NOVA CACHOEIRINHA",
    "CEI LEOPOLDINA": "CEI VILA LEOPOLDINA",
    "CEI BELEM": "CEI BELÉM",
    "CEI BELÉM": "CEI BELÉM",
    "CEI MONTE AZUL": "CEI MONTE AZUL",
    "CEI V GUSTAVO": "CEI VILA GUSTAVO",
    "CEI VILA GUSTAVO": "CEI VILA GUSTAVO",
    "CEI LIBERDADE": "CEI LIBERDADE",
    "CASA PORTO": "CASA PORTO SEGURO",
    "CRIAR E TOCAR": "CRIAR & TOCAR",
    "AEB- SIAT II ARMENIA": "SIAT II ARMÊNIA",
    "AEB- SIAT II ARMÊNIA": "SIAT II ARMÊNIA",
    "SIAT II": "SIAT II ARMÊNIA",
    "CTA CANINDE": "CTA 18 – CANINDÉ",
    "CTA LIBERDADE": "CTA 17 – LIBERDADE",
    "CAEF HOTEL RIVOLI": "CAE F RIVOLI",
    "CAEF PAULICEIA": "CAE F PAULICEIA",
}


@dataclass
class ItemPatrimonioPlanilha:
    aba: str
    unidade_planilha: str
    nome_canonico: str
    escopo_unidade: str
    descricao: str
    numero_etiqueta: Optional[str] = None
    localizacao: Optional[str] = None
    departamento: Optional[str] = None
    propriedade: str = "aeb"
    data_aquisicao: Optional[date] = None
    documento_nf: Optional[str] = None
    forma_aquisicao: Optional[str] = None
    valor_centavos: Optional[int] = None
    situacao: str = "bom"
    motivo_baixa: Optional[str] = None
    data_baixa: Optional[date] = None
    origem: str = PATRIMONIO_ORIGEM_INVENTARIO


def _cel(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def nome_unidade_inventario(aba: str, titulo_celula: Optional[str], projeto_celula: Optional[str] = None) -> str:
    for candidato in (titulo_celula, projeto_celula, aba):
        bruto = (candidato or "").strip()
        if not bruto or bruto.upper().startswith("F") and bruto[1:].replace("T", "").isdigit():
            continue
        chave = _norm_chave(bruto)
        for k, v in MAPA_INVENTARIO_UNIDADE.items():
            if _norm_chave(k) == chave:
                return v
        rotulo = rotulo_nfp_canonico(bruto)
        if rotulo:
            return rotulo
    return (titulo_celula or aba or "").strip()


def unidade_eh_sede(nome: str) -> bool:
    chave = _norm_chave(nome)
    return chave in {"SEDE", "SEDE AEB", "AEB SEDE"} or chave.startswith("SEDE")


def _doc_nf(valor) -> Optional[str]:
    texto = _cel(valor)
    if not texto or texto in {"*", "-", "."}:
        return None
    return texto[:120]


def _linha_item(row, aba: str, unidade_aba: str) -> Optional[ItemPatrimonioPlanilha]:
    material = _cel(row[2] if len(row) > 2 else "")
    if not material or material.upper() in {"MATERIAL", "DESCRIÇÃO DO ITEM", "DESCRICAO DO ITEM"}:
        return None
    if material.upper().startswith("ASSOCIA"):
        return None
    projeto_txt = _cel(row[5] if len(row) > 5 else "")
    nome = nome_unidade_inventario(aba, unidade_aba, projeto_txt)
    if unidade_eh_sede(nome) or unidade_eh_sede(unidade_aba):
        escopo = ESCOPO_SEDE
        nome = "SEDE AEB"
    else:
        escopo = ESCOPO_PROJETO
    data_baixa = parse_data_aquisicao(row[13] if len(row) > 13 else None)
    motivo = _cel(row[12] if len(row) > 12 else "") or None
    forma = _cel(row[9] if len(row) > 9 else "") or None
    return ItemPatrimonioPlanilha(
        aba=aba,
        unidade_planilha=unidade_aba,
        nome_canonico=nome,
        escopo_unidade=escopo,
        descricao=material[:240],
        numero_etiqueta=etiqueta_texto(row[1] if len(row) > 1 else None),
        localizacao=_cel(row[3] if len(row) > 3 else "") or None,
        departamento=_cel(row[4] if len(row) > 4 else "") or None,
        propriedade=normalizar_propriedade(_cel(row[6] if len(row) > 6 else "")),
        data_aquisicao=parse_data_aquisicao(row[7] if len(row) > 7 else None),
        documento_nf=_doc_nf(row[8] if len(row) > 8 else None),
        forma_aquisicao=forma,
        valor_centavos=reais_para_centavos(row[10] if len(row) > 10 else None),
        situacao=normalizar_situacao(_cel(row[11] if len(row) > 11 else ""), data_baixa=data_baixa),
        motivo_baixa=motivo,
        data_baixa=data_baixa,
        origem=normalizar_origem(None, forma=forma) if forma else PATRIMONIO_ORIGEM_INVENTARIO,
    )


def extrair_itens_inventario(conteudo: bytes, nome_arquivo: str = "") -> list[ItemPatrimonioPlanilha]:
    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    saida: list[ItemPatrimonioPlanilha] = []
    for aba in wb.sheetnames:
        if "ndice" in aba.lower():
            continue
        ws = wb[aba]
        unidade_aba = aba
        for i, row in enumerate(ws.iter_rows(max_col=14, values_only=True)):
            if i == 0:
                unidade_aba = _cel(row[11] if len(row) > 11 else "") or aba
                continue
            if i < 3:
                continue
            item = _linha_item(row, aba, unidade_aba)
            if item:
                saida.append(item)
    wb.close()
    return saida


def caminho_inventario_padrao() -> Path:
    return Path.home() / "Downloads" / "Arquivos Módulo Compras" / "INVENTÁRIO PATRIMONIO AEB E PUBLICO (1).xlsx"
