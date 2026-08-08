"""Extracao de valores Digitados da planilha METAS NFP (abas mensais).

Nao depende de celulas fixas tipo F35/B42 — localiza por rotulos, porque
o bloco Soulcial/Captador muda de linha entre meses.
"""

from __future__ import annotations

import re
from typing import Any, Optional

from openpyxl.worksheet.worksheet import Worksheet

from nfp_metas_utils import codigo_projeto_metas, ref_credito_padrao

MESES_PT = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARCO": 3,
    "MARÇO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}

COMPETENCIA_POR_ABA = {
    "MARCO": "2026-03",
    "MARÇO": "2026-03",
    "ABRIL": "2026-04",
    "MAIO": "2026-05",
    "JUNHO": "2026-06",
    "JULHO": "2026-07",
    "AGOSTO": "2026-08",
}


def _norm(texto: Any) -> str:
    s = str(texto or "").strip().upper()
    s = s.replace("Ç", "C").replace("Ã", "A").replace("É", "E").replace("Ê", "E")
    s = s.replace("\n", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def _float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _int(v: Any, default: int = 0) -> int:
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except (TypeError, ValueError):
        return default


def competencia_da_aba(nome_aba: str) -> Optional[str]:
    n = _norm(nome_aba)
    for chave, comp in COMPETENCIA_POR_ABA.items():
        if _norm(chave) in n and "CONSOLID" not in n:
            return comp
    return None


def parse_ref_do_titulo(titulo: Any, competencia: str) -> str:
    """Le 'JULHO 2026 - REF MARCO' / 'MAIO 2026 - REF JANEIRO 2026'."""
    texto = str(titulo or "")
    m = re.search(r"REF\s*([A-ZÇÃÉÊÕÁÍÓÚ]+)?\s*(\d{4})?", _norm(texto), flags=re.I)
    if not m or not (m.group(1) or "").strip():
        return ref_credito_padrao(competencia)

    mes_nome = m.group(1).strip()
    # normaliza MARCO sem cedilha
    mes_num = None
    for nome, num in MESES_PT.items():
        if _norm(nome) == _norm(mes_nome) or _norm(nome).startswith(_norm(mes_nome)[:3]):
            mes_num = num
            break
    if not mes_num:
        return ref_credito_padrao(competencia)

    ano_comp = int(competencia[:4])
    mes_comp = int(competencia[5:7])
    if m.group(2):
        ano_ref = int(m.group(2))
    else:
        ano_ref = ano_comp - 1 if mes_num > mes_comp else ano_comp
    return f"{ano_ref:04d}-{mes_num:02d}"


def _valor_ao_lado_do_rotulo(ws: Worksheet, rotulo_exato: str, max_r: int = 55) -> Optional[float]:
    alvo = _norm(rotulo_exato)
    for r in range(1, max_r + 1):
        for c in range(1, 12):
            if _norm(ws.cell(r, c).value) == alvo:
                return _float(ws.cell(r, c + 1).value)
    return None


def _digitadas_diego(ws: Worksheet, max_r: int = 40) -> int:
    for r in range(1, max_r + 1):
        if _norm(ws.cell(r, 1).value) == "DIEGO":
            return _int(ws.cell(r, 2).value)
    return 0


def _bloco_soulcial_captador(ws: Worksheet) -> tuple[float, float]:
    """Retorna (soulcial_base, total_captador) a partir do bloco SOUCIAL."""
    for r in range(1, 55):
        b = _norm(ws.cell(r, 2).value)
        if "SOU" in b and "CIAL" in b and "20%" not in b:
            for r2 in range(r + 1, r + 6):
                if "TOTAL GERAL" in _norm(ws.cell(r2, 2).value):
                    # linha de valores = proxima
                    rv = r2 + 1
                    return _float(ws.cell(rv, 2).value), _float(ws.cell(rv, 10).value)
    return 0.0, 0.0


def _linhas_projetos(ws: Worksheet) -> list[dict]:
    linhas = []
    for r in range(3, 40):
        nome = ws.cell(r, 1).value
        if not nome:
            continue
        s = str(nome).strip()
        nu = _norm(s)
        if nu in {"TOTAL", "DIEGO"} or nu.startswith("TOTAL"):
            break
        codigo = codigo_projeto_metas(s) or s
        linhas.append(
            {
                "codigo_projeto": codigo,
                "digitadas": _int(ws.cell(r, 2).value),
                "doadas": _int(ws.cell(r, 4).value),
                "soulcial": _float(ws.cell(r, 9).value),
                "soulcial_campanhas": _float(ws.cell(r, 10).value),
            }
        )
    return linhas


def extrair_competencia_planilha(ws: Worksheet, competencia: str, titulo_aba: str = "") -> dict:
    titulo = ws["A1"].value or titulo_aba
    f35 = _valor_ao_lado_do_rotulo(ws, "DIGITADO")
    f36 = _valor_ao_lado_do_rotulo(ws, "DOADO")
    soulcial_base, total_captador = _bloco_soulcial_captador(ws)
    return {
        "competencia": competencia,
        "ref_credito": parse_ref_do_titulo(titulo, competencia),
        "titulo": str(titulo or "").strip() or f"METAS NFP {competencia}",
        "cabecalho": {
            "f35_digitado": float(f35 or 0),
            "f36_doado": float(f36 or 0),
            "soulcial_base": soulcial_base,
            "total_captador": total_captador,
            "digitadas_diego": _digitadas_diego(ws),
        },
        "linhas": _linhas_projetos(ws),
    }


def listar_abas_mensais(wb) -> list[tuple[str, str]]:
    """[(competencia, nome_aba), ...] ordenado."""
    achados = []
    for nome in wb.sheetnames:
        comp = competencia_da_aba(nome)
        if comp:
            achados.append((comp, nome))
    return sorted(achados, key=lambda x: x[0])
