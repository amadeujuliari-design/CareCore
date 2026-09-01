"""Serviço de conferência NFS-e emitidas × PDFs locais (Finanças)."""

from __future__ import annotations

import json
import re
import shutil
import threading
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from time_operacional import agora_operacional_naive

_CONFIG_DIR = Path("uploads/financeiro/nfse_conferencia")
_CONFIG_DIR.mkdir(parents=True, exist_ok=True)

_RE_DATA_HORA = re.compile(
    r"(\d{2}/\d{2}/\d{4})\s*(?:às\s*)?(\d{2}:\d{2}:\d{2})",
    re.IGNORECASE,
)
_RE_CHAVE_ACESSO = re.compile(r"\d{50}")
_RE_NUMERO_NFSE_CHAVE = re.compile(r"0{8,}(\d{3})260")
_RE_CNPJ = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}")

_job_lock = threading.Lock()
_job: dict[str, Any] = {
    "status": "idle",
    "mensagem": "Aguardando.",
    "log": [],
    "resumo": {
        "registros_site": 0,
        "ok": 0,
        "cancelada": 0,
        "cancelada_sem_pdf": 0,
        "nao_encontrada": 0,
        "ja_conferida": 0,
        "pdf_sem_portal": 0,
        "pdf_sem_chave": 0,
        "total": 0,
    },
    "browser_aberto": False,
    "aguardando_login": False,
}


def _config_path(org_id: str) -> Path:
    return _CONFIG_DIR / f"config_{org_id}.json"


RITMOS_NFSE: dict[str, dict[str, float | int]] = {
    "normal": {
        "intervalo_notas_min_seg": 3,
        "intervalo_notas_max_seg": 6,
        "pausa_a_cada_notas": 10,
        "pausa_longa_min_seg": 45,
        "pausa_longa_max_seg": 90,
        "intervalo_pagina_min_seg": 5,
        "intervalo_pagina_max_seg": 10,
        "intervalo_mes_min_seg": 12,
        "intervalo_mes_max_seg": 20,
        "slow_mo_ms": 120,
    },
    "lento": {
        "intervalo_notas_min_seg": 6,
        "intervalo_notas_max_seg": 12,
        "pausa_a_cada_notas": 8,
        "pausa_longa_min_seg": 90,
        "pausa_longa_max_seg": 150,
        "intervalo_pagina_min_seg": 10,
        "intervalo_pagina_max_seg": 18,
        "intervalo_mes_min_seg": 20,
        "intervalo_mes_max_seg": 35,
        "slow_mo_ms": 200,
    },
    "muito_lento": {
        "intervalo_notas_min_seg": 12,
        "intervalo_notas_max_seg": 20,
        "pausa_a_cada_notas": 5,
        "pausa_longa_min_seg": 120,
        "pausa_longa_max_seg": 180,
        "intervalo_pagina_min_seg": 15,
        "intervalo_pagina_max_seg": 25,
        "intervalo_mes_min_seg": 30,
        "intervalo_mes_max_seg": 45,
        "slow_mo_ms": 350,
    },
}


def resolver_ritmo_nfse(config: dict[str, Any]) -> dict[str, float | int]:
    nome = str(config.get("ritmo") or "lento").lower().replace("-", "_")
    if nome not in RITMOS_NFSE:
        nome = "lento"
    return dict(RITMOS_NFSE[nome])


def resolver_periodo_meses(config: dict[str, Any]) -> tuple[int, int, int]:
    """Retorna (ano, mês_inicial, mês_final) normalizado para o calendário operacional."""
    ano = int(config.get("ano") or date.today().year)
    mes_inicio = max(1, min(12, int(config.get("mes_inicio") or 1)))
    mes_fim = max(1, min(12, int(config.get("mes_fim") or 12)))
    if mes_inicio > mes_fim:
        mes_inicio, mes_fim = mes_fim, mes_inicio
    hoje = date.today()
    if ano == hoje.year:
        mes_fim = min(mes_fim, hoje.month)
    if ano > hoje.year:
        mes_fim = mes_inicio
    if mes_fim < mes_inicio:
        mes_fim = mes_inicio
    return ano, mes_inicio, mes_fim


def pdf_pertence_periodo(
    pdf: Path,
    indice: "IndicePdf",
    *,
    ano: int,
    mes_inicio: int,
    mes_fim: int,
) -> bool:
    """Filtra PDFs do período conferido (evita falso «sem portal» em rodada parcial)."""
    for chave_data, caminho in indice.mapa_data_hora.items():
        if caminho != pdf:
            continue
        try:
            _dia, mes, ano_pdf = chave_data.split()[0].split("/")
            if int(ano_pdf) == ano and mes_inicio <= int(mes) <= mes_fim:
                return True
        except (ValueError, IndexError):
            continue
    nome = pdf.name
    sufixo_ano = str(ano)[2:]
    for mes in range(mes_inicio, mes_fim + 1):
        if f"{mes:02d}-{ano}" in nome or f"{mes:02d}-{sufixo_ano}" in nome:
            return True
    return False


def carregar_config(org_id: str) -> dict[str, Any]:
    caminho = _config_path(org_id)
    if not caminho.is_file():
        return {
            "pastas_origem": [],
            "pasta_destino": "",
            "ano": datetime.now().year,
            "mes_inicio": 1,
            "mes_fim": 12,
            "ritmo": "lento",
        }
    try:
        dados = json.loads(caminho.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        dados = {}
    ritmo = str(dados.get("ritmo") or "lento").lower()
    if ritmo not in RITMOS_NFSE:
        ritmo = "lento"
    return {
        "pastas_origem": list(dados.get("pastas_origem") or []),
        "pasta_destino": str(dados.get("pasta_destino") or ""),
        "ano": int(dados.get("ano") or datetime.now().year),
        "mes_inicio": int(dados.get("mes_inicio") or 1),
        "mes_fim": int(dados.get("mes_fim") or 12),
        "ritmo": ritmo,
    }


def salvar_config(org_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    ritmo = str(payload.get("ritmo") or "lento").lower()
    if ritmo not in RITMOS_NFSE:
        ritmo = "lento"
    config = {
        "pastas_origem": [
            str(p).strip()
            for p in (payload.get("pastas_origem") or [])
            if str(p).strip()
        ],
        "pasta_destino": str(payload.get("pasta_destino") or "").strip(),
        "ano": int(payload.get("ano") or datetime.now().year),
        "mes_inicio": int(payload.get("mes_inicio") or 1),
        "mes_fim": int(payload.get("mes_fim") or 12),
        "ritmo": ritmo,
    }
    _, mi, mf = resolver_periodo_meses(config)
    config["mes_inicio"] = mi
    config["mes_fim"] = mf
    _config_path(org_id).write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return config


def snap_job() -> dict[str, Any]:
    with _job_lock:
        return dict(_job)


def set_job(**kwargs: Any) -> None:
    with _job_lock:
        _job.update(kwargs)
        if "mensagem" in kwargs:
            logs = list(_job.get("log") or [])
            logs.append(str(kwargs["mensagem"]))
            _job["log"] = logs[-200:]


def reset_job() -> None:
    with _job_lock:
        _job.clear()
        _job.update(
            {
                "status": "idle",
                "mensagem": "Aguardando.",
                "log": [],
                "resumo": {
                    "registros_site": 0,
                    "ok": 0,
                    "cancelada": 0,
                    "cancelada_sem_pdf": 0,
                    "nao_encontrada": 0,
                    "ja_conferida": 0,
                    "pdf_sem_portal": 0,
                    "pdf_sem_chave": 0,
                    "total": 0,
                },
                "browser_aberto": False,
                "aguardando_login": False,
            }
        )


def marcar_parar() -> None:
    set_job(parar=True, mensagem="Parada solicitada…")


def deve_parar() -> bool:
    with _job_lock:
        return bool(_job.get("parar"))


def normalizar_data_hora(texto: str | None) -> str | None:
    if not texto:
        return None
    match = _RE_DATA_HORA.search(texto.replace("-03:00", "").replace("-02:00", ""))
    if not match:
        return None
    return f"{match.group(1)} {match.group(2)}"


def extrair_chave_acesso(texto: str | None) -> str | None:
    """Extrai chave de acesso NFS-e (50 dígitos) de portal ou PDF."""
    if not texto:
        return None
    compacto = re.sub(r"\s+", "", texto)
    match = _RE_CHAVE_ACESSO.search(compacto)
    return match.group(0) if match else None


def extrair_numero_nfse(chave: str | None) -> int | None:
    """Extrai o número sequencial da NFS-e a partir da chave de acesso."""
    if not chave:
        return None
    match = _RE_NUMERO_NFSE_CHAVE.search(chave)
    return int(match.group(1)) if match else None


def extrair_tomador_cnpj(texto: str | None) -> str | None:
    """CNPJ/CPF do tomador na DANFSe ou na página do portal."""
    if not texto:
        return None
    compacto = re.sub(r"\s+", " ", texto)
    idx = compacto.lower().find("tomador")
    trecho = compacto[idx : idx + 420] if idx >= 0 else compacto
    cnpjs = _RE_CNPJ.findall(trecho)
    return cnpjs[0] if cnpjs else None


def montar_observacao_cancelada_sem_pdf(
    sugestao: Path | None,
    *,
    indice: "IndicePdf",
) -> str:
    base = "Cancelada no portal; PDF desta chave não existe nas pastas de origem."
    if not sugestao:
        return base
    chave = indice.mapa_pdf_chave.get(sugestao)
    numero = extrair_numero_nfse(chave)
    rotulo = f" (nota {numero})" if numero else ""
    return f"{base} Possível substituta: {sugestao.name}{rotulo}."


def sugerir_pdf_substituto(
    indice: "IndicePdf",
    *,
    texto_portal: str,
    chave_atual: str | None,
    janela_segundos: int = 600,
) -> Path | None:
    """Sugere PDF substituto por horário de emissão e, se possível, mesmo tomador."""
    datas = extrair_datas_portal(texto_portal)
    if not datas:
        return None
    try:
        dt_ref = datetime.strptime(datas[0], "%d/%m/%Y %H:%M:%S")
    except ValueError:
        return None

    tomador = extrair_tomador_cnpj(texto_portal)
    melhor: tuple[int, Path] | None = None

    for chave_data, pdf in indice.mapa_data_hora.items():
        chave_pdf = indice.mapa_pdf_chave.get(pdf)
        if chave_pdf and chave_pdf == chave_atual:
            continue
        try:
            dt_pdf = datetime.strptime(chave_data, "%d/%m/%Y %H:%M:%S")
        except ValueError:
            continue
        delta = abs(int((dt_pdf - dt_ref).total_seconds()))
        if delta > janela_segundos:
            continue
        score = delta
        tom_pdf = indice.mapa_tomador.get(pdf)
        if tomador and tom_pdf == tomador:
            score -= 120
        if melhor is None or score < melhor[0]:
            melhor = (score, pdf)

    return melhor[1] if melhor else None


def extrair_datas_portal(texto_pagina: str) -> list[str]:
    encontradas: list[str] = []
    for rotulo in (
        "Data de geração",
        "Data de emissão",
        "Data e Hora da emissão da NFS-e",
        "Data e Hora da emissão da DPS",
    ):
        idx = texto_pagina.find(rotulo)
        if idx < 0:
            continue
        trecho = texto_pagina[idx : idx + 120]
        normalizada = normalizar_data_hora(trecho)
        if normalizada and normalizada not in encontradas:
            encontradas.append(normalizada)
    if not encontradas:
        for match in _RE_DATA_HORA.finditer(texto_pagina):
            candidato = f"{match.group(1)} {match.group(2)}"
            if candidato not in encontradas:
                encontradas.append(candidato)
                if len(encontradas) >= 2:
                    break
    return encontradas


_RE_SITUACAO_CANCELADA = re.compile(
    r"cancelad|substituíd|substituid|anulad|inutiliz",
    re.IGNORECASE,
)
_RE_SITUACAO_ATIVA = re.compile(
    r"autorizad|emitid|normal|válid|valid|ativa",
    re.IGNORECASE,
)


def nota_esta_cancelada(texto_pagina: str) -> bool:
    """Lê a situação na página de visualização da NFS-e (fonte confiável)."""
    compacto = re.sub(r"\s+", " ", texto_pagina or "")
    for rotulo in (
        "Situação da NFS-e",
        "Situação da NFSe",
        "Situação",
        "Situacao",
    ):
        idx = compacto.find(rotulo)
        if idx < 0:
            continue
        trecho = compacto[idx : idx + 160]
        if _RE_SITUACAO_CANCELADA.search(trecho):
            return True
        if _RE_SITUACAO_ATIVA.search(trecho):
            return False
    return False


def pdf_indica_cancelada(caminho: Path | None) -> bool:
    if not caminho:
        return False
    return "cancelad" in caminho.name.lower()


def resolver_situacao_cancelada(
    *,
    cancelada_lista: bool,
    texto_pagina: str,
    pdf: Path | None,
) -> bool:
    if cancelada_lista:
        return True
    if nota_esta_cancelada(texto_pagina):
        return True
    return pdf_indica_cancelada(pdf)


def _ler_texto_pdf(caminho: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError(
            "Biblioteca pypdf não instalada. Rode: pip install pypdf"
        ) from exc

    reader = PdfReader(str(caminho))
    partes: list[str] = []
    for pagina in reader.pages:
        partes.append(pagina.extract_text() or "")
    return "\n".join(partes)


def _extrair_datas_pdf(texto: str) -> list[str]:
    encontradas: list[str] = []
    compacto = re.sub(r"\s+", " ", texto or "")

    for rotulo in (
        "Data e Hora da emissão da NFS-e",
        "Data e Hora da emissão da DPS",
        "Data de geração",
        "Data de emissão",
    ):
        idx = compacto.find(rotulo)
        if idx < 0:
            continue
        trecho = compacto[idx : idx + 140]
        normalizada = normalizar_data_hora(trecho)
        if normalizada and normalizada not in encontradas:
            encontradas.append(normalizada)

    for match in _RE_DATA_HORA.finditer(compacto):
        chave = f"{match.group(1)} {match.group(2)}"
        if chave not in encontradas:
            encontradas.append(chave)

    return encontradas


@dataclass
class IndicePdf:
    arquivos: list[Path] = field(default_factory=list)
    mapa_data_hora: dict[str, Path] = field(default_factory=dict)
    mapa_chave: dict[str, Path] = field(default_factory=dict)
    mapa_pdf_chave: dict[Path, str] = field(default_factory=dict)
    mapa_tomador: dict[Path, str] = field(default_factory=dict)
    sem_chave: list[Path] = field(default_factory=list)

    @classmethod
    def montar(cls, pastas: list[str]) -> IndicePdf:
        indice = cls()
        for pasta_raw in pastas:
            pasta = Path(pasta_raw)
            if not pasta.is_dir():
                continue
            for caminho in sorted(pasta.rglob("*.pdf")):
                if not caminho.is_file():
                    continue
                indice.arquivos.append(caminho)
                try:
                    texto = _ler_texto_pdf(caminho)
                except Exception:
                    continue
                chave = extrair_chave_acesso(texto)
                if chave:
                    indice.mapa_chave.setdefault(chave, caminho)
                    indice.mapa_pdf_chave.setdefault(caminho, chave)
                else:
                    indice.sem_chave.append(caminho)
                tomador = extrair_tomador_cnpj(texto)
                if tomador:
                    indice.mapa_tomador.setdefault(caminho, tomador)
                for chave_data in _extrair_datas_pdf(texto):
                    indice.mapa_data_hora.setdefault(chave_data, caminho)
        return indice

    def buscar_por_chave(self, chave: str | None) -> Path | None:
        if not chave:
            return None
        return self.mapa_chave.get(chave)

    def buscar(self, datas: list[str]) -> Path | None:
        for data in datas:
            if data in self.mapa_data_hora:
                return self.mapa_data_hora[data]
        return None


@dataclass
class ConferenciaTracker:
    """Acumula registros do portal e gera relatório cruzado site × pastas."""

    linhas: list[dict[str, Any]] = field(default_factory=list)
    chaves_vistas: set[str] = field(default_factory=set)
    chaves_confirmadas: set[str] = field(default_factory=set)

    def registrar(
        self,
        *,
        chave: str | None,
        cancelada: bool,
        resultado: str,
        pdf: Path | None = None,
        observacao: str = "",
    ) -> None:
        if chave:
            self.chaves_vistas.add(chave)
        self.linhas.append(
            {
                "chave": chave or "",
                "cancelada": cancelada,
                "resultado": resultado,
                "pdf": pdf.name if pdf else "",
                "observacao": observacao,
            }
        )

    def finalizar(
        self,
        indice: IndicePdf,
        pasta_destino: Path,
        *,
        ano: int,
        mes_inicio: int,
        mes_fim: int,
    ) -> dict[str, Any]:
        orfaos: list[dict[str, str]] = []
        sem_chave_periodo: list[str] = []
        for chave, pdf in indice.mapa_chave.items():
            if chave in self.chaves_vistas:
                continue
            if not pdf_pertence_periodo(
                pdf, indice, ano=ano, mes_inicio=mes_inicio, mes_fim=mes_fim
            ):
                continue
            orfaos.append({"chave": chave, "arquivo": pdf.name})
        for pdf in indice.sem_chave:
            if pdf_pertence_periodo(
                pdf, indice, ano=ano, mes_inicio=mes_inicio, mes_fim=mes_fim
            ):
                sem_chave_periodo.append(pdf.name)

        resumo = {
            "registros_site": len(self.linhas),
            "ok": sum(1 for l in self.linhas if l["resultado"] == "ok"),
            "cancelada": sum(1 for l in self.linhas if l["resultado"] == "cancelada"),
            "cancelada_sem_pdf": sum(
                1 for l in self.linhas if l["resultado"] == "cancelada_sem_pdf"
            ),
            "nao_encontrada": sum(
                1 for l in self.linhas if l["resultado"] == "nao_encontrada"
            ),
            "ja_conferida": sum(
                1 for l in self.linhas if l["resultado"] == "ja_conferida"
            ),
            "pdf_sem_portal": len(orfaos),
            "pdf_sem_chave": len(sem_chave_periodo),
            "pdfs_indexados": len(indice.arquivos),
            "chaves_indexadas": len(indice.mapa_chave),
            "periodo": f"{mes_inicio:02d}/{ano} a {mes_fim:02d}/{ano}",
            "total": len(self.linhas),
        }

        caminho_relatorio = salvar_relatorio_xlsx(
            pasta_destino,
            ano=ano,
            mes_inicio=mes_inicio,
            mes_fim=mes_fim,
            resumo=resumo,
            linhas=self.linhas,
            pdfs_sem_portal=orfaos,
            pdfs_sem_chave=sem_chave_periodo,
        )

        return {
            "resumo": resumo,
            "relatorio_xlsx": str(caminho_relatorio),
            "pdfs_sem_portal": orfaos,
            "pdfs_sem_chave": sem_chave_periodo,
        }


_ROTULOS_RESUMO: dict[str, str] = {
    "registros_site": "Notas lidas no portal",
    "ok": "Confirmadas (OK)",
    "cancelada": "Canceladas com PDF",
    "cancelada_sem_pdf": "Canceladas sem PDF local",
    "nao_encontrada": "Autorizadas sem PDF",
    "ja_conferida": "Já conferidas nesta execução",
    "pdf_sem_portal": "PDFs locais sem registro no portal",
    "pdf_sem_chave": "PDFs sem chave legível",
    "pdfs_indexados": "PDFs indexados nas pastas",
    "chaves_indexadas": "Chaves de acesso indexadas",
    "periodo": "Período conferido",
    "total": "Total de registros no portal",
}

_ROTULOS_RESULTADO: dict[str, str] = {
    "ok": "OK",
    "cancelada": "Cancelada (com PDF)",
    "cancelada_sem_pdf": "Cancelada sem PDF",
    "nao_encontrada": "Não encontrada",
    "ja_conferida": "Já conferida",
}

_CORES_RESULTADO: dict[str, str] = {
    "ok": "C6EFCE",
    "cancelada": "FFC7CE",
    "cancelada_sem_pdf": "FCE4D6",
    "nao_encontrada": "FFEB9C",
    "ja_conferida": "D9D9D9",
}


def _nome_arquivo_relatorio(
    *,
    ano: int,
    mes_inicio: int,
    mes_fim: int,
    extensao: str,
) -> str:
    stamp = agora_operacional_naive().strftime("%Y%m%d_%H%M")
    return (
        f"relatorio-conferencia-nfse_{ano}_{mes_inicio:02d}-{mes_fim:02d}_{stamp}.{extensao}"
    )


def _estilo_cabecalho_planilha(ws, linha: int, colunas: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, colunas + 1):
        celula = ws.cell(row=linha, column=col)
        celula.fill = fill
        celula.font = font
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _ajustar_largura_colunas(ws, larguras: dict[int, float]) -> None:
    from openpyxl.utils import get_column_letter

    for col, largura in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = largura


def salvar_relatorio_xlsx(
    pasta_destino: Path,
    *,
    ano: int,
    mes_inicio: int,
    mes_fim: int,
    resumo: dict[str, Any],
    linhas: list[dict[str, Any]],
    pdfs_sem_portal: list[dict[str, str]],
    pdfs_sem_chave: list[str],
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "Biblioteca openpyxl não instalada. Rode: pip install openpyxl"
        ) from exc

    pasta_destino.mkdir(parents=True, exist_ok=True)
    caminho = pasta_destino / _nome_arquivo_relatorio(
        ano=ano,
        mes_inicio=mes_inicio,
        mes_fim=mes_fim,
        extensao="xlsx",
    )

    wb = Workbook()
    gerado_em = agora_operacional_naive().strftime("%d/%m/%Y %H:%M:%S")
    periodo = resumo.get("periodo") or f"{mes_inicio:02d}/{ano} a {mes_fim:02d}/{ano}"

    # —— Aba Resumo ——
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo["A1"] = "Conferência NFS-e — CareCore+ Finanças"
    ws_resumo["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws_resumo["A2"] = f"Período: {periodo}"
    ws_resumo["A3"] = f"Gerado em: {gerado_em}"
    ws_resumo.append([])
    ws_resumo.append(["Indicador", "Valor"])
    _estilo_cabecalho_planilha(ws_resumo, 5, 2)
    ordem_resumo = (
        "periodo",
        "registros_site",
        "ok",
        "cancelada",
        "cancelada_sem_pdf",
        "nao_encontrada",
        "ja_conferida",
        "pdf_sem_portal",
        "pdf_sem_chave",
        "pdfs_indexados",
        "chaves_indexadas",
        "total",
    )
    for chave in ordem_resumo:
        if chave not in resumo:
            continue
        ws_resumo.append([_ROTULOS_RESUMO.get(chave, chave), resumo[chave]])
    for row in range(6, ws_resumo.max_row + 1):
        ws_resumo.cell(row=row, column=1).font = Font(bold=True)
        ws_resumo.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    _ajustar_largura_colunas(ws_resumo, {1: 42, 2: 18})

    # —— Aba Portal ——
    ws_portal = wb.create_sheet("Portal")
    cab_portal = [
        "Nº",
        "Nº NFS-e",
        "Chave de acesso",
        "Cancelada",
        "Resultado",
        "PDF / substituta sugerida",
        "Observação",
    ]
    ws_portal.append(cab_portal)
    _estilo_cabecalho_planilha(ws_portal, 1, len(cab_portal))

    linhas_ordenadas = sorted(
        linhas,
        key=lambda item: extrair_numero_nfse(item.get("chave")) or 0,
        reverse=True,
    )
    for idx, linha in enumerate(linhas_ordenadas, start=1):
        chave = linha.get("chave") or ""
        resultado = str(linha.get("resultado") or "")
        ws_portal.append(
            [
                idx,
                extrair_numero_nfse(chave) or "",
                chave,
                "Sim" if linha.get("cancelada") else "Não",
                _ROTULOS_RESULTADO.get(resultado, resultado),
                linha.get("pdf", ""),
                linha.get("observacao", ""),
            ]
        )
        cor = _CORES_RESULTADO.get(resultado)
        if cor:
            fill = PatternFill("solid", fgColor=cor)
            for col in range(1, len(cab_portal) + 1):
                ws_portal.cell(row=idx + 1, column=col).fill = fill
        ws_portal.cell(row=idx + 1, column=3).alignment = Alignment(wrap_text=False)
        ws_portal.cell(row=idx + 1, column=7).alignment = Alignment(wrap_text=True)

    ws_portal.freeze_panes = "A2"
    ws_portal.auto_filter.ref = ws_portal.dimensions
    _ajustar_largura_colunas(
        ws_portal,
        {1: 6, 2: 10, 3: 52, 4: 11, 5: 22, 6: 44, 7: 56},
    )

    # —— Aba PDF sem portal ——
    ws_orfaos = wb.create_sheet("PDF sem portal")
    ws_orfaos.append(["Chave de acesso", "Arquivo PDF"])
    _estilo_cabecalho_planilha(ws_orfaos, 1, 2)
    if pdfs_sem_portal:
        for item in pdfs_sem_portal:
            ws_orfaos.append([item.get("chave", ""), item.get("arquivo", "")])
    else:
        ws_orfaos.append(["—", "Nenhum PDF local ficou fora do portal neste período."])
    ws_orfaos.freeze_panes = "A2"
    _ajustar_largura_colunas(ws_orfaos, {1: 52, 2: 48})

    # —— Aba PDF sem chave ——
    ws_sem_chave = wb.create_sheet("PDF sem chave")
    ws_sem_chave.append(["Arquivo PDF"])
    _estilo_cabecalho_planilha(ws_sem_chave, 1, 1)
    if pdfs_sem_chave:
        for nome in pdfs_sem_chave:
            ws_sem_chave.append([nome])
    else:
        ws_sem_chave.append(["Nenhum PDF sem chave legível no período."])
    ws_sem_chave.freeze_panes = "A2"
    _ajustar_largura_colunas(ws_sem_chave, {1: 64})

    wb.save(caminho)
    return caminho


def salvar_relatorio_csv(
    pasta_destino: Path,
    *,
    ano: int,
    mes_inicio: int,
    mes_fim: int,
    resumo: dict[str, Any],
    linhas: list[dict[str, Any]],
    pdfs_sem_portal: list[dict[str, str]],
    pdfs_sem_chave: list[str],
) -> Path:
    """Compatibilidade legada — preferir salvar_relatorio_xlsx."""
    return salvar_relatorio_xlsx(
        pasta_destino,
        ano=ano,
        mes_inicio=mes_inicio,
        mes_fim=mes_fim,
        resumo=resumo,
        linhas=linhas,
        pdfs_sem_portal=pdfs_sem_portal,
        pdfs_sem_chave=pdfs_sem_chave,
    )


def _nome_destino_pdf(origem: Path, sufixo: str) -> str:
    stem = origem.stem
    ext = origem.suffix or ".pdf"
    return f"{stem} {sufixo}{ext}"


def copiar_pdf_confirmado(
    origem: Path,
    pasta_destino: Path,
    *,
    cancelada: bool,
) -> Path:
    pasta_destino.mkdir(parents=True, exist_ok=True)
    sufixo = "CANCELADA" if cancelada else "OK"
    destino = pasta_destino / _nome_destino_pdf(origem, sufixo)
    contador = 1
    while destino.exists():
        destino = pasta_destino / f"{origem.stem} {sufixo} ({contador}){origem.suffix}"
        contador += 1
    shutil.copy2(origem, destino)
    return destino


def salvar_print_nao_encontrada(
    pasta_destino: Path,
    bytes_png: bytes,
    *,
    chave: str | None,
    indice: int,
) -> Path:
    pasta_destino.mkdir(parents=True, exist_ok=True)
    if chave:
        rotulo = chave[-12:]
        nome = f"NF-nao-encontrada_chave-{rotulo}_{indice:03d}.png"
    else:
        nome = f"NF-nao-encontrada_sem-chave_{indice:03d}.png"
    destino = pasta_destino / nome
    destino.write_bytes(bytes_png)
    return destino


def salvar_print_pessoas(
    pasta_destino: Path,
    bytes_png: bytes,
    *,
    data_ref: str | None,
    indice: int,
) -> Path:
    """Compatibilidade — preferir salvar_print_nao_encontrada."""
    return salvar_print_nao_encontrada(
        pasta_destino, bytes_png, chave=data_ref, indice=indice
    )


def aplicar_resumo_final(resumo: dict[str, Any]) -> None:
    with _job_lock:
        _job["resumo"] = resumo


def validar_config(config: dict[str, Any]) -> None:
    if not config.get("pastas_origem"):
        raise ValueError("Informe ao menos uma pasta de origem com PDFs.")
    if not config.get("pasta_destino"):
        raise ValueError("Informe a pasta de destino para confirmações.")
    destino = Path(str(config["pasta_destino"]).replace("/", "\\"))
    destino.mkdir(parents=True, exist_ok=True)
    for pasta in config["pastas_origem"]:
        caminho = Path(str(pasta).replace("/", "\\"))
        if not caminho.is_dir():
            raise ValueError(f"Pasta de origem não encontrada: {pasta}")
    mi = int(config.get("mes_inicio") or 1)
    mf = int(config.get("mes_fim") or 12)
    if not (1 <= mi <= 12 and 1 <= mf <= 12):
        raise ValueError("Mês inicial e final devem estar entre 1 e 12.")
    if mi > mf:
        raise ValueError("Mês inicial não pode ser posterior ao mês final.")


def registrar_inicio(org_id: str) -> dict[str, Any]:
    config = carregar_config(org_id)
    validar_config(config)
    browser_aberto = bool(snap_job().get("browser_aberto"))
    reset_job()
    set_job(
        status="preparando",
        mensagem=f"Conferência NFS-e iniciada ({agora_operacional_naive():%d/%m/%Y %H:%M}).",
        parar=False,
        org_id=org_id,
        config=config,
        browser_aberto=browser_aberto,
    )
    return config
