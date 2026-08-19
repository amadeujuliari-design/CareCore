# -*- coding: utf-8 -*-
"""Leitura do cadastro proposto de itens de consumo (CSV/XLSX)."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from typing import Optional

from openpyxl import load_workbook

from compras_itens_consumo_utils import chave_item_consumo

_COL_DESC = frozenset({"descricao", "descrição", "descricao do item", "descrição do item", "item", "produto"})
_COL_CAT = frozenset({"categoria", "grupo"})
_COL_UN = frozenset({"unidade_medida", "unidade de medida", "unidade", "un", "und"})
_COL_MARCA = frozenset({"marca_preferencial", "marca preferencial", "marca preferencial (quando houver)", "marca"})
_COL_OBS = frozenset({"observacao", "observação", "observacao operacional", "observação operacional", "obs"})
_COL_ATIVO = frozenset({"ativo", "status"})


def _norm_header(valor: object) -> str:
    texto = str(valor or "").strip().lower()
    texto = texto.replace("\n", " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def _cel(valor: object) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).replace("\n", " ").strip()


def _ativo(valor: object) -> bool:
    texto = _cel(valor).lower()
    if texto in {"nao", "não", "n", "0", "false", "inativo"}:
        return False
    return True


def _idx(headers: list[str], aliases: frozenset[str]) -> Optional[int]:
    for i, h in enumerate(headers):
        if h in aliases:
            return i
    return None


@dataclass
class LinhaItemConsumo:
    descricao: str
    categoria: str = ""
    unidade_medida: str = ""
    marca_preferencial: str = ""
    observacao: str = ""
    ativo: bool = True

    @property
    def chave(self) -> str:
        return chave_item_consumo(self.descricao)


def _linhas_de_matriz(headers: list[str], linhas: list[list[object]]) -> list[LinhaItemConsumo]:
    i_desc = _idx(headers, _COL_DESC)
    if i_desc is None:
        return []
    i_cat = _idx(headers, _COL_CAT)
    i_un = _idx(headers, _COL_UN)
    i_marca = _idx(headers, _COL_MARCA)
    i_obs = _idx(headers, _COL_OBS)
    i_ativo = _idx(headers, _COL_ATIVO)
    saida: list[LinhaItemConsumo] = []
    for row in linhas:
        if not row:
            continue
        def _get(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row):
                return ""
            return _cel(row[idx])

        descricao = _get(i_desc)
        if not descricao:
            continue
        saida.append(
            LinhaItemConsumo(
                descricao=descricao,
                categoria=_get(i_cat),
                unidade_medida=_get(i_un),
                marca_preferencial=_get(i_marca),
                observacao=_get(i_obs),
                ativo=_ativo(_get(i_ativo)) if i_ativo is not None else True,
            )
        )
    return saida


def extrair_itens_consumo(conteudo: bytes, nome_arquivo: str = "") -> list[LinhaItemConsumo]:
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".csv") or (not nome.endswith(".xlsx") and b"," in conteudo[:200]):
        texto = conteudo.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(texto))
        rows = list(reader)
        if not rows:
            return []
        headers = [_norm_header(h) for h in rows[0]]
        return _linhas_de_matriz(headers, rows[1:])

    bio = io.BytesIO(conteudo)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if not first:
            return []
        headers = [_norm_header(h) for h in first]
        linhas = [list(r or []) for r in rows_iter]
        return _linhas_de_matriz(headers, linhas)
    finally:
        wb.close()
