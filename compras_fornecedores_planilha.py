# -*- coding: utf-8 -*-
"""Leitura das planilhas de fornecedores da Sede AEB (CSV operacional e XLSX 2025)."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from compras_telefone_utils import formatar_telefone_compras, sanitizar_telefone_compras

from openpyxl import load_workbook

_COL_NOME = frozenset({"fornecedor", "nome", "razao social", "razão social"})
_COL_SEGMENTO = frozenset({"segmento", "tipo de servico", "tipo de serviço", "categoria"})
_COL_CONTATO = frozenset({"contato", "representante", "repres"})
_COL_EMAIL = frozenset({"e-mail", "email", "e mail"})
_COL_EMAIL_EMPRESA = frozenset({"email da empresa", "email empresa", "e-mail da empresa"})
_COL_TELEFONE = frozenset({"telefone", "tel", "celular"})
_COL_CIDADE = frozenset({"cidade", "municipio", "município"})
_COL_CNPJ = frozenset({"cnpj"})
_COL_PROJETOS = frozenset({"projetos", "projeto", "unidades"})
_COL_STATUS = frozenset({"status", "situacao", "situação"})
_COL_OBS = frozenset({"observacao", "observação", "obs", "avaliacao", "avaliação"})


def _norm_header(valor: object) -> str:
    texto = str(valor or "").strip().lower()
    texto = texto.replace("\n", " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def _norm_nome(valor: object) -> str:
    return re.sub(r"\s+", " ", str(valor or "").strip())


def _norm_chave_nome(valor: object) -> str:
    return _norm_nome(valor).upper()


def _cel(valor: object) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).strip()
    return str(valor).strip()


def _status_ativo(valor: object) -> Optional[bool]:
    texto = _cel(valor).lower()
    if not texto:
        return None
    if texto in {"ativo", "sim", "s", "1", "true"}:
        return True
    if texto in {"inativo", "nao", "não", "n", "0", "false", "bloqueado"}:
        return False
    return None


@dataclass
class LinhaFornecedorPlanilha:
    nome: str
    segmento: str = ""
    categoria: str = ""
    contato: str = ""
    telefone: str = ""
    email: str = ""
    email_empresa: str = ""
    cidade: str = ""
    cnpj: str = ""
    projetos_atendidos: str = ""
    observacao: str = ""
    ativo: Optional[bool] = None
    origem: str = ""
    aba: str = ""


def _mapear_colunas(headers: list[str]) -> dict[str, int]:
    mapa: dict[str, int] = {}
    for idx, bruto in enumerate(headers):
        h = _norm_header(bruto)
        if not h:
            continue
        if h in _COL_NOME:
            mapa.setdefault("nome", idx)
        elif h in _COL_SEGMENTO:
            mapa.setdefault("segmento", idx)
        elif h in _COL_CONTATO:
            mapa.setdefault("contato", idx)
        elif h in _COL_EMAIL_EMPRESA:
            mapa.setdefault("email_empresa", idx)
        elif h in _COL_EMAIL:
            mapa.setdefault("email", idx)
        elif h in _COL_TELEFONE:
            mapa.setdefault("telefone", idx)
        elif h in _COL_CIDADE:
            mapa.setdefault("cidade", idx)
        elif h in _COL_CNPJ:
            mapa.setdefault("cnpj", idx)
        elif h in _COL_PROJETOS:
            mapa.setdefault("projetos", idx)
        elif h in _COL_STATUS:
            mapa.setdefault("status", idx)
        elif h in _COL_OBS:
            mapa.setdefault("observacao", idx)
    return mapa


def _valor_coluna(row: list[str], mapa: dict[str, int], chave: str) -> str:
    idx = mapa.get(chave)
    if idx is None or idx >= len(row):
        return ""
    return _cel(row[idx])


def _linha_de_registro(row: list[str], mapa: dict[str, int], origem: str, aba: str = "") -> Optional[LinhaFornecedorPlanilha]:
    nome = _valor_coluna(row, mapa, "nome")
    if not nome:
        return None
    segmento = _valor_coluna(row, mapa, "segmento")
    categoria = segmento
    if not segmento and _valor_coluna(row, mapa, "categoria"):
        categoria = _valor_coluna(row, mapa, "categoria")
    return LinhaFornecedorPlanilha(
        nome=nome,
        segmento=segmento or categoria,
        categoria=categoria,
        contato=_valor_coluna(row, mapa, "contato"),
        telefone=_valor_coluna(row, mapa, "telefone"),
        email=_valor_coluna(row, mapa, "email"),
        email_empresa=_valor_coluna(row, mapa, "email_empresa"),
        cidade=_valor_coluna(row, mapa, "cidade"),
        cnpj=_valor_coluna(row, mapa, "cnpj"),
        projetos_atendidos=_valor_coluna(row, mapa, "projetos"),
        observacao=_valor_coluna(row, mapa, "observacao"),
        ativo=_status_ativo(_valor_coluna(row, mapa, "status")),
        origem=origem,
        aba=aba,
    )


def _parse_csv(conteudo: bytes) -> list[LinhaFornecedorPlanilha]:
    texto = conteudo.decode("latin-1", errors="replace")
    linhas_brutas = list(csv.reader(io.StringIO(texto), delimiter=";"))
    header_idx = None
    for idx, row in enumerate(linhas_brutas):
        joined = _norm_header(" ".join(_cel(c) for c in row))
        if "fornecedor" in joined and ("contato" in joined or "telefone" in joined):
            header_idx = idx
            break
    if header_idx is None:
        raise ValueError("Cabeçalho de fornecedores não encontrado no CSV (esperado: Fornecedor;Contato;Telefone…).")

    headers = [_cel(c) for c in linhas_brutas[header_idx]]
    mapa = _mapear_colunas(headers)
    if "nome" not in mapa:
        raise ValueError("Coluna Fornecedor não encontrada no CSV.")

    saida: list[LinhaFornecedorPlanilha] = []
    for row in linhas_brutas[header_idx + 1 :]:
        cells = [_cel(c) for c in row]
        if not any(cells):
            continue
        item = _linha_de_registro(cells, mapa, origem="csv")
        if item:
            saida.append(item)
    return saida


def _parse_xlsx(conteudo: bytes) -> list[LinhaFornecedorPlanilha]:
    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    saida: list[LinhaFornecedorPlanilha] = []
    for aba in wb.sheetnames:
        ws = wb[aba]
        header_idx = None
        headers: list[str] = []
        preview: list[list[str]] = []
        for idx, row in enumerate(ws.iter_rows(max_row=12, values_only=True), 1):
            cells = [_cel(c) for c in row]
            if not any(cells):
                continue
            preview.append(cells)
            joined = _norm_header(" ".join(cells))
            if "fornecedor" in joined and ("segmento" in joined or "representante" in joined or "contato" in joined):
                header_idx = idx
                headers = cells
                break
        if header_idx is None:
            continue
        mapa = _mapear_colunas(headers)
        if "nome" not in mapa:
            continue
        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            cells = [_cel(c) for c in row]
            if not any(cells):
                continue
            item = _linha_de_registro(cells, mapa, origem="xlsx", aba=aba)
            if item:
                saida.append(item)
    if not saida:
        raise ValueError("Nenhuma linha de fornecedor encontrada no XLSX.")
    return saida


def extrair_linhas_fornecedores(conteudo: bytes, nome_arquivo: str) -> list[LinhaFornecedorPlanilha]:
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".csv"):
        return _parse_csv(conteudo)
    if nome.endswith((".xlsx", ".xlsm", ".xltx")):
        return _parse_xlsx(conteudo)
    if conteudo[:2] == b"PK":
        return _parse_xlsx(conteudo)
    return _parse_csv(conteudo)


def _concat_unico(atual: str, novo: str, separador: str = "; ") -> str:
    atual_limpo = _norm_nome(atual)
    novo_limpo = _norm_nome(novo)
    if not novo_limpo:
        return atual_limpo
    if not atual_limpo:
        return novo_limpo
    partes = [p.strip() for p in re.split(r"[;,/|]", atual_limpo) if p.strip()]
    chaves = {p.upper() for p in partes}
    if novo_limpo.upper() in chaves:
        return atual_limpo
    partes.append(novo_limpo)
    return separador.join(partes)


def mesclar_payload_fornecedor(existente: dict, linha: LinhaFornecedorPlanilha) -> dict:
    """Mescla linha da planilha em payload de fornecedor (upsert por nome)."""
    payload = dict(existente)
    payload["nome"] = _norm_nome(linha.nome)
    payload["segmento"] = _concat_unico(payload.get("segmento") or "", linha.segmento)
    payload["contato"] = payload.get("contato") or linha.contato or None
    tel_principal, tel_extras = sanitizar_telefone_compras(linha.telefone or payload.get("telefone"))
    payload["telefone"] = tel_principal
    if tel_extras:
        extras_txt = " / ".join(formatar_telefone_compras(t) for t in tel_extras)
        obs = payload.get("observacao") or ""
        if extras_txt not in obs:
            payload["observacao"] = f"{obs} | Tel. adicional: {extras_txt}".strip(" |") if obs else f"Tel. adicional: {extras_txt}"
    payload["email"] = payload.get("email") or linha.email or None
    payload["email_empresa"] = payload.get("email_empresa") or linha.email_empresa or None
    payload["cidade"] = payload.get("cidade") or linha.cidade or None
    payload["cnpj"] = payload.get("cnpj") or linha.cnpj or None
    payload["projetos_atendidos"] = _concat_unico(payload.get("projetos_atendidos") or "", linha.projetos_atendidos)
    if linha.observacao:
        payload["observacao"] = _concat_unico(payload.get("observacao") or "", linha.observacao, " | ")
    if linha.ativo is not None and payload.get("ativo") is None:
        payload["ativo"] = linha.ativo
    return payload


def linha_para_payload(linha: LinhaFornecedorPlanilha) -> dict:
    tel_principal, tel_extras = sanitizar_telefone_compras(linha.telefone)
    observacao = linha.observacao or None
    if tel_extras:
        extras_txt = " / ".join(formatar_telefone_compras(t) for t in tel_extras)
        observacao = f"{observacao} | Tel. adicional: {extras_txt}".strip(" |") if observacao else f"Tel. adicional: {extras_txt}"
    return {
        "nome": _norm_nome(linha.nome),
        "segmento": _norm_nome(linha.segmento) or None,
        "contato": linha.contato or None,
        "telefone": tel_principal,
        "email": linha.email or None,
        "email_empresa": linha.email_empresa or None,
        "cidade": linha.cidade or None,
        "cnpj": linha.cnpj or None,
        "projetos_atendidos": _norm_nome(linha.projetos_atendidos) or None,
        "observacao": observacao,
        "ativo": True if linha.ativo is None else bool(linha.ativo),
        "bloqueado": False,
    }
