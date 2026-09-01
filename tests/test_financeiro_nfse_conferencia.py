"""Testes do serviço de conferência NFS-e."""

from pathlib import Path

from financeiro_nfse_conferencia_service import (
    ConferenciaTracker,
    IndicePdf,
    extrair_chave_acesso,
    extrair_datas_portal,
    extrair_numero_nfse,
    montar_observacao_cancelada_sem_pdf,
    normalizar_data_hora,
    nota_esta_cancelada,
    resolver_situacao_cancelada,
    resolver_periodo_meses,
    salvar_relatorio_xlsx,
    sugerir_pdf_substituto,
)


def test_normalizar_data_hora_portal():
    assert normalizar_data_hora("29/01/2026 às 19:51:26-03:00") == "29/01/2026 19:51:26"
    assert normalizar_data_hora("06/01/2026 19:47:41") == "06/01/2026 19:47:41"


def test_extrair_datas_portal():
    texto = """
    Identificação da NFS-e
    Data de geração
    29/01/2026 às 19:51:26-03:00
    Identificação do DPS
    Data de emissão
    29/01/2026 às 19:51:26-03:00
    """
    datas = extrair_datas_portal(texto)
    assert "29/01/2026 19:51:26" in datas


def test_nota_esta_cancelada_apenas_situacao():
    autorizada = """
    Situação da NFS-e
    Autorizada
    Data de geração
    29/01/2026 às 19:51:26-03:00
    """
    assert nota_esta_cancelada(autorizada) is False

    cancelada = """
    Situação da NFS-e
    Cancelada
    Data de geração
    28/01/2026 às 10:00:00-03:00
    """
    assert nota_esta_cancelada(cancelada) is True

    menu_falso_positivo = """
    Ações
    Cancelar NFS-e
    Substituir NFS-e
    Situação da NFS-e
    Autorizada
    """
    assert nota_esta_cancelada(menu_falso_positivo) is False


def test_extrair_chave_acesso():
    chave = "35503082254166611000133000000000055026085143597594"
    texto = f"Identificação da NFS-e\nChave de acesso\n{chave}\n"
    assert extrair_chave_acesso(texto) == chave


def test_resolver_situacao_cancelada_lista():
    assert (
        resolver_situacao_cancelada(
            cancelada_lista=True,
            texto_pagina="Situação da NFS-e Autorizada",
            pdf=None,
        )
        is True
    )


def test_resolver_periodo_meses():
    ano, mi, mf = resolver_periodo_meses(
        {"ano": 2026, "mes_inicio": 3, "mes_fim": 3}
    )
    assert ano == 2026
    assert mi == 3
    assert mf == 3


def test_extrair_numero_nfse():
    chave = "35503082254166611000133000000000040126039937713862"
    assert extrair_numero_nfse(chave) == 401


def test_sugerir_pdf_substituto_por_horario():
    indice = IndicePdf()
    pdf_substituta = Path("NF VILA PARI 03-2026.pdf")
    chave_substituta = "35503082254166611000133000000000040226030265850400"
    chave_cancelada = "35503082254166611000133000000000040126039937713862"
    indice.mapa_data_hora["02/03/2026 08:36:12"] = pdf_substituta
    indice.mapa_pdf_chave[pdf_substituta] = chave_substituta
    indice.mapa_tomador[pdf_substituta] = "61.705.877/0001-72"

    texto_portal = """
    Situação da NFS-e
    Cancelada
    Data de geração
    02/03/2026 às 08:33:10-03:00
    Tomador do Serviço
    CNPJ / CPF / NIF 61.705.877/0001-72
    """
    sugestao = sugerir_pdf_substituto(
        indice,
        texto_portal=texto_portal,
        chave_atual=chave_cancelada,
    )
    assert sugestao == pdf_substituta


def test_montar_observacao_cancelada_sem_pdf():
    indice = IndicePdf()
    pdf = Path("NF VILA PARI 03-2026.pdf")
    indice.mapa_pdf_chave[pdf] = (
        "35503082254166611000133000000000040226030265850400"
    )
    obs = montar_observacao_cancelada_sem_pdf(pdf, indice=indice)
    assert "Possível substituta" in obs
    assert "nota 402" in obs


def test_tracker_conta_cancelada_sem_pdf(tmp_path):
    tracker = ConferenciaTracker()
    tracker.registrar(
        chave="35503082254166611000133000000000040126039937713862",
        cancelada=True,
        resultado="cancelada_sem_pdf",
    )
    tracker.registrar(
        chave="35503082254166611000133000000000039726033662186552",
        cancelada=False,
        resultado="nao_encontrada",
    )
    resultado = tracker.finalizar(
        IndicePdf(),
        tmp_path,
        ano=2026,
        mes_inicio=3,
        mes_fim=3,
    )
    resumo = resultado["resumo"]
    assert resumo["cancelada_sem_pdf"] == 1
    assert resumo["nao_encontrada"] == 1
    assert resultado["relatorio_xlsx"].endswith(".xlsx")
    assert Path(resultado["relatorio_xlsx"]).is_file()


def test_salvar_relatorio_xlsx_abas(tmp_path):
    from openpyxl import load_workbook

    caminho = salvar_relatorio_xlsx(
        tmp_path,
        ano=2026,
        mes_inicio=3,
        mes_fim=3,
        resumo={
            "periodo": "03/2026 a 03/2026",
            "registros_site": 1,
            "ok": 1,
            "cancelada": 0,
            "cancelada_sem_pdf": 0,
            "nao_encontrada": 0,
            "ja_conferida": 0,
            "pdf_sem_portal": 0,
            "pdf_sem_chave": 0,
            "pdfs_indexados": 10,
            "chaves_indexadas": 10,
            "total": 1,
        },
        linhas=[
            {
                "chave": "35503082254166611000133000000000039726033662186552",
                "cancelada": False,
                "resultado": "ok",
                "pdf": "NF Atende 3 03-2026.pdf",
                "observacao": "",
            }
        ],
        pdfs_sem_portal=[],
        pdfs_sem_chave=[],
    )
    wb = load_workbook(caminho)
    assert wb.sheetnames == ["Resumo", "Portal", "PDF sem portal", "PDF sem chave"]
    assert wb["Portal"]["E2"].value == "OK"
