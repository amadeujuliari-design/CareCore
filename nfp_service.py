"""Servicos do modulo NFP – Creditos."""
from __future__ import annotations

import csv
import io
import sys
from collections import defaultdict
from typing import Any, BinaryIO, Iterable, Optional, Sequence, Union

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from models import (
    NfpAgenteCaptadorDB,
    NfpBatimentoDB,
    NfpCnpjCaptacaoCompetenciaDB,
    NfpCnpjLojaDB,
    NfpDoacaoAutomaticaDB,
    NfpDoadorDB,
    NfpRateioDB,
    NfpSefazCreditoDB,
)
from nfp_utils import (
    AGENTES_CAPTACAO_PADRAO,
    CAPTADORES_PADRAO,
    NOME_GENERICO_CONFERIR,
    achar_coluna,
    centavos_para_float,
    chave_base,
    chave_com_ocorrencia,
    cnpj_valido,
    competencia_referencia_das_datas,
    competencia_valida,
    cpf_valido,
    limpar_documento,
    limpar_nota,
    nome_eh_generico,
    nome_loja_para_cadastro,
    normalizar_agente_captacao,
    origem_doador_auto_agente,
    origem_eh_rateio_agente,
    origem_rateio_agente,
    percentual_agente_padrao,
    rateio_centavos,
    tipo_eh_doacao_automatica,
    valor_para_centavos,
)
from time_operacional import agora_operacional_naive


# Exportacoes do site podem trazer campos longos; o padrao do csv (128 KB) estoura.
csv.field_size_limit(min(sys.maxsize, 2**31 - 1))


CAMPOS_ENDERECO = (
    "cep",
    "logradouro",
    "numero",
    "complemento",
    "bairro",
    "cidade",
    "uf",
)


def _texto_opcional(valor) -> Optional[str]:
    texto = str(valor or "").strip()
    return texto or None


def _bool_payload(valor, default: bool = True) -> bool:
    if valor is None:
        return default
    if isinstance(valor, bool):
        return valor
    texto = str(valor).strip().lower()
    if texto in {"1", "true", "sim", "s", "yes"}:
        return True
    if texto in {"0", "false", "nao", "não", "n", "no"}:
        return False
    return default


def _int_percentual(valor, default: int = 0) -> int:
    try:
        pct = int(valor)
    except (TypeError, ValueError):
        pct = default
    return max(0, min(100, pct))


def serializar_agente(row: NfpAgenteCaptadorDB) -> dict:
    return {
        "id": row.id,
        "numero_cadastro": int(getattr(row, "numero_cadastro", 0) or 0),
        "codigo": row.codigo,
        "tipo": row.tipo,
        "nome": row.nome,
        "nome_fantasia": row.nome_fantasia,
        "cpf": row.cpf,
        "cnpj": row.cnpj,
        "email": row.email,
        "telefone": row.telefone,
        "cep": row.cep,
        "logradouro": row.logradouro,
        "numero": row.numero,
        "complemento": row.complemento,
        "bairro": row.bairro,
        "cidade": row.cidade,
        "uf": row.uf,
        "percentual_agente": int(row.percentual_agente or 0),
        "percentual_aeb": 100 - int(row.percentual_agente or 0),
        "ativo": bool(row.ativo),
        "observacoes": row.observacoes,
        "criado_em": row.criado_em.isoformat() if row.criado_em else None,
        "atualizado_em": row.atualizado_em.isoformat() if row.atualizado_em else None,
    }


def serializar_doador(row: NfpDoadorDB) -> dict:
    return {
        "id": row.id,
        "numero_cadastro": int(getattr(row, "numero_cadastro", 0) or 0),
        "nome": row.nome,
        "email": row.email,
        "telefone": row.telefone,
        "cpf": row.cpf,
        "data_nascimento": row.data_nascimento,
        "unidade_captador": row.unidade_captador,
        "origem_cadastro": getattr(row, "origem_cadastro", None),
        "cep": row.cep,
        "logradouro": row.logradouro,
        "numero": row.numero,
        "complemento": row.complemento,
        "bairro": row.bairro,
        "cidade": row.cidade,
        "uf": row.uf,
        "ativo": bool(getattr(row, "ativo", True)),
        "observacoes": getattr(row, "observacoes", None),
        "criado_em": row.criado_em.isoformat() if row.criado_em else None,
        "atualizado_em": row.atualizado_em.isoformat() if getattr(row, "atualizado_em", None) else None,
    }


def serializar_cnpj(row: NfpCnpjLojaDB) -> dict:
    return {
        "id": row.id,
        "numero_cadastro": int(getattr(row, "numero_cadastro", 0) or 0),
        "cnpj": row.cnpj,
        "loja": row.loja,
        "razao_social": getattr(row, "razao_social", None),
        "inscricao_estadual": getattr(row, "inscricao_estadual", None),
        "captador": row.captador,
        "email": getattr(row, "email", None),
        "telefone": getattr(row, "telefone", None),
        "cep": getattr(row, "cep", None),
        "logradouro": getattr(row, "logradouro", None),
        "numero": getattr(row, "numero", None),
        "complemento": getattr(row, "complemento", None),
        "bairro": getattr(row, "bairro", None),
        "cidade": getattr(row, "cidade", None),
        "uf": getattr(row, "uf", None),
        "cnpj_conferir": bool(row.cnpj_conferir),
        "ativo": bool(getattr(row, "ativo", True)),
        "observacoes": getattr(row, "observacoes", None),
        "criado_em": row.criado_em.isoformat() if row.criado_em else None,
        "atualizado_em": row.atualizado_em.isoformat() if getattr(row, "atualizado_em", None) else None,
    }


def aplicar_endereco(obj, payload: dict) -> None:
    for campo in CAMPOS_ENDERECO:
        if campo not in payload:
            continue
        valor = _texto_opcional(payload.get(campo))
        if campo == "cep" and valor:
            valor = limpar_documento(valor)[:8]
        if campo == "uf" and valor:
            valor = valor.upper()[:2]
        setattr(obj, campo, valor)


async def mapa_percentual_agentes(db: AsyncSession, organizacao_id: str) -> dict[str, int]:
    rows = (
        await db.execute(
            select(NfpAgenteCaptadorDB).where(NfpAgenteCaptadorDB.organizacao_id == organizacao_id)
        )
    ).scalars().all()
    mapa = {normalizar_agente_captacao(r.codigo): int(r.percentual_agente or 0) for r in rows}
    for codigo in CAPTADORES_PADRAO:
        chave = normalizar_agente_captacao(codigo)
        if chave not in mapa:
            mapa[chave] = percentual_agente_padrao(chave)
    return mapa


async def proximo_numero_cadastro_agente(db: AsyncSession, organizacao_id: str) -> int:
    atual = (
        await db.execute(
            select(func.coalesce(func.max(NfpAgenteCaptadorDB.numero_cadastro), 0)).where(
                NfpAgenteCaptadorDB.organizacao_id == organizacao_id
            )
        )
    ).scalar_one()
    return int(atual or 0) + 1


async def garantir_agentes_padrao(db: AsyncSession, organizacao_id: str) -> dict:
    existentes = (
        await db.execute(
            select(NfpAgenteCaptadorDB).where(NfpAgenteCaptadorDB.organizacao_id == organizacao_id)
        )
    ).scalars().all()
    por_codigo = {normalizar_agente_captacao(r.codigo): r for r in existentes}
    proximo = await proximo_numero_cadastro_agente(db, organizacao_id)
    criados = 0
    agora = agora_operacional_naive()
    for codigo in CAPTADORES_PADRAO:
        chave = normalizar_agente_captacao(codigo)
        if chave in por_codigo:
            continue
        db.add(
            NfpAgenteCaptadorDB(
                organizacao_id=organizacao_id,
                numero_cadastro=proximo,
                codigo=chave,
                tipo="PJ",
                nome=codigo.strip(),
                percentual_agente=percentual_agente_padrao(chave),
                ativo=True,
                criado_em=agora,
                atualizado_em=agora,
            )
        )
        proximo += 1
        criados += 1
    if criados:
        await db.commit()
    return {"criados": criados, "total_padrao": len(CAPTADORES_PADRAO)}


def _cel(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _bytes_arquivo(arquivo: BinaryIO | bytes) -> bytes:
    if hasattr(arquivo, "read"):
        raw = arquivo.read()
        if hasattr(arquivo, "seek"):
            try:
                arquivo.seek(0)
            except Exception:
                pass
        return raw if isinstance(raw, (bytes, bytearray)) else bytes(raw or b"")
    return arquivo if isinstance(arquivo, (bytes, bytearray)) else bytes(arquivo or b"")


def _decodificar_csv(raw: bytes) -> str:
    """Decodifica CSV do site SEFAZ (UTF-16 com BOM, UTF-8 ou Latin-1/cp1252)."""
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        tentativas = ("utf-16",)
    elif raw[:4000].count(b"\x00") > len(raw[:4000]) // 4:
        # Muitos bytes nulos = UTF-16 sem BOM.
        tentativas = ("utf-16-le", "utf-16-be")
    else:
        tentativas = ("utf-8-sig", "utf-8", "cp1252", "latin-1")

    for enc in tentativas:
        try:
            return raw.decode(enc)
        except Exception:
            continue
    return raw.decode("latin-1", errors="replace")


def _ler_planilha_csv(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = _decodificar_csv(raw)
    if not text:
        raise ValueError("Nao foi possivel ler o CSV (encoding).")

    # Remove BOM residual
    if text.startswith("\ufeff"):
        text = text.lstrip("\ufeff")

    first = next((ln for ln in text.splitlines() if ln.strip()), "")
    if not first:
        return [], []
    if first.count("\t") >= first.count(";") and first.count("\t") >= first.count(","):
        delim = "\t"
    elif first.count(";") >= first.count(","):
        delim = ";"
    else:
        delim = ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    headers = [(_cel(h) or f"col_{i}") for i, h in enumerate(reader.fieldnames or [])]
    # DictReader keys are original fieldnames
    dados: list[dict[str, Any]] = []
    for row in reader:
        if not row or not any(_cel(v) for v in row.values()):
            continue
        item = {}
        for h in reader.fieldnames or []:
            chave = _cel(h) or h
            item[chave] = row.get(h)
        dados.append(item)
    # normalize headers list to match item keys
    if dados:
        headers = list(dados[0].keys())
    return headers, dados


def _ler_planilha(arquivo: BinaryIO | bytes) -> tuple[list[str], list[dict[str, Any]]]:
    raw = _bytes_arquivo(arquivo)
    if not raw:
        return [], []
    # XLSX/OLE signature
    if raw[:2] == b"PK" or raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            wb.close()
            return [], []
        headers = [_cel(h) or f"col_{i}" for i, h in enumerate(header_row)]
        dados = []
        for row in rows_iter:
            if not row or not any(_cel(c) for c in row):
                continue
            item = {}
            for i, h in enumerate(headers):
                item[h] = row[i] if i < len(row) else None
            dados.append(item)
        wb.close()
        return headers, dados
    return _ler_planilha_csv(raw)


def _ler_varios_arquivos(arquivos: Sequence[BinaryIO | bytes]) -> tuple[list[str], list[dict[str, Any]]]:
    headers_ref: list[str] = []
    dados: list[dict[str, Any]] = []
    for arq in arquivos:
        headers, parte = _ler_planilha(arq)
        if not parte:
            continue
        if not headers_ref:
            headers_ref = headers
        dados.extend(parte)
    return headers_ref, dados


def _adicionar_ocorrencias(linhas: list[dict], grupo_keys: list[str]) -> list[dict]:
    contadores: dict[tuple, int] = defaultdict(int)
    out = []
    for linha in linhas:
        chave = tuple(linha.get(k) for k in grupo_keys)
        contadores[chave] += 1
        nova = dict(linha)
        nova["ocorrencia"] = contadores[chave]
        out.append(nova)
    return out


async def enriquecer_nomes_cnpjs_genericos(
    db: AsyncSession,
    organizacao_id: str,
    pares_cnpj_nome: Iterable[tuple[str, str]],
) -> int:
    """Substitui nome generico no cadastro quando a SEFAZ traz nome real."""
    atualizados = 0
    mapa: dict[str, str] = {}
    for cnpj, nome in pares_cnpj_nome:
        c = limpar_documento(cnpj)
        n = (nome or "").strip()
        if c and n and not nome_eh_generico(n):
            mapa[c] = n
    if not mapa:
        return 0

    resultado = await db.execute(
        select(NfpCnpjLojaDB).where(NfpCnpjLojaDB.organizacao_id == organizacao_id)
    )
    for loja in resultado.scalars().all():
        if not nome_eh_generico(loja.loja):
            continue
        novo = mapa.get(limpar_documento(loja.cnpj))
        if not novo:
            continue
        loja.loja = novo
        if cnpj_valido(loja.cnpj):
            loja.cnpj_conferir = False
        atualizados += 1
    return atualizados


ORIGEM_DOADOR_MANUAL = "MANUAL"
ORIGEM_DOADOR_PLANILHA = "PLANILHA"
ORIGEM_DOADOR_DOACAO = "DOACAO_AUTOMATICA"
UNIDADE_DOADOR_DIRETO_AEB = "AEB"


async def proximo_numero_cadastro_doador(db: AsyncSession, organizacao_id: str) -> int:
    atual = (
        await db.execute(
            select(func.coalesce(func.max(NfpDoadorDB.numero_cadastro), 0)).where(
                NfpDoadorDB.organizacao_id == organizacao_id
            )
        )
    ).scalar_one()
    return int(atual or 0) + 1


async def proximo_numero_cadastro_cnpj(db: AsyncSession, organizacao_id: str) -> int:
    atual = (
        await db.execute(
            select(func.coalesce(func.max(NfpCnpjLojaDB.numero_cadastro), 0)).where(
                NfpCnpjLojaDB.organizacao_id == organizacao_id
            )
        )
    ).scalar_one()
    return int(atual or 0) + 1


def nome_placeholder_doador(cpf: str) -> str:
    digitos = limpar_documento(cpf)
    if len(digitos) == 11:
        return f"Doador {digitos[:3]}.***.***-{digitos[-2:]}"
    return f"Doador CPF {digitos or 'pendente'}"


async def garantir_doador_por_cpf(
    db: AsyncSession,
    organizacao_id: str,
    cpf: str,
    *,
    nome: Optional[str] = None,
    unidade_captador: Optional[str] = UNIDADE_DOADOR_DIRETO_AEB,
    origem_cadastro: str = ORIGEM_DOADOR_DOACAO,
    existentes: Optional[dict[str, NfpDoadorDB]] = None,
    contador_numero: Optional[list[int]] = None,
) -> tuple[Optional[NfpDoadorDB], bool]:
    """Garante ficha de doador. Retorna (row, criado)."""
    cpf_limpo = limpar_documento(cpf)
    if not cpf_limpo or not cpf_valido(cpf_limpo):
        return None, False

    row = None
    if existentes is not None:
        row = existentes.get(cpf_limpo)
    else:
        row = (
            await db.execute(
                select(NfpDoadorDB).where(
                    NfpDoadorDB.organizacao_id == organizacao_id,
                    NfpDoadorDB.cpf == cpf_limpo,
                )
            )
        ).scalar_one_or_none()

    if row:
        return row, False

    if contador_numero is not None:
        numero = int(contador_numero[0])
        contador_numero[0] = numero + 1
    else:
        numero = await proximo_numero_cadastro_doador(db, organizacao_id)

    agora = agora_operacional_naive()
    nome_final = (nome or "").strip() or nome_placeholder_doador(cpf_limpo)
    row = NfpDoadorDB(
        organizacao_id=organizacao_id,
        numero_cadastro=numero,
        nome=nome_final,
        cpf=cpf_limpo,
        unidade_captador=normalizar_agente_captacao(unidade_captador) if unidade_captador else UNIDADE_DOADOR_DIRETO_AEB,
        origem_cadastro=origem_cadastro,
        ativo=True,
        criado_em=agora,
        atualizado_em=agora,
    )
    db.add(row)
    if existentes is not None:
        existentes[cpf_limpo] = row
    return row, True


async def sincronizar_doadores_de_doacoes(
    db: AsyncSession,
    organizacao_id: str,
    competencia: Optional[str] = None,
) -> dict:
    """Inclui no cadastro CPFs de doacao automatica (doadores diretos AEB)."""
    q = (
        select(NfpDoacaoAutomaticaDB.cpf_doador_cadastrador)
        .where(NfpDoacaoAutomaticaDB.organizacao_id == organizacao_id)
        .distinct()
    )
    if competencia and competencia_valida(competencia):
        q = q.where(NfpDoacaoAutomaticaDB.competencia == competencia)

    cpfs_raw = (await db.execute(q)).scalars().all()
    cpfs = sorted({limpar_documento(c) for c in cpfs_raw if limpar_documento(c)})

    existentes_rows = (
        await db.execute(
            select(NfpDoadorDB).where(NfpDoadorDB.organizacao_id == organizacao_id)
        )
    ).scalars().all()
    existentes = {limpar_documento(r.cpf): r for r in existentes_rows if limpar_documento(r.cpf)}
    contador_numero = [await proximo_numero_cadastro_doador(db, organizacao_id)]

    criados = 0
    ja_existiam = 0
    invalidos = 0
    for cpf in cpfs:
        if not cpf_valido(cpf):
            invalidos += 1
            continue
        _, criado = await garantir_doador_por_cpf(
            db,
            organizacao_id,
            cpf,
            unidade_captador=UNIDADE_DOADOR_DIRETO_AEB,
            origem_cadastro=ORIGEM_DOADOR_DOACAO,
            existentes=existentes,
            contador_numero=contador_numero,
        )
        if criado:
            criados += 1
        else:
            ja_existiam += 1

    if criados:
        await db.commit()

    return {
        "cpfs_analisados": len(cpfs),
        "criados": criados,
        "ja_existiam": ja_existiam,
        "invalidos": invalidos,
        "total_cadastro": len(existentes),
    }


async def importar_doadores(
    db: AsyncSession,
    organizacao_id: str,
    arquivo: BinaryIO,
) -> dict:
    headers, dados = _ler_planilha(arquivo)
    col_nome = achar_coluna(headers, ["nome"])
    col_cpf = achar_coluna(headers, ["cpf"])
    col_email = achar_coluna(headers, ["email", "e-mail"])
    col_telefone = achar_coluna(headers, ["telefone", "fone", "celular"])
    col_unidade = achar_coluna(headers, ["unidade_captador", "unidade / captador", "unidade", "captador"])
    if not col_nome or not col_cpf:
        raise ValueError("Nao encontrei as colunas obrigatorias: nome e cpf.")

    inseridos = 0
    ignorados = 0
    proximo = await proximo_numero_cadastro_doador(db, organizacao_id)
    for row in dados:
        nome = _cel(row.get(col_nome))
        cpf = limpar_documento(row.get(col_cpf))
        if not nome or not cpf:
            continue
        existe = await db.execute(
            select(NfpDoadorDB.id).where(
                NfpDoadorDB.organizacao_id == organizacao_id,
                NfpDoadorDB.cpf == cpf,
            )
        )
        if existe.scalar_one_or_none():
            ignorados += 1
            continue
        db.add(
            NfpDoadorDB(
                organizacao_id=organizacao_id,
                numero_cadastro=proximo,
                nome=nome,
                email=_cel(row.get(col_email)) if col_email else None,
                telefone=_cel(row.get(col_telefone)) if col_telefone else None,
                cpf=cpf,
                unidade_captador=_cel(row.get(col_unidade)) if col_unidade else None,
                origem_cadastro=ORIGEM_DOADOR_PLANILHA,
                ativo=True,
                criado_em=agora_operacional_naive(),
                atualizado_em=agora_operacional_naive(),
            )
        )
        proximo += 1
        inseridos += 1
    await db.commit()
    return {"inseridos": inseridos, "ignorados": ignorados}


async def importar_cnpjs(
    db: AsyncSession,
    organizacao_id: str,
    arquivo: BinaryIO,
    captador_padrao: str = "DIEGO",
    competencia: Optional[str] = None,
) -> dict:
    """Importa CNPJs do agente.

    Cadastro mestre: so acrescenta CNPJs novos (nao apaga historico de cadastro).

    Com competencia: a planilha vira o conjunto vigente do captador naquele mes
    (substitui os vinculos anteriores da mesma competencia). Esse conjunto vale
    nos fechamentos seguintes ate uma nova importacao do agente.
    """
    headers, dados = _ler_planilha(arquivo)
    col_cnpj = achar_coluna(headers, ["cnpj"])
    col_loja = achar_coluna(headers, ["nome da loja", "loja", "estabelecimento", "emitente"])
    col_captador = achar_coluna(headers, ["captador", "unidade/captador", "unidade"])
    if not col_cnpj:
        raise ValueError("Nao encontrei a coluna obrigatoria: CNPJ.")

    if competencia and not competencia_valida(competencia):
        raise ValueError("Competencia invalida. Use YYYY-MM.")

    inseridos = 0
    ignorados = 0
    conferir = 0
    vinculos = 0
    vinculos_substituidos = 0
    saidas = 0
    entradas = 0
    competencia_anterior: Optional[str] = None
    linhas_arquivo = 0
    proximo = await proximo_numero_cadastro_cnpj(db, organizacao_id)
    captador_padrao_n = normalizar_agente_captacao(captador_padrao) or "DIEGO"

    # Normaliza linhas da planilha (um CNPJ pode repetir: fica a ultima ocorrencia).
    planilha: dict[str, dict[str, str]] = {}
    for row in dados:
        cnpj = limpar_documento(row.get(col_cnpj))
        if not cnpj:
            continue
        linhas_arquivo += 1
        loja_raw = _cel(row.get(col_loja)) if col_loja else ""
        captador = normalizar_agente_captacao(
            _cel(row.get(col_captador)) if col_captador else captador_padrao_n
        ) or captador_padrao_n
        loja = nome_loja_para_cadastro(cnpj, loja_raw)
        planilha[cnpj] = {"loja": loja, "captador": captador, "loja_raw": loja_raw}

    for cnpj, info in planilha.items():
        loja = info["loja"]
        captador = info["captador"]
        precisa_conferir = loja == NOME_GENERICO_CONFERIR or not cnpj_valido(cnpj)

        atual = (
            await db.execute(
                select(NfpCnpjLojaDB).where(
                    NfpCnpjLojaDB.organizacao_id == organizacao_id,
                    NfpCnpjLojaDB.cnpj == cnpj,
                )
            )
        ).scalar_one_or_none()
        if atual:
            if nome_eh_generico(atual.loja) and not nome_eh_generico(loja) and loja != NOME_GENERICO_CONFERIR:
                atual.loja = loja
                atual.cnpj_conferir = False
            # Cadastro mestre nao remove quem saiu da planilha — o rateio usa o historico.
            ignorados += 1
        else:
            db.add(
                NfpCnpjLojaDB(
                    organizacao_id=organizacao_id,
                    numero_cadastro=proximo,
                    cnpj=cnpj,
                    loja=loja,
                    captador=captador,
                    cnpj_conferir=precisa_conferir,
                    ativo=True,
                    criado_em=agora_operacional_naive(),
                    atualizado_em=agora_operacional_naive(),
                )
            )
            proximo += 1
            inseridos += 1
            if precisa_conferir:
                conferir += 1

    if competencia:
        captadores_planilha = {info["captador"] for info in planilha.values()} or {captador_padrao_n}
        anterior_comp, anterior_cnpjs = await _lista_captacao_anterior(
            db, organizacao_id, competencia, captadores_planilha
        )
        competencia_anterior = anterior_comp
        saidas = len(anterior_cnpjs - set(planilha))
        entradas = len(set(planilha) - anterior_cnpjs)

        # Substitui o conjunto do(s) captador(es) nesta competencia.
        result = await db.execute(
            delete(NfpCnpjCaptacaoCompetenciaDB).where(
                NfpCnpjCaptacaoCompetenciaDB.organizacao_id == organizacao_id,
                NfpCnpjCaptacaoCompetenciaDB.competencia == competencia,
                NfpCnpjCaptacaoCompetenciaDB.captador.in_(sorted(captadores_planilha)),
            )
        )
        vinculos_substituidos = int(result.rowcount or 0)

        agora = agora_operacional_naive()
        for cnpj, info in planilha.items():
            db.add(
                NfpCnpjCaptacaoCompetenciaDB(
                    organizacao_id=organizacao_id,
                    competencia=competencia,
                    cnpj=cnpj,
                    captador=info["captador"],
                    loja=info["loja"],
                    criado_em=agora,
                    atualizado_em=agora,
                )
            )
            vinculos += 1

    await db.commit()
    return {
        "inseridos": inseridos,
        "ignorados": ignorados,
        "cnpj_conferir": conferir,
        "competencia": competencia,
        "linhas_arquivo": linhas_arquivo,
        "vinculos_competencia": vinculos,
        "competencia_anterior": competencia_anterior,
        "saidas": saidas,
        "entradas": entradas,
        "vinculos_substituidos": vinculos_substituidos,
    }


async def _lista_captacao_anterior(
    db: AsyncSession,
    organizacao_id: str,
    competencia: str,
    captadores: Iterable[str],
) -> tuple[Optional[str], set[str]]:
    """Ultima lista do(s) captador(es) antes da competencia informada."""
    alvo = sorted({c for c in captadores if c})
    if not alvo:
        return None, set()

    rows = (
        await db.execute(
            select(
                NfpCnpjCaptacaoCompetenciaDB.competencia,
                NfpCnpjCaptacaoCompetenciaDB.cnpj,
            ).where(
                NfpCnpjCaptacaoCompetenciaDB.organizacao_id == organizacao_id,
                NfpCnpjCaptacaoCompetenciaDB.competencia < competencia,
                NfpCnpjCaptacaoCompetenciaDB.captador.in_(alvo),
            )
        )
    ).all()
    if not rows:
        return None, set()

    ultima = max(r[0] for r in rows)
    return ultima, {limpar_documento(r[1]) for r in rows if r[0] == ultima and limpar_documento(r[1])}


async def mapa_cnpjs_captacao_vigente(
    db: AsyncSession,
    organizacao_id: str,
    competencia: str,
) -> dict[str, dict[str, str]]:
    """Retorna CNPJ -> {captador, loja} pela ultima planilha vigente de cada captador.

    Vigencia: maior competencia <= competencia do fechamento com vinculos daquele captador.
    """
    if not competencia_valida(competencia):
        return {}

    rows = (
        await db.execute(
            select(NfpCnpjCaptacaoCompetenciaDB).where(
                NfpCnpjCaptacaoCompetenciaDB.organizacao_id == organizacao_id,
                NfpCnpjCaptacaoCompetenciaDB.competencia <= competencia,
            )
        )
    ).scalars().all()
    if not rows:
        return {}

    # Ultima competencia com importacao por captador.
    ultima_por_captador: dict[str, str] = {}
    for v in rows:
        captador = normalizar_agente_captacao(v.captador)
        comp = (v.competencia or "").strip()
        if not captador or not competencia_valida(comp):
            continue
        atual = ultima_por_captador.get(captador)
        if not atual or comp > atual:
            ultima_por_captador[captador] = comp

    mapa: dict[str, dict[str, str]] = {}
    for v in rows:
        captador = normalizar_agente_captacao(v.captador)
        if not captador:
            continue
        if v.competencia != ultima_por_captador.get(captador):
            continue
        cnpj = limpar_documento(v.cnpj)
        if not cnpj:
            continue
        mapa[cnpj] = {
            "captador": captador,
            "loja": v.loja or "",
            "competencia_vigencia": v.competencia,
        }
    return mapa


async def importar_doacoes_sefaz(
    db: AsyncSession,
    organizacao_id: str,
    arquivo: BinaryIO,
    competencia: Optional[str] = None,
) -> dict:
    headers, dados = _ler_planilha(arquivo)
    col_numero = achar_coluna(headers, ["número da nota", "numero da nota", "no.", "nota"])
    col_valor = achar_coluna(headers, ["valor da nota", "valor nota"])
    col_data = achar_coluna(headers, ["data da nota", "data nota"])
    col_entidade = achar_coluna(headers, ["cnpj entidade social"])
    col_cpf = achar_coluna(headers, ["cpf doador/cadastrador", "cpf doador", "cpf cadastrador"])
    col_pedido = achar_coluna(headers, ["data do pedido"])
    col_status = achar_coluna(headers, ["status do pedido", "status"])
    col_tipo = achar_coluna(headers, ["tipo da doação", "tipo da doacao", "tipo"])
    col_estab = achar_coluna(headers, ["cnpj estabelecimento", "cnpj do estabelecimento"])
    if not col_numero or not col_estab:
        raise ValueError("Planilha de doacao automatica sem colunas obrigatorias.")

    linhas_brutas = 0
    ignorados_tipo = 0
    linhas = []
    datas_ref = []
    for row in dados:
        linhas_brutas += 1
        tipo = _cel(row.get(col_tipo)) if col_tipo else ""
        if not tipo_eh_doacao_automatica(tipo):
            ignorados_tipo += 1
            continue
        numero = limpar_nota(row.get(col_numero))
        cnpj_estab = limpar_documento(row.get(col_estab))
        if not numero and not cnpj_estab:
            continue
        data_nota = _cel(row.get(col_data)) if col_data else ""
        if " " in data_nota:
            data_nota = data_nota.split(" ")[0]
        if data_nota:
            datas_ref.append(data_nota)
        base = chave_base(cnpj_estab, numero, data_nota)
        linhas.append(
            {
                "numero_nota": numero,
                "valor_nota_cent": valor_para_centavos(row.get(col_valor)) if col_valor else 0,
                "data_nota": data_nota,
                "cnpj_entidade_social": limpar_documento(row.get(col_entidade)) if col_entidade else "",
                "cpf_doador": limpar_documento(row.get(col_cpf)) if col_cpf else "",
                "data_pedido": _cel(row.get(col_pedido)) if col_pedido else "",
                "status_pedido": _cel(row.get(col_status)) if col_status else "",
                "tipo_doacao": tipo,
                "cnpj_estabelecimento": cnpj_estab,
                "chave_base": base,
            }
        )

    if competencia and competencia_valida(competencia):
        competencia_final = competencia
    else:
        competencia_final = competencia_referencia_das_datas(datas_ref)

    for linha in linhas:
        linha["competencia"] = competencia_final

    linhas = _adicionar_ocorrencias(linhas, ["competencia", "chave_base"])
    await db.execute(
        delete(NfpDoacaoAutomaticaDB).where(
            NfpDoacaoAutomaticaDB.organizacao_id == organizacao_id,
            NfpDoacaoAutomaticaDB.competencia == competencia_final,
        )
    )
    await db.execute(
        delete(NfpBatimentoDB).where(
            NfpBatimentoDB.organizacao_id == organizacao_id,
            NfpBatimentoDB.competencia == competencia_final,
        )
    )
    for linha in linhas:
        cent = int(linha["valor_nota_cent"] or 0)
        db.add(
            NfpDoacaoAutomaticaDB(
                organizacao_id=organizacao_id,
                numero_nota=linha["numero_nota"],
                valor_nota=centavos_para_float(cent),
                valor_nota_centavos=cent,
                data_nota=linha["data_nota"],
                cnpj_entidade_social=linha["cnpj_entidade_social"],
                cpf_doador_cadastrador=linha["cpf_doador"],
                data_pedido=linha["data_pedido"],
                status_pedido=linha["status_pedido"],
                tipo_doacao=linha["tipo_doacao"],
                cnpj_estabelecimento=linha["cnpj_estabelecimento"],
                chave=chave_com_ocorrencia(linha["chave_base"], linha["ocorrencia"]),
                competencia=competencia_final,
            )
        )
    await db.commit()
    batidos = await gerar_batimento(db, organizacao_id, competencia_final)
    sync = await sincronizar_doadores_de_doacoes(db, organizacao_id, competencia=competencia_final)
    return {
        "inseridos": len(linhas),
        "linhas_arquivo": linhas_brutas,
        "ignorados_tipo": ignorados_tipo,
        "competencia": competencia_final,
        "batimentos": batidos,
        "doadores_sincronizados": sync,
    }


async def importar_sefaz_creditos(
    db: AsyncSession,
    organizacao_id: str,
    arquivo: Union[BinaryIO, Sequence[BinaryIO | bytes], None] = None,
    competencia: Optional[str] = None,
    arquivos: Optional[Sequence[BinaryIO | bytes]] = None,
) -> dict:
    lista: list[BinaryIO | bytes] = []
    if arquivos:
        lista.extend(list(arquivos))
    if arquivo is not None and not isinstance(arquivo, (list, tuple)):
        lista.append(arquivo)
    if not lista:
        raise ValueError("Nenhum arquivo de creditos informado.")

    headers, dados = _ler_varios_arquivos(lista)
    col_cnpj = achar_coluna(headers, ["cnpj emit.", "cnpj emitente", "cnpj do emitente", "cnpj estabelecimento", "cnpj"])
    col_emitente = achar_coluna(headers, ["emitente", "nome emitente", "estabelecimento", "loja"])
    col_numero = achar_coluna(headers, ["número da nota", "numero da nota", "número nf", "numero nf", "nota", "no."])
    col_emissao = achar_coluna(headers, ["data emissão", "data emissao", "data da emissão", "data da emissao", "data"])
    col_valor = achar_coluna(headers, ["valor nf", "valor da nf", "valor nota", "valor da nota", "valor total"])
    col_registro = achar_coluna(headers, ["data registro", "data do registro", "registro"])
    col_creditos = achar_coluna(headers, ["créditos", "creditos", "valor do crédito", "valor do credito", "crédito", "credito"])
    col_situacao = achar_coluna(headers, ["situação do crédito", "situacao do credito", "situação", "situacao", "status"])
    if not col_cnpj or not col_numero:
        raise ValueError("Planilha SEFAZ sem colunas obrigatorias (CNPJ emitente e numero da nota).")

    linhas = []
    pares_nome = []
    datas_ref = []
    for row in dados:
        cnpj = limpar_documento(row.get(col_cnpj))
        numero = limpar_nota(row.get(col_numero))
        if not cnpj and not numero:
            continue
        emitente = _cel(row.get(col_emitente)) if col_emitente else ""
        data_emissao = _cel(row.get(col_emissao)) if col_emissao else ""
        if " " in data_emissao:
            data_emissao = data_emissao.split(" ")[0]
        if data_emissao:
            datas_ref.append(data_emissao)
        pares_nome.append((cnpj, emitente))
        base = chave_base(cnpj, numero, data_emissao)
        linhas.append(
            {
                "cnpj_emitente": cnpj,
                "emitente": emitente,
                "numero_nota": numero,
                "data_emissao": data_emissao,
                "valor_nf_cent": valor_para_centavos(row.get(col_valor)) if col_valor else 0,
                "data_registro": _cel(row.get(col_registro)) if col_registro else "",
                "creditos_cent": valor_para_centavos(row.get(col_creditos)) if col_creditos else 0,
                "situacao_credito": _cel(row.get(col_situacao)) if col_situacao else "",
                "chave_base": base,
            }
        )

    if competencia and competencia_valida(competencia):
        competencia_final = competencia
    else:
        competencia_final = competencia_referencia_das_datas(datas_ref)

    for linha in linhas:
        linha["competencia"] = competencia_final

    linhas = _adicionar_ocorrencias(linhas, ["competencia", "chave_base"])
    await db.execute(
        delete(NfpSefazCreditoDB).where(
            NfpSefazCreditoDB.organizacao_id == organizacao_id,
            NfpSefazCreditoDB.competencia == competencia_final,
        )
    )
    await db.execute(
        delete(NfpBatimentoDB).where(
            NfpBatimentoDB.organizacao_id == organizacao_id,
            NfpBatimentoDB.competencia == competencia_final,
        )
    )
    for linha in linhas:
        vcent = int(linha["valor_nf_cent"] or 0)
        ccent = int(linha["creditos_cent"] or 0)
        db.add(
            NfpSefazCreditoDB(
                organizacao_id=organizacao_id,
                cnpj_emitente=linha["cnpj_emitente"],
                emitente=linha["emitente"],
                numero_nota=linha["numero_nota"],
                data_emissao=linha["data_emissao"],
                valor_nf=centavos_para_float(vcent),
                valor_nf_centavos=vcent,
                data_registro=linha["data_registro"],
                creditos=centavos_para_float(ccent),
                creditos_centavos=ccent,
                situacao_credito=linha["situacao_credito"],
                chave=chave_com_ocorrencia(linha["chave_base"], linha["ocorrencia"]),
                competencia=competencia_final,
            )
        )
    nomes_ok = await enriquecer_nomes_cnpjs_genericos(db, organizacao_id, pares_nome)
    await db.commit()
    batidos = await gerar_batimento(db, organizacao_id, competencia_final)
    sync = await sincronizar_doadores_de_doacoes(db, organizacao_id, competencia=competencia_final)
    return {
        "inseridos": len(linhas),
        "arquivos": len(lista),
        "competencia": competencia_final,
        "nomes_enriquecidos": nomes_ok,
        "batimentos": batidos,
        "doadores_sincronizados": sync,
    }


async def gerar_batimento(db: AsyncSession, organizacao_id: str, competencia: str) -> int:
    await db.execute(
        delete(NfpBatimentoDB).where(
            NfpBatimentoDB.organizacao_id == organizacao_id,
            NfpBatimentoDB.competencia == competencia,
        )
    )
    doacoes = (
        await db.execute(
            select(NfpDoacaoAutomaticaDB).where(
                NfpDoacaoAutomaticaDB.organizacao_id == organizacao_id,
                NfpDoacaoAutomaticaDB.competencia == competencia,
            )
        )
    ).scalars().all()
    sefaz = (
        await db.execute(
            select(NfpSefazCreditoDB).where(
                NfpSefazCreditoDB.organizacao_id == organizacao_id,
                NfpSefazCreditoDB.competencia == competencia,
            )
        )
    ).scalars().all()

    mapa_sefaz: dict[tuple[str, str], list[NfpSefazCreditoDB]] = defaultdict(list)
    for s in sefaz:
        mapa_sefaz[(limpar_documento(s.cnpj_emitente), limpar_nota(s.numero_nota))].append(s)

    total = 0
    for d in doacoes:
        chave = (limpar_documento(d.cnpj_estabelecimento), limpar_nota(d.numero_nota))
        matches = mapa_sefaz.get(chave) or []
        for idx, s in enumerate(matches, start=1):
            db.add(
                NfpBatimentoDB(
                    organizacao_id=organizacao_id,
                    competencia=competencia,
                    id_doacao=d.id,
                    id_sefaz=s.id,
                    cpf_doador_cadastrador=d.cpf_doador_cadastrador,
                    cnpj_estabelecimento=d.cnpj_estabelecimento,
                    emitente=s.emitente,
                    numero_nota=d.numero_nota,
                    data_emissao=s.data_emissao,
                    data_nota=d.data_nota,
                    ocorrencia=idx,
                    valor_nota_centavos=int(d.valor_nota_centavos or 0),
                    valor_nf_centavos=int(s.valor_nf_centavos or 0),
                    creditos_centavos=int(s.creditos_centavos or 0),
                )
            )
            total += 1
    await db.commit()
    return total


async def listar_agentes_captacao(db: AsyncSession, organizacao_id: str) -> list[str]:
    """Codigos de agentes: cadastro oficial, captadores em CNPJs e padroes."""
    await garantir_agentes_padrao(db, organizacao_id)
    rows_cadastro = (
        await db.execute(
            select(NfpAgenteCaptadorDB.codigo)
            .where(
                NfpAgenteCaptadorDB.organizacao_id == organizacao_id,
                NfpAgenteCaptadorDB.ativo.is_(True),
            )
        )
    ).scalars().all()
    rows_cnpj = (
        await db.execute(
            select(NfpCnpjLojaDB.captador)
            .where(
                NfpCnpjLojaDB.organizacao_id == organizacao_id,
                NfpCnpjLojaDB.captador.is_not(None),
                NfpCnpjLojaDB.captador != "",
            )
            .distinct()
        )
    ).scalars().all()
    agentes = {
        normalizar_agente_captacao(r)
        for r in list(rows_cadastro) + list(rows_cnpj)
        if normalizar_agente_captacao(r)
    }
    for padrao in AGENTES_CAPTACAO_PADRAO:
        agentes.add(padrao)
    return sorted(agentes)


async def percentual_do_agente(
    db: AsyncSession,
    organizacao_id: str,
    codigo: str,
    mapa: Optional[dict[str, int]] = None,
) -> int:
    chave = normalizar_agente_captacao(codigo)
    if not chave:
        return 0
    if mapa is not None:
        if chave in mapa:
            return int(mapa[chave])
        return percentual_agente_padrao(chave)
    mapa_local = await mapa_percentual_agentes(db, organizacao_id)
    return int(mapa_local.get(chave, percentual_agente_padrao(chave)))


async def calcular_rateio(db: AsyncSession, organizacao_id: str, competencia: str) -> dict:
    if not competencia_valida(competencia):
        raise ValueError("Competencia invalida.")

    await garantir_agentes_padrao(db, organizacao_id)
    mapa_pct = await mapa_percentual_agentes(db, organizacao_id)

    cnpjs = (
        await db.execute(
            select(NfpCnpjLojaDB).where(NfpCnpjLojaDB.organizacao_id == organizacao_id)
        )
    ).scalars().all()
    mapa_cnpjs = {}
    for item in cnpjs:
        c = limpar_documento(item.cnpj)
        if not c:
            continue
        captador = normalizar_agente_captacao(item.captador)
        mapa_cnpjs[c] = {"loja": item.loja or "", "captador": captador}

    # Lista vigente por captador: ultima planilha com competencia <= fechamento.
    cnpjs_agente: dict[str, str] = {}
    vigentes = await mapa_cnpjs_captacao_vigente(db, organizacao_id, competencia)
    if vigentes:
        for c, info in vigentes.items():
            captador = normalizar_agente_captacao(info.get("captador"))
            if not c or not captador:
                continue
            cnpjs_agente[c] = captador
            base = mapa_cnpjs.get(c, {"loja": "", "captador": captador})
            if info.get("loja") and not nome_eh_generico(info.get("loja")):
                base["loja"] = info["loja"]
            base["captador"] = captador
            mapa_cnpjs[c] = base
    else:
        # Compatibilidade: bases antigas sem historico por competencia.
        for c, info in mapa_cnpjs.items():
            captador = normalizar_agente_captacao(info.get("captador"))
            if captador:
                cnpjs_agente[c] = captador

    doacoes = (
        await db.execute(
            select(NfpDoacaoAutomaticaDB).where(
                NfpDoacaoAutomaticaDB.organizacao_id == organizacao_id,
                NfpDoacaoAutomaticaDB.competencia == competencia,
            )
        )
    ).scalars().all()
    chaves_auto = {
        (limpar_documento(d.cnpj_estabelecimento), limpar_nota(d.numero_nota))
        for d in doacoes
    }

    sefaz = (
        await db.execute(
            select(NfpSefazCreditoDB).where(
                NfpSefazCreditoDB.organizacao_id == organizacao_id,
                NfpSefazCreditoDB.competencia == competencia,
            )
        )
    ).scalars().all()

    await db.execute(
        delete(NfpRateioDB).where(
            NfpRateioDB.organizacao_id == organizacao_id,
            NfpRateioDB.competencia == competencia,
        )
    )

    grupos: dict[tuple, dict] = {}
    for s in sefaz:
        cnpj = limpar_documento(s.cnpj_emitente)
        numero = limpar_nota(s.numero_nota)
        cred = int(s.creditos_centavos or 0)
        info = mapa_cnpjs.get(cnpj, {})
        loja = info.get("loja") or s.emitente or ""
        captador = normalizar_agente_captacao(info.get("captador"))
        eh_agente = cnpj in cnpjs_agente
        agente = cnpjs_agente.get(cnpj) or captador
        eh_auto = (cnpj, numero) in chaves_auto
        if eh_agente and eh_auto:
            origem = origem_doador_auto_agente(agente)
        elif eh_agente:
            origem = origem_rateio_agente(agente)
        elif eh_auto:
            origem = "DOADOR_AUTOMATICO_AEB"
        else:
            origem = "DIRETO_AEB"
        chave = (cnpj, loja, captador or agente, origem)
        if chave not in grupos:
            grupos[chave] = {"retorno_centavos": 0, "qtd": 0}
        grupos[chave]["retorno_centavos"] += cred
        grupos[chave]["qtd"] += 1

    for (cnpj, loja, captador, origem), dados in grupos.items():
        retorno = int(dados["retorno_centavos"] or 0)
        qtd = int(dados["qtd"] or 0)
        if origem_eh_rateio_agente(origem):
            pct = mapa_pct.get(normalizar_agente_captacao(captador or origem), percentual_agente_padrao(origem))
            valor_agente, valor_aeb = rateio_centavos(retorno, pct)
        else:
            valor_agente = 0
            valor_aeb = retorno
        db.add(
            NfpRateioDB(
                organizacao_id=organizacao_id,
                cnpj=cnpj,
                loja=loja,
                captador=captador or None,
                origem=origem,
                retorno=centavos_para_float(retorno),
                retorno_centavos=retorno,
                qtd=qtd,
                aeb=centavos_para_float(valor_aeb),
                aeb_centavos=valor_aeb,
                credito_liquido=centavos_para_float(retorno),
                credito_liquido_centavos=retorno,
                # Coluna historica valor_diego_* = parte do agente de captacao.
                valor_diego=centavos_para_float(valor_agente),
                valor_diego_centavos=valor_agente,
                valor_aeb=centavos_para_float(valor_aeb),
                valor_aeb_centavos=valor_aeb,
                final=centavos_para_float(retorno),
                final_centavos=retorno,
                competencia=competencia,
            )
        )
    await db.commit()
    return {"grupos": len(grupos), "competencia": competencia}


async def resumo_dashboard(
    db: AsyncSession,
    organizacao_id: str,
    competencia: Optional[str] = None,
    agente: Optional[str] = None,
) -> dict:
    async def count(model, **filtros):
        q = select(func.count()).select_from(model).where(model.organizacao_id == organizacao_id)
        for k, v in filtros.items():
            q = q.where(getattr(model, k) == v)
        return int((await db.execute(q)).scalar_one() or 0)

    comps = (
        await db.execute(
            select(NfpSefazCreditoDB.competencia)
            .where(NfpSefazCreditoDB.organizacao_id == organizacao_id)
            .distinct()
            .order_by(NfpSefazCreditoDB.competencia.desc())
        )
    ).scalars().all()
    competencia_atual = competencia if competencia_valida(competencia or "") else (comps[0] if comps else None)

    agentes = await listar_agentes_captacao(db, organizacao_id)
    agente_in = (agente or "").strip()
    visao_todos = (not agente_in) or agente_in.upper() in {"TODOS", "TODO", "ALL", "*"}
    if visao_todos:
        agente_sel = ""
    else:
        agente_sel = normalizar_agente_captacao(agente_in)
        if agente_sel not in agentes:
            agente_sel = "DIEGO" if "DIEGO" in agentes else (agentes[0] if agentes else "")

    percentual_agente = await percentual_do_agente(db, organizacao_id, agente_sel) if agente_sel else 0

    total_creditos = 0
    parte_agente = 0
    parte_aeb = 0
    bruto_lojas_agente = 0
    doador_aeb_em_loja_agente = 0
    doador_automatico_total = 0
    direto_aeb = 0
    cnpjs_agente = 0

    if agente_sel:
        cnpjs_agente = await count(NfpCnpjLojaDB, captador=agente_sel)
    else:
        cnpjs_agente = await count(NfpCnpjLojaDB)

    origem_agente = origem_rateio_agente(agente_sel) if agente_sel else ""
    origem_auto_agente = origem_doador_auto_agente(agente_sel) if agente_sel else ""

    if competencia_atual:
        rows = (
            await db.execute(
                select(
                    NfpRateioDB.origem,
                    func.coalesce(func.sum(NfpRateioDB.retorno_centavos), 0),
                    func.coalesce(func.sum(NfpRateioDB.valor_diego_centavos), 0),
                    func.coalesce(func.sum(NfpRateioDB.valor_aeb_centavos), 0),
                )
                .where(
                    NfpRateioDB.organizacao_id == organizacao_id,
                    NfpRateioDB.competencia == competencia_atual,
                )
                .group_by(NfpRateioDB.origem)
            )
        ).all()

        for origem, retorno_c, agente_c, aeb_c in rows:
            origem_n = normalizar_agente_captacao(origem)
            retorno_c = int(retorno_c or 0)
            agente_c = int(agente_c or 0)
            aeb_c = int(aeb_c or 0)
            total_creditos += retorno_c
            parte_aeb += aeb_c

            if agente_sel:
                if origem_n == origem_agente:
                    bruto_lojas_agente += retorno_c
                    parte_agente += agente_c
                elif origem_n == origem_auto_agente:
                    doador_aeb_em_loja_agente += retorno_c
                    doador_automatico_total += retorno_c
                elif origem_n.startswith("DOADOR_AUTOMATICO"):
                    doador_automatico_total += retorno_c
                elif origem_n == "DIRETO_AEB":
                    direto_aeb += retorno_c
            else:
                if origem_eh_rateio_agente(origem_n):
                    bruto_lojas_agente += retorno_c
                    parte_agente += agente_c
                elif origem_n.startswith("DOADOR_AUTOMATICO") and origem_n != "DOADOR_AUTOMATICO_AEB":
                    doador_aeb_em_loja_agente += retorno_c
                    doador_automatico_total += retorno_c
                elif origem_n.startswith("DOADOR_AUTOMATICO"):
                    doador_automatico_total += retorno_c
                elif origem_n == "DIRETO_AEB":
                    direto_aeb += retorno_c

    return {
        "competencia": competencia_atual,
        "competencias": list(comps),
        "agentes_captacao": agentes,
        "agente_captacao": agente_sel or None,
        "visao_todos": not bool(agente_sel),
        "percentual_agente": percentual_agente,
        "percentual_aeb": 100 - int(percentual_agente or 0) if agente_sel else None,
        "doadores": await count(NfpDoadorDB),
        "doadores_diretos_aeb": await count(NfpDoadorDB),
        "cnpjs": await count(NfpCnpjLojaDB),
        "cnpjs_agente": cnpjs_agente,
        "cnpjs_conferir": await count(NfpCnpjLojaDB, cnpj_conferir=True),
        "sefaz_creditos": await count(NfpSefazCreditoDB, competencia=competencia_atual) if competencia_atual else 0,
        "doacoes_automaticas": await count(NfpDoacaoAutomaticaDB, competencia=competencia_atual) if competencia_atual else 0,
        "batimentos": await count(NfpBatimentoDB, competencia=competencia_atual) if competencia_atual else 0,
        "rateio_grupos": await count(NfpRateioDB, competencia=competencia_atual) if competencia_atual else 0,
        "total_creditos_centavos": total_creditos,
        "total_diego_centavos": parte_agente,
        "total_aeb_centavos": parte_aeb,
        "total_creditos": centavos_para_float(total_creditos),
        "total_diego": centavos_para_float(parte_agente),
        "total_aeb": centavos_para_float(parte_aeb),
        "bruto_lojas_agente_centavos": bruto_lojas_agente,
        "bruto_lojas_agente": centavos_para_float(bruto_lojas_agente),
        "parte_agente_centavos": parte_agente,
        "parte_agente": centavos_para_float(parte_agente),
        "parte_agente_50_centavos": parte_agente,
        "parte_agente_50": centavos_para_float(parte_agente),
        "parte_aeb_lojas_agente_centavos": max(0, bruto_lojas_agente - parte_agente),
        "parte_aeb_lojas_agente": centavos_para_float(max(0, bruto_lojas_agente - parte_agente)),
        "parte_aeb_50": centavos_para_float(max(0, bruto_lojas_agente - parte_agente)),
        "doador_aeb_loja_agente_centavos": doador_aeb_em_loja_agente,
        "doador_aeb_loja_agente": centavos_para_float(doador_aeb_em_loja_agente),
        "doador_automatico_total_centavos": doador_automatico_total,
        "doador_automatico_total": centavos_para_float(doador_automatico_total),
        "doador_automatico_aeb_centavos": max(0, doador_automatico_total - doador_aeb_em_loja_agente),
        "doador_automatico_aeb": centavos_para_float(max(0, doador_automatico_total - doador_aeb_em_loja_agente)),
        "direto_aeb_centavos": direto_aeb,
        "direto_aeb": centavos_para_float(direto_aeb),
        "aeb_no_rateio_agente_centavos": max(0, bruto_lojas_agente - parte_agente),
        "aeb_no_rateio_agente": centavos_para_float(max(0, bruto_lojas_agente - parte_agente)),
        "aeb_fora_rateio_centavos": doador_automatico_total + direto_aeb,
        "aeb_fora_rateio": centavos_para_float(doador_automatico_total + direto_aeb),
        "aeb_total_competencia_centavos": parte_aeb,
        "aeb_total_competencia": centavos_para_float(parte_aeb),
        "parte_aeb_consolidada_agente_centavos": max(0, bruto_lojas_agente - parte_agente) + doador_aeb_em_loja_agente,
        "parte_aeb_consolidada_agente": centavos_para_float(
            max(0, bruto_lojas_agente - parte_agente) + doador_aeb_em_loja_agente
        ),
        # Compatibilidade com cards anteriores
        "bruto_lojas_diego": centavos_para_float(bruto_lojas_agente),
        "parte_diego_50": centavos_para_float(parte_agente),
        "doador_aeb_loja_diego": centavos_para_float(doador_aeb_em_loja_agente),
    }


def exportar_rateio_xlsx(linhas: list[NfpRateioDB]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "rateio"
    ws.append(
        [
            "cnpj", "loja", "captador", "origem", "qtd",
            "retorno", "valor_diego", "valor_aeb", "final", "competencia",
        ]
    )
    for r in linhas:
        ws.append(
            [
                r.cnpj, r.loja, r.captador, r.origem, r.qtd,
                r.retorno, r.valor_diego, r.valor_aeb, r.final, r.competencia,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bucket_origem_rateio(origem: Optional[str], agente_filtro: str = "") -> str:
    o = normalizar_agente_captacao(origem)
    ag = normalizar_agente_captacao(agente_filtro)
    if not o:
        return "outros"
    if o == "DIRETO_AEB":
        return "direto_aeb"
    if o.startswith("DOADOR_AUTOMATICO"):
        if ag and o == origem_doador_auto_agente(ag):
            return "doador_auto_agente"
        if o == "DOADOR_AUTOMATICO_AEB" or o.startswith("DOADOR_AUTOMATICO"):
            return "doador_auto"
        return "doador_auto"
    if origem_eh_rateio_agente(o):
        if ag and o != ag:
            return "outros_agentes"
        return "rateio_agente"
    return "outros"


async def relatorio_rateio_consolidado(
    db: AsyncSession,
    organizacao_id: str,
    competencia_inicio: Optional[str] = None,
    competencia_fim: Optional[str] = None,
    agente: Optional[str] = None,
) -> dict:
    agente_sel = normalizar_agente_captacao(agente)
    q = select(NfpRateioDB).where(NfpRateioDB.organizacao_id == organizacao_id)
    if competencia_inicio and competencia_valida(competencia_inicio):
        q = q.where(NfpRateioDB.competencia >= competencia_inicio)
    if competencia_fim and competencia_valida(competencia_fim):
        q = q.where(NfpRateioDB.competencia <= competencia_fim)
    rows = (await db.execute(q.order_by(NfpRateioDB.competencia, NfpRateioDB.captador))).scalars().all()

    por_comp: dict[str, dict] = {}
    por_agente: dict[str, dict] = {}
    totais = {
        "total_creditos": 0.0,
        "parte_agente": 0.0,
        "parte_aeb": 0.0,
        "doador_auto": 0.0,
        "direto_aeb": 0.0,
        "qtd_linhas": 0,
    }

    def _acc(destino: dict, bucket: str, retorno: float, parte_ag: float, parte_aeb: float) -> None:
        destino["total_creditos"] = float(destino.get("total_creditos") or 0) + retorno
        destino["parte_agente"] = float(destino.get("parte_agente") or 0) + parte_ag
        destino["parte_aeb"] = float(destino.get("parte_aeb") or 0) + parte_aeb
        if bucket in {"doador_auto", "doador_auto_agente"}:
            destino["doador_auto"] = float(destino.get("doador_auto") or 0) + retorno
        if bucket == "direto_aeb":
            destino["direto_aeb"] = float(destino.get("direto_aeb") or 0) + retorno
        destino["qtd_linhas"] = int(destino.get("qtd_linhas") or 0) + 1

    for r in rows:
        origem_n = normalizar_agente_captacao(r.origem)
        captador_n = normalizar_agente_captacao(r.captador) or origem_n
        if agente_sel:
            # Inclui linhas do agente (rateio e doador auto na loja dele) + opcionalmente ignora outros.
            if origem_n not in {agente_sel, origem_doador_auto_agente(agente_sel)} and captador_n != agente_sel:
                continue

        bucket = _bucket_origem_rateio(r.origem, agente_sel)
        retorno = float(r.retorno or 0)
        parte_ag = float(r.valor_diego or 0)
        parte_aeb = float(r.valor_aeb or 0)
        competencia = r.competencia or ""

        if competencia not in por_comp:
            por_comp[competencia] = {
                "competencia": competencia,
                "total_creditos": 0.0,
                "parte_agente": 0.0,
                "parte_aeb": 0.0,
                "doador_auto": 0.0,
                "direto_aeb": 0.0,
                "qtd_linhas": 0,
            }
        _acc(por_comp[competencia], bucket, retorno, parte_ag, parte_aeb)
        _acc(totais, bucket, retorno, parte_ag, parte_aeb)

        chave_ag = captador_n or "SEM_CAPTADOR"
        if chave_ag not in por_agente:
            por_agente[chave_ag] = {
                "agente": chave_ag,
                "total_creditos": 0.0,
                "parte_agente": 0.0,
                "parte_aeb": 0.0,
                "doador_auto": 0.0,
                "direto_aeb": 0.0,
                "qtd_linhas": 0,
            }
        _acc(por_agente[chave_ag], bucket, retorno, parte_ag, parte_aeb)

    return {
        "competencia_inicio": competencia_inicio,
        "competencia_fim": competencia_fim,
        "agente": agente_sel or None,
        "totais": totais,
        "por_competencia": [por_comp[k] for k in sorted(por_comp.keys())],
        "por_agente": sorted(por_agente.values(), key=lambda x: (-float(x["total_creditos"]), x["agente"])),
    }


async def relatorio_rateio_detalhado(
    db: AsyncSession,
    organizacao_id: str,
    competencia: str,
    agente: Optional[str] = None,
    origem: Optional[str] = None,
    busca: Optional[str] = None,
    limite: int = 2000,
) -> dict:
    if not competencia_valida(competencia):
        raise ValueError("Competencia invalida. Use YYYY-MM.")

    agente_sel = normalizar_agente_captacao(agente)
    origem_sel = normalizar_agente_captacao(origem)
    q = select(NfpRateioDB).where(
        NfpRateioDB.organizacao_id == organizacao_id,
        NfpRateioDB.competencia == competencia,
    )
    if agente_sel:
        q = q.where(
            (NfpRateioDB.captador == agente_sel)
            | (NfpRateioDB.origem == agente_sel)
            | (NfpRateioDB.origem == origem_doador_auto_agente(agente_sel))
        )
    if origem_sel:
        q = q.where(NfpRateioDB.origem == origem_sel)
    if busca:
        termo = f"%{busca.strip()}%"
        q = q.where((NfpRateioDB.loja.ilike(termo)) | (NfpRateioDB.cnpj.ilike(termo)))

    rows = (
        await db.execute(
            q.order_by(NfpRateioDB.origem, NfpRateioDB.loja).limit(max(1, min(limite, 5000)))
        )
    ).scalars().all()

    linhas = []
    totais = {
        "total_creditos": 0.0,
        "parte_agente": 0.0,
        "parte_aeb": 0.0,
        "qtd_linhas": 0,
        "qtd_notas": 0,
    }
    for r in rows:
        retorno = float(r.retorno or 0)
        parte_ag = float(r.valor_diego or 0)
        parte_aeb = float(r.valor_aeb or 0)
        qtd = int(r.qtd or 0)
        linhas.append(
            {
                "id": r.id,
                "cnpj": r.cnpj,
                "loja": r.loja,
                "captador": r.captador,
                "origem": r.origem,
                "qtd": qtd,
                "retorno": retorno,
                "valor_agente": parte_ag,
                "valor_diego": parte_ag,
                "valor_aeb": parte_aeb,
                "final": float(r.final or 0),
                "competencia": r.competencia,
            }
        )
        totais["total_creditos"] += retorno
        totais["parte_agente"] += parte_ag
        totais["parte_aeb"] += parte_aeb
        totais["qtd_linhas"] += 1
        totais["qtd_notas"] += qtd

    return {
        "competencia": competencia,
        "agente": agente_sel or None,
        "origem": origem_sel or None,
        "busca": (busca or "").strip() or None,
        "totais": totais,
        "linhas": linhas,
    }


async def listar_origens_rateio(
    db: AsyncSession,
    organizacao_id: str,
    competencia: str,
    agente: Optional[str] = None,
) -> dict:
    if not competencia_valida(competencia):
        raise ValueError("Competencia invalida. Use YYYY-MM.")

    agente_sel = normalizar_agente_captacao(agente)
    q = (
        select(NfpRateioDB.origem)
        .where(
            NfpRateioDB.organizacao_id == organizacao_id,
            NfpRateioDB.competencia == competencia,
            NfpRateioDB.origem.is_not(None),
            NfpRateioDB.origem != "",
        )
        .distinct()
        .order_by(NfpRateioDB.origem)
    )
    if agente_sel:
        q = q.where(
            (NfpRateioDB.captador == agente_sel)
            | (NfpRateioDB.origem == agente_sel)
            | (NfpRateioDB.origem == origem_doador_auto_agente(agente_sel))
        )

    rows = (await db.execute(q)).scalars().all()
    origens = [normalizar_agente_captacao(o) for o in rows if normalizar_agente_captacao(o)]
    return {
        "competencia": competencia,
        "agente": agente_sel or None,
        "origens": origens,
    }
