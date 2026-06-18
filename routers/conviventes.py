# =====================================================================
import json
import os
import re
import uuid
import calendar
import unicodedata
from io import BytesIO
from datetime import date, datetime, timedelta
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Form, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import and_, case, cast, delete, exists, func, or_, String

from audit_log import registrar_evento_auditoria
from database import get_db
from models import (
    ConviventeDB, MotivoInativacaoDB, OrigemEncaminhamentoDB, 
    QuartoDB, LeitoDB, DocumentoConviventeDB, OcorrenciaConviventeDB, UsuarioDB,
    InteracaoOcorrenciaDB, ObservadorOcorrenciaDB, RegistroRotinaDB,
    FechamentoMensalDB,
    SisaLancamentoDB,
    SisaImportacaoDB,
    SisaPresencaImportadaDB,
    SisaDivergenciaDB,
    RegistroPIADB,
    HistoricoConviventeDB,
    HistoricoLegadoSIATDB,
    HistoricoLegadoRotinaSIATDB,
    AusenciaJustificadaConfirmacaoDB,
    LavanderiaRegistroDB,
    PertenceRecolhidoBaixaDB,
    PertenceRecolhidoDB,
)

ORIGEM_HISTORICO_ROTINA_OPERACIONAL = "Rotina operacional"
REGISTROS_POR_PAGINA_PADRAO = 30
LISTAGEM_OPERACIONAL_DIAS_PADRAO = 7
EXPORT_ROTINA_HISTORICO_LIMITE_MAX = 5000


def _normalizar_periodo_listagem_operacional(
    data_inicio: str | None,
    data_fim: str | None,
    *,
    aplicar_padrao: bool = True,
):
    if data_inicio or data_fim or not aplicar_padrao:
        return data_inicio, data_fim

    agora = agora_sao_paulo()
    fim = agora.date()
    inicio = fim - timedelta(days=LISTAGEM_OPERACIONAL_DIAS_PADRAO - 1)
    return inicio.isoformat(), fim.isoformat()
from schemas import (
    ConviventeCreate, ConviventeUpdate, ConviventeResponse,
    MotivoInativacaoCreate, MotivoInativacaoResponse,
    OrigemEncaminhamentoCreate, OrigemEncaminhamentoResponse,
    DocumentoResponse, OcorrenciaResponse, OcorrenciaConviventeListaResponse, InteracaoOcorrenciaCreate, OcorrenciaCreate,
    OcorrenciaRelatorioPrioridades, OcorrenciaPrioridadeResumo,
    RegistroPIACreate, RegistroPIAResponse, RegistroPIAListaResponse,
    DocumentoListaResponse,
    HistoricoConviventeCreate, HistoricoConviventeListaResponse, HistoricoConviventeResponse, HistoricoConviventeUpdate,
    RotinaHistoricoListaResponse, RotinaHistoricoResumoPeriodo,
    RelatorioPiaListaResponse,
    SisaImportacaoListaResponse, FechamentoMensalListaResponse,
    RegistroRotinaCreate,
    RegistroRotinaResponse,
    RegistroRotinaEdicao,
    RegistroRotinaCancelamento,
    RegistroRotinaDesfazerRapido,
    FechamentoMensalCreate,
    FechamentoMensalReabertura,
    FechamentoMensalResponse,
    SisaLancamentoCreate,
    SisaLancamentoResponse,
    SisaImportacaoResponse,
    SisaImportacaoDetalheResponse,
    SisaPreviaImportacaoResponse,
    SisaAcoesImportacao,
    SisaDivergenciaResponse,
    SisaDivergenciaListaResponse,
    SisaDivergenciaStatusUpdate,
    AusenciaJustificadaPendenteResponse,
    AusenciaJustificadaResposta,
)
from security import (
    bloquear_usuario_global_puro,
    get_usuario_logado,
    normalizar_perfil_acesso,
    PERFIL_GLOBAL,
    PERFIL_MANUTENCAO,
    PERFIL_ORIENTADOR,
    PERFIL_TECNICO,
    usuario_tem_perfil,
    usuario_eh_gestor,
    usuario_eh_manutencao,
)
from routers.conviventes_helpers import (
    PESO_PRIORIDADE,
    PRIORIDADES_OCORRENCIA,
    agora_sao_paulo,
    aplicar_credenciais_convivente_salvar,
    convivente_para_response,
    normalizar_prioridade_ocorrencia,
    usuario_pode_resolver_ocorrencia,
    usuario_pode_ver_credenciais_cofre_convivente,
)
from routers.conviventes_documentos import (
    TAMANHO_MAXIMO_DOCUMENTO_BYTES,
    UPLOAD_DIR,
    UPLOAD_DIR_ABSOLUTO,
    caminho_absoluto_documento,
    remover_documentos_foto_perfil,
    validar_upload_documento,
)
from routers.conviventes_xlsx import (
    gerar_xlsx_convenio_sisa_mensal,
    gerar_xlsx_historico,
)
from imagem_upload import eh_arquivo_imagem, padronizar_upload_imagem
from tenant_scope import obter_instituicao_escopo

router = APIRouter(prefix="/api", tags=["Conviventes e Ocorrencias"])

TIPOS_ROTINA_PRINCIPAIS = {"Entrada", "Saída"}
TIPOS_ROTINA_REFEICOES = {"Café da manhã", "Almoço", "Jantar", "Lanche noturno"}
TIPOS_ROTINA_INTERACOES_SIMPLES = {"Banho"}
TIPOS_ROTINA_PARES = {
    "Retirada de Cobertor",
    "Entrega de Cobertor",
    "Retirada de Toalha",
    "Entrega de Toalha",
}
TIPOS_ROTINA_COM_OBSERVACAO_OBRIGATORIA = {
    "Movimentação de Bagageiro",
    "Bipar documentos guardados",
    "Bipar documentos retirados",
}
TIPOS_ROTINA_VALIDOS = (
    TIPOS_ROTINA_PRINCIPAIS
    | TIPOS_ROTINA_REFEICOES
    | TIPOS_ROTINA_INTERACOES_SIMPLES
    | TIPOS_ROTINA_PARES
    | TIPOS_ROTINA_COM_OBSERVACAO_OBRIGATORIA
)
DEPENDENCIAS_EXCLUSAO_CONVIVENTE = (
    (RegistroRotinaDB, "registros de rotina"),
    (OcorrenciaConviventeDB, "ocorrências"),
    (HistoricoConviventeDB, "históricos do prontuário"),
    (RegistroPIADB, "registros PIA"),
    (DocumentoConviventeDB, "documentos/anexos"),
    (LavanderiaRegistroDB, "registros de lavanderia"),
    (PertenceRecolhidoBaixaDB, "baixas de pertences"),
    (HistoricoLegadoRotinaSIATDB, "histórico legado de rotina"),
    (SisaLancamentoDB, "lançamentos SISA"),
    (SisaPresencaImportadaDB, "presenças SISA importadas"),
    (SisaDivergenciaDB, "divergências SISA"),
    (AusenciaJustificadaConfirmacaoDB, "confirmações de ausência justificada"),
)

ROTULOS_REFEICOES_EXTRAS = {
    "Café da manhã": "café da manhã",
    "Almoço": "almoço",
    "Jantar": "jantar",
    "Lanche noturno": "lanche noturno",
}


async def validar_leito_do_projeto(
    db: AsyncSession,
    leito_id: str | None,
    instituicao_id: str,
) -> None:
    if not leito_id:
        return

    leito = (
        await db.execute(
            select(LeitoDB)
            .join(QuartoDB, QuartoDB.id == LeitoDB.quarto_id)
            .where(
                LeitoDB.id == leito_id,
                QuartoDB.instituicao_id == instituicao_id,
            )
        )
    ).scalar_one_or_none()

    if not leito:
        raise HTTPException(status_code=404, detail="Leito não encontrado neste projeto.")


async def validar_usuario_do_projeto(
    db: AsyncSession,
    usuario_id: str | None,
    instituicao_id: str,
    *,
    detail: str = "Usuário não encontrado neste projeto.",
    exigir_ativo: bool = False,
    perfis_permitidos: list[str] | None = None,
) -> None:
    if not usuario_id:
        return

    filtros = [
        UsuarioDB.id == usuario_id,
        UsuarioDB.instituicao_id == instituicao_id,
    ]
    if exigir_ativo:
        filtros.append(UsuarioDB.ativo == True)  # noqa: E712
    if perfis_permitidos:
        filtros.append(UsuarioDB.perfil_acesso.in_(perfis_permitidos))

    usuario = (
        await db.execute(
            select(UsuarioDB.id).where(*filtros)
        )
    ).scalar_one_or_none()

    if not usuario:
        raise HTTPException(status_code=404, detail=detail)


async def validar_usuarios_do_projeto(
    db: AsyncSession,
    usuarios_ids: list[str] | None,
    instituicao_id: str,
    *,
    detail: str = "Um ou mais usuários não pertencem ao projeto atual.",
) -> None:
    ids = list(dict.fromkeys([usuario_id for usuario_id in (usuarios_ids or []) if usuario_id]))
    if not ids:
        return

    encontrados = (
        await db.execute(
            select(UsuarioDB.id).where(
                UsuarioDB.instituicao_id == instituicao_id,
                UsuarioDB.id.in_(ids),
            )
        )
    ).scalars().all()

    if len(set(encontrados)) != len(ids):
        raise HTTPException(status_code=400, detail=detail)


def normalizar_id_tecnico_ocorrencia(valor: str | None) -> str | None:
    tecnico_id = str(valor or "").strip()
    if not tecnico_id or tecnico_id.lower() in {"null", "none", "undefined"}:
        return None
    return tecnico_id


def tecnico_responsavel_ocorrencia_convivente(convivente: ConviventeDB, tecnico_responsavel_id: str | None) -> str | None:
    tecnico_convivente = normalizar_id_tecnico_ocorrencia(getattr(convivente, "tecnico_id", None))
    tecnico_informado = normalizar_id_tecnico_ocorrencia(tecnico_responsavel_id)

    if not tecnico_convivente:
        return None

    if tecnico_informado and str(tecnico_informado) != str(tecnico_convivente):
        raise HTTPException(
            status_code=400,
            detail="Técnico responsável da ocorrência deve ser o técnico vinculado ao convivente.",
        )

    return str(tecnico_convivente)


def status_inicial_ocorrencia_manual() -> str:
    return "Pendente"


async def validar_tecnico_responsavel_ocorrencia(
    db: AsyncSession,
    convivente: ConviventeDB,
    tecnico_responsavel_id: str | None,
    instituicao_id: str,
) -> str | None:
    tecnico_id = tecnico_responsavel_ocorrencia_convivente(
        convivente,
        tecnico_responsavel_id,
    )
    if not tecnico_id:
        return None

    tecnico_ativo = (
        await db.execute(
            select(UsuarioDB.id).where(
                UsuarioDB.id == tecnico_id,
                UsuarioDB.instituicao_id == instituicao_id,
                UsuarioDB.ativo == True,  # noqa: E712
                UsuarioDB.perfil_acesso.in_(["Técnico", "Gestor"]),
            )
        )
    ).scalar_one_or_none()

    return str(tecnico_ativo) if tecnico_ativo else None


async def validar_referencias_convivente_do_projeto(
    db: AsyncSession,
    dados: dict,
    instituicao_id: str,
) -> None:
    await validar_leito_do_projeto(db, dados.get("leito_id"), instituicao_id)
    await validar_usuario_do_projeto(
        db,
        dados.get("tecnico_id"),
        instituicao_id,
        detail="Técnico responsável não está ativo neste projeto.",
        exigir_ativo=True,
        perfis_permitidos=["Técnico", "Gestor"],
    )

    motivo_id = dados.get("motivo_inativacao_id")
    if motivo_id:
        motivo = (
            await db.execute(
                select(MotivoInativacaoDB.id).where(
                    MotivoInativacaoDB.id == motivo_id,
                    MotivoInativacaoDB.instituicao_id == instituicao_id,
                )
            )
        ).scalar_one_or_none()
        if not motivo:
            raise HTTPException(status_code=404, detail="Motivo de inativação não encontrado neste projeto.")

    origem_id = dados.get("origem_encaminhamento_id")
    if origem_id:
        origem = (
            await db.execute(
                select(OrigemEncaminhamentoDB.id).where(
                    OrigemEncaminhamentoDB.id == origem_id,
                    OrigemEncaminhamentoDB.instituicao_id == instituicao_id,
                )
            )
        ).scalar_one_or_none()
        if not origem:
            raise HTTPException(status_code=404, detail="Origem de encaminhamento não encontrada neste projeto.")


async def verificar_mes_fechado(
    db: AsyncSession,
    usuario_atual: dict,
    data_referencia: datetime,
    acao: str = "alterar registros"
):
    """
    Bloqueia alterações operacionais em períodos já fechados.

    Considera mês fechado apenas quando existe registro em fechamentos_mensais
    com status = 'Fechado'. Um mês reaberto não bloqueia novas alterações.
    """
    fechamento = (
        await db.execute(
            select(FechamentoMensalDB).where(
                FechamentoMensalDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                FechamentoMensalDB.ano == data_referencia.year,
                FechamentoMensalDB.mes == data_referencia.month,
                FechamentoMensalDB.status == "Fechado"
            )
        )
    ).scalar_one_or_none()

    if fechamento:
        raise HTTPException(
            status_code=403,
            detail=(
                f"Mês {str(data_referencia.month).zfill(2)}/{data_referencia.year} "
                f"está fechado. Não é possível {acao}. "
                "Reabra o mês pela gestão antes de prosseguir."
            )
        )

    return None

@router.get("/tecnicos")
async def listar_tecnicos(db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    query = select(
        UsuarioDB.id,
        UsuarioDB.nome,
        UsuarioDB.perfil_acesso,
        UsuarioDB.avatar_url,
    ).where(
        UsuarioDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
        UsuarioDB.ativo == True,  # noqa: E712
        UsuarioDB.perfil_acesso.in_(["Técnico", "Gestor"]),
    )
    result = await db.execute(query)
    return [
        {
            "id": t.id,
            "nome": t.nome,
            "perfil_acesso": t.perfil_acesso,
            "avatar_url": t.avatar_url,
        }
        for t in result.all()
    ]


@router.get("/equipe")
async def listar_equipe_operacional(
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    query = select(
        UsuarioDB.id,
        UsuarioDB.nome,
        UsuarioDB.perfil_acesso,
        UsuarioDB.avatar_url,
        UsuarioDB.email,
    ).where(
        UsuarioDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
        UsuarioDB.ativo == True,  # noqa: E712
        UsuarioDB.is_global == False,  # noqa: E712
        ~func.lower(func.trim(UsuarioDB.perfil_acesso)).in_(
            ["global", "executivo", "manutenção", "manutencao"]
        ),
    ).order_by(UsuarioDB.nome.asc())
    result = await db.execute(query)
    return [
        {
            "id": usuario.id,
            "nome": usuario.nome,
            "perfil_acesso": usuario.perfil_acesso,
            "avatar_url": usuario.avatar_url,
            "email": usuario.email,
        }
        for usuario in result.all()
        if usuario_visivel_como_equipe(usuario)
    ]


@router.get("/motivos-inteligentes")
async def listar_motivos_inteligentes(db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    query = select(OcorrenciaConviventeDB.motivo).where(OcorrenciaConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual)).distinct()
    return [m for m in (await db.execute(query)).scalars().all() if m and m.strip()]

@router.get("/motivos-inativacao", response_model=List[MotivoInativacaoResponse])
async def listar_motivos(db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    return (await db.execute(select(MotivoInativacaoDB).where(MotivoInativacaoDB.instituicao_id == obter_instituicao_escopo(usuario_atual)))).scalars().all()

@router.post("/motivos-inativacao", response_model=MotivoInativacaoResponse)
async def criar_motivo(motivo: MotivoInativacaoCreate, db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    bloquear_usuario_global_puro(usuario_atual)
    novo_motivo = MotivoInativacaoDB(instituicao_id=obter_instituicao_escopo(usuario_atual), descricao=motivo.descricao)
    db.add(novo_motivo)
    await db.commit()
    await db.refresh(novo_motivo)
    return novo_motivo

@router.get("/origens-encaminhamento", response_model=List[OrigemEncaminhamentoResponse])
async def listar_origens(db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    return (await db.execute(
        select(OrigemEncaminhamentoDB)
        .where(OrigemEncaminhamentoDB.instituicao_id == obter_instituicao_escopo(usuario_atual))
        .order_by(OrigemEncaminhamentoDB.descricao.asc())
    )).scalars().all()

@router.post("/origens-encaminhamento", response_model=OrigemEncaminhamentoResponse)
async def criar_origem(origem: OrigemEncaminhamentoCreate, db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    bloquear_usuario_global_puro(usuario_atual)
    nova_origem = OrigemEncaminhamentoDB(instituicao_id=obter_instituicao_escopo(usuario_atual), descricao=origem.descricao)
    db.add(nova_origem)
    await db.commit()
    await db.refresh(nova_origem)
    return nova_origem


@router.post("/conviventes", response_model=ConviventeResponse)
async def criar_convivente(convivente: ConviventeCreate, db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    bloquear_usuario_global_puro(usuario_atual)
    try:
        obs_status = getattr(convivente, "observacao_status", None)
        dados = convivente.model_dump(exclude_unset=True, exclude={"observacao_status"})
        
        if dados.get("status") in ["Inativado", "Bloqueado", "Saída qualificada"]:
            dados["leito_id"] = None
            dados["inativado_em"] = agora_sao_paulo()
            dados["ausencia_justificada_desde"] = None
        elif dados.get("status") == "Ausência justificada":
            dados["ausencia_justificada_desde"] = agora_sao_paulo().date()
            dados["inativado_em"] = None
        elif dados.get("status") == "Ativo":
            dados["ausencia_justificada_desde"] = None

        await validar_referencias_convivente_do_projeto(
            db,
            dados,
            obter_instituicao_escopo(usuario_atual),
        )

        aplicar_credenciais_convivente_salvar(dados)

        novo_convivente = ConviventeDB(instituicao_id=obter_instituicao_escopo(usuario_atual), **dados)
        maior_numero = (await db.execute(select(func.max(ConviventeDB.numero_institucional)).where(ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual)))).scalar()
        novo_convivente.numero_institucional = (maior_numero or 0) + 1
        db.add(novo_convivente)
        await db.flush() 

        if novo_convivente.leito_id and novo_convivente.status == "Ativo":
            leito = (await db.execute(select(LeitoDB).where(LeitoDB.id == novo_convivente.leito_id))).scalar_one_or_none()
            if leito: leito.status = "Ocupado"

        if novo_convivente.status != "Ativo" and obs_status:
            db.add(OcorrenciaConviventeDB(
                instituicao_id=obter_instituicao_escopo(usuario_atual), convivente_id=novo_convivente.id,
                usuario_criador_id=usuario_atual["sub"], tecnico_responsavel_id=novo_convivente.tecnico_id,
                tipo_ocorrencia="Mudança de Status Institucional", motivo=f"Registro Inicial: {novo_convivente.status}",
                descricao=obs_status, requer_acao_tecnica=False, status_resolucao="Resolvido",
                parecer_tecnico="Ação registrada diretamente na criação da ficha."
            ))

        await db.commit()
        await db.refresh(novo_convivente)
        registrar_evento_auditoria(
            "convivente_criado",
            usuario_atual=usuario_atual,
            convivente_id=novo_convivente.id,
            status=novo_convivente.status,
            tecnico_id=novo_convivente.tecnico_id,
        )
        return convivente_para_response(novo_convivente, usuario_atual)
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Não foi possível criar o convivente. Verifique os dados informados.",
        ) from e

@router.get("/conviventes", response_model=List[ConviventeResponse])
async def listar_conviventes(db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    rows = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalars().all()

    return [convivente_para_response(conv, usuario_atual) for conv in rows]


@router.get("/ausencias-justificadas/pendencias", response_model=List[AusenciaJustificadaPendenteResponse])
async def listar_ausencias_justificadas_pendentes(
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    bloquear_usuario_global_puro(usuario_atual)
    hoje = agora_sao_paulo().date()

    confirmados_hoje = (
        select(AusenciaJustificadaConfirmacaoDB.convivente_id)
        .where(
            AusenciaJustificadaConfirmacaoDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            AusenciaJustificadaConfirmacaoDB.data_referencia == hoje,
        )
    )

    resultado = await db.execute(
        select(ConviventeDB, UsuarioDB.nome.label("tecnico_nome"))
        .outerjoin(UsuarioDB, UsuarioDB.id == ConviventeDB.tecnico_id)
        .where(
            ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            ConviventeDB.status == "Ausência justificada",
            ~ConviventeDB.id.in_(confirmados_hoje),
            or_(
                ConviventeDB.ausencia_justificada_desde.is_(None),
                ConviventeDB.ausencia_justificada_desde < hoje,
            ),
        )
        .order_by(ConviventeDB.nome_completo.asc())
    )

    pendencias = []
    for convivente, tecnico_nome in resultado.all():
        desde = convivente.ausencia_justificada_desde or hoje
        pendencias.append({
            "convivente_id": convivente.id,
            "nome": convivente.nome_social or convivente.nome_completo,
            "prontuario": convivente.numero_institucional,
            "tecnico_id": convivente.tecnico_id,
            "tecnico_nome": tecnico_nome,
            "ausencia_justificada_desde": convivente.ausencia_justificada_desde,
            "dias_em_ausencia": max((hoje - desde).days + 1, 1),
        })

    return pendencias


@router.post("/ausencias-justificadas/{convivente_id}/responder")
async def responder_ausencia_justificada(
    convivente_id: str,
    payload: AusenciaJustificadaResposta,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    bloquear_usuario_global_puro(usuario_atual)
    hoje = agora_sao_paulo().date()

    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                ConviventeDB.status == "Ausência justificada",
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente em ausência justificada não encontrado.")

    status_atribuido = payload.status_atribuido
    justificativa = (payload.justificativa or "").strip()

    if not payload.continua_ausente:
        if status_atribuido not in {"Ativo", "Inativado"}:
            raise HTTPException(status_code=400, detail="Escolha Ativo ou Inativado para encerrar a ausência justificada.")

        if not usuario_pode_alterar_status_convivente(usuario_atual, convivente):
            raise HTTPException(
                status_code=403,
                detail=(
                    "Apenas Gestores, o Técnico responsável ou um Técnico quando "
                    "o convivente estiver sem técnico atrelado podem encerrar a ausência justificada."
                ),
            )

        if status_atribuido == "Inativado" and not justificativa:
            raise HTTPException(status_code=400, detail="Informe a justificativa para inativar o convivente.")

    confirmacao = (
        await db.execute(
            select(AusenciaJustificadaConfirmacaoDB).where(
                AusenciaJustificadaConfirmacaoDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                AusenciaJustificadaConfirmacaoDB.convivente_id == convivente_id,
                AusenciaJustificadaConfirmacaoDB.data_referencia == hoje,
            )
        )
    ).scalar_one_or_none()

    if not confirmacao:
        confirmacao = AusenciaJustificadaConfirmacaoDB(
            instituicao_id=obter_instituicao_escopo(usuario_atual),
            convivente_id=convivente_id,
            usuario_id=usuario_atual["sub"],
            data_referencia=hoje,
            continua_ausente=payload.continua_ausente,
            status_atribuido=status_atribuido,
            justificativa=justificativa or None,
            respondido_em=agora_sao_paulo(),
        )
        db.add(confirmacao)
    else:
        confirmacao.usuario_id = usuario_atual["sub"]
        confirmacao.continua_ausente = payload.continua_ausente
        confirmacao.status_atribuido = status_atribuido
        confirmacao.justificativa = justificativa or None
        confirmacao.respondido_em = agora_sao_paulo()

    if not payload.continua_ausente:
        status_antigo = convivente.status
        convivente.status = status_atribuido
        convivente.ausencia_justificada_desde = None

        if status_atribuido == "Ativo":
            convivente.inativado_em = None
        else:
            leito_id_antigo = convivente.leito_id
            convivente.leito_id = None
            convivente.inativado_em = agora_sao_paulo()

            if leito_id_antigo:
                leito = (
                    await db.execute(select(LeitoDB).where(LeitoDB.id == leito_id_antigo))
                ).scalar_one_or_none()
                if leito:
                    leito.status = "Livre"

        db.add(OcorrenciaConviventeDB(
            instituicao_id=obter_instituicao_escopo(usuario_atual),
            convivente_id=convivente.id,
            usuario_criador_id=usuario_atual["sub"],
            tecnico_responsavel_id=convivente.tecnico_id,
            tipo_ocorrencia="Mudança de Status Institucional",
            motivo=f"Alerta de ausência: alterado de {status_antigo} para {status_atribuido}",
            descricao=justificativa or "Ausência justificada encerrada pelo alerta diário.",
            requer_acao_tecnica=False,
            status_resolucao="Resolvido",
            parecer_tecnico="Ação registrada pelo alerta diário de ausência justificada.",
        ))

    await db.commit()

    return {"status": "sucesso"}


@router.get("/conviventes/resumo")
async def listar_conviventes_resumo(
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    rows = (
        await db.execute(
            select(
                ConviventeDB.id,
                ConviventeDB.numero_institucional,
                ConviventeDB.status,
                ConviventeDB.nome_completo,
                ConviventeDB.nome_social,
                ConviventeDB.cpf,
                ConviventeDB.leito_id,
                ConviventeDB.tecnico_id,
                ConviventeDB.foto_url,
            ).where(
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            ).order_by(
                ConviventeDB.status.asc(),
                ConviventeDB.nome_completo.asc(),
            )
        )
    ).all()

    return [
        {
            "id": row.id,
            "numero_institucional": row.numero_institucional,
            "status": row.status,
            "nome_completo": row.nome_completo,
            "nome_social": row.nome_social,
            "cpf": row.cpf,
            "leito_id": row.leito_id,
            "tecnico_id": row.tecnico_id,
            "foto_url": row.foto_url,
        }
        for row in rows
    ]

@router.get("/conviventes/{convivente_id}", response_model=ConviventeResponse)
async def obtener_convivente(convivente_id: str, db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    convivente = (await db.execute(select(ConviventeDB).where(ConviventeDB.id == convivente_id, ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual)))).scalar_one_or_none()
    if not convivente: raise HTTPException(status_code=404, detail="Convivente não encontrado.")
    return convivente_para_response(convivente, usuario_atual)


@router.get("/conviventes/{convivente_id}/pia", response_model=RegistroPIAListaResponse)
async def listar_registros_pia(
    convivente_id: str,
    limite: int = Query(REGISTROS_POR_PAGINA_PADRAO, ge=1, le=200),
    deslocamento: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    instituicao_id = obter_instituicao_escopo(usuario_atual)
    base = select(RegistroPIADB).where(
        RegistroPIADB.convivente_id == convivente_id,
        RegistroPIADB.instituicao_id == instituicao_id,
    )
    total = int((await db.execute(
        select(func.count()).select_from(base.subquery())
    )).scalar_one() or 0)

    registros = (
        await db.execute(
            select(RegistroPIADB, UsuarioDB.nome.label("usuario_nome"))
            .join(UsuarioDB, UsuarioDB.id == RegistroPIADB.usuario_id)
            .where(
                RegistroPIADB.convivente_id == convivente_id,
                RegistroPIADB.instituicao_id == instituicao_id,
            )
            .order_by(RegistroPIADB.data_registro.desc())
            .offset(deslocamento)
            .limit(limite)
        )
    ).all()

    items = [
        {
            **{coluna.name: getattr(registro, coluna.name) for coluna in registro.__table__.columns},
            "usuario_nome": usuario_nome,
        }
        for registro, usuario_nome in registros
    ]

    return {
        "registros": items,
        "total": total,
        "limite": limite,
        "deslocamento": deslocamento,
        "has_more": deslocamento + len(items) < total,
    }


@router.post("/conviventes/{convivente_id}/pia", response_model=RegistroPIAResponse)
async def criar_registro_pia(
    convivente_id: str,
    payload: RegistroPIACreate,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)

    if not usuario_pode_gerenciar_pia_convivente(usuario_atual):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Apenas Gestores, Técnicos ou Manutenção podem criar ou evoluir PIA.",
        )

    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    exigir_convivente_ativo_para_registro(convivente)

    if not payload.titulo.strip() or not payload.descricao.strip():
        raise HTTPException(status_code=400, detail="Informe título e descrição do registro PIA.")

    registro_pai_id = payload.registro_pai_id or None
    if registro_pai_id:
        registro_pai = (
            await db.execute(
                select(RegistroPIADB).where(
                    RegistroPIADB.id == registro_pai_id,
                    RegistroPIADB.convivente_id == convivente_id,
                    RegistroPIADB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                )
            )
        ).scalar_one_or_none()

        if not registro_pai:
            raise HTTPException(status_code=404, detail="Registro PIA principal não encontrado.")

        if registro_pai.registro_pai_id:
            raise HTTPException(status_code=400, detail="Evolua sempre o registro principal do PIA.")

        if not (payload.subtitulo or "").strip():
            raise HTTPException(status_code=400, detail="Informe o subtítulo/tema da evolução.")

    registro = RegistroPIADB(
        instituicao_id=obter_instituicao_escopo(usuario_atual),
        convivente_id=convivente_id,
        usuario_id=usuario_atual["sub"],
        registro_pai_id=registro_pai_id,
        tipo_registro="Evolução" if registro_pai_id else (payload.tipo_registro.strip() or "PIA"),
        titulo="Evolução" if registro_pai_id else payload.titulo.strip(),
        subtitulo=(payload.subtitulo or "").strip() or None,
        descricao=payload.descricao.strip(),
        objetivos=(payload.objetivos or "").strip() or None,
        encaminhamentos=(payload.encaminhamentos or "").strip() or None,
        status=payload.status.strip() or "Em acompanhamento",
        data_registro=agora_sao_paulo(),
    )

    db.add(registro)
    await db.commit()
    await db.refresh(registro)

    return {
        **{coluna.name: getattr(registro, coluna.name) for coluna in registro.__table__.columns},
        "usuario_nome": usuario_atual.get("nome"),
    }


def usuario_pode_criar_historico_convivente(usuario_atual: dict) -> bool:
    return usuario_eh_gestor(usuario_atual) or usuario_tem_perfil(usuario_atual, {PERFIL_TECNICO})


def usuario_pode_gerenciar_pia_convivente(usuario_atual: dict) -> bool:
    return (
        usuario_eh_gestor(usuario_atual)
        or usuario_eh_manutencao(usuario_atual)
        or usuario_tem_perfil(usuario_atual, {PERFIL_TECNICO})
    )


def usuario_pode_editar_historico_convivente(usuario_atual: dict, convivente: ConviventeDB) -> bool:
    return bool(
        usuario_eh_gestor(usuario_atual)
        or (
            usuario_tem_perfil(usuario_atual, {PERFIL_TECNICO})
            and str(convivente.tecnico_id) == str(usuario_atual.get("sub"))
        )
    )


def usuario_pode_alterar_status_convivente(usuario_atual: dict, convivente: ConviventeDB) -> bool:
    if usuario_eh_gestor(usuario_atual):
        return True

    if not usuario_tem_perfil(usuario_atual, {PERFIL_TECNICO}):
        return False

    tecnico_id = getattr(convivente, "tecnico_id", None)
    return not tecnico_id or str(tecnico_id) == str(usuario_atual.get("sub"))


def usuario_pode_excluir_convivente_sem_vinculos(usuario_atual: dict) -> bool:
    return usuario_eh_gestor(usuario_atual) or usuario_eh_manutencao(usuario_atual)


def usuario_pode_excluir_importacao_sisa(usuario_atual: dict) -> bool:
    return usuario_eh_gestor(usuario_atual) or usuario_eh_manutencao(usuario_atual)


def usuario_visivel_como_equipe(usuario: UsuarioDB | dict | None) -> bool:
    if not usuario:
        return False

    perfil = getattr(usuario, "perfil_acesso", None)
    is_global = getattr(usuario, "is_global", False)
    if isinstance(usuario, dict):
        perfil = usuario.get("perfil_acesso")
        is_global = usuario.get("is_global", False)

    perfil_normalizado = normalizar_perfil_acesso(str(perfil or "").strip())
    perfil_chave = perfil_normalizado.lower()
    return (
        not bool(is_global)
        and perfil_normalizado not in {PERFIL_GLOBAL, PERFIL_MANUTENCAO}
        and perfil_chave not in {"global", "executivo", "manutenção", "manutencao"}
    )


def convivente_esta_ativo(convivente: ConviventeDB) -> bool:
    return getattr(convivente, "status", None) == "Ativo"


def exigir_convivente_ativo_para_registro(convivente: ConviventeDB) -> None:
    if convivente_esta_ativo(convivente):
        return

    raise HTTPException(
        status_code=status.HTTP_409_CONFLICT,
        detail=(
            "Convivente inativo. Ative o convivente antes de registrar "
            "movimentações, ocorrências, interações ou novos registros."
        ),
    )


async def listar_bloqueios_exclusao_convivente(
    db: AsyncSession,
    convivente_id: str,
) -> list[str]:
    bloqueios = []

    for modelo, rotulo in DEPENDENCIAS_EXCLUSAO_CONVIVENTE:
        total = (
            await db.execute(
                select(func.count(modelo.id)).where(modelo.convivente_id == convivente_id)
            )
        ).scalar_one()
        if total:
            bloqueios.append(rotulo)

    return bloqueios


@router.get("/conviventes/{convivente_id}/historicos", response_model=HistoricoConviventeListaResponse)
async def listar_historicos_convivente(
    convivente_id: str,
    data_inicio: str = None,
    data_fim: str = None,
    origem_informacao: str = None,
    busca: str = None,
    limite: int = 30,
    deslocamento: int = 0,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    instituicao_id = obter_instituicao_escopo(usuario_atual)
    condicoes = [
        HistoricoConviventeDB.instituicao_id == instituicao_id,
        HistoricoConviventeDB.convivente_id == convivente_id,
        HistoricoConviventeDB.origem_informacao != ORIGEM_HISTORICO_ROTINA_OPERACIONAL,
    ]

    if data_inicio:
        try:
            inicio = date.fromisoformat(data_inicio)
            condicoes.append(HistoricoConviventeDB.data_origem >= inicio)
        except ValueError:
            raise HTTPException(status_code=400, detail="Data inicial inválida.")

    if data_fim:
        try:
            fim = date.fromisoformat(data_fim)
            condicoes.append(HistoricoConviventeDB.data_origem <= fim)
        except ValueError:
            raise HTTPException(status_code=400, detail="Data final inválida.")

    if origem_informacao:
        termo_origem = f"%{origem_informacao.strip()}%"
        condicoes.append(HistoricoConviventeDB.origem_informacao.ilike(termo_origem))

    if busca:
        termo = f"%{busca.strip()}%"
        condicoes.append(
            or_(
                HistoricoConviventeDB.titulo.ilike(termo),
                HistoricoConviventeDB.descricao.ilike(termo),
                HistoricoConviventeDB.origem_informacao.ilike(termo),
            )
        )

    limite_seguro = min(max(int(limite or 30), 1), 200)
    deslocamento_seguro = max(int(deslocamento or 0), 0)

    total = (
        await db.execute(
            select(func.count(HistoricoConviventeDB.id))
            .select_from(HistoricoConviventeDB)
            .where(*condicoes)
        )
    ).scalar_one()

    registros = (
        await db.execute(
            select(HistoricoConviventeDB, UsuarioDB.nome.label("usuario_nome"))
            .join(UsuarioDB, UsuarioDB.id == HistoricoConviventeDB.usuario_id)
            .where(*condicoes)
            .order_by(
                HistoricoConviventeDB.data_origem.desc(),
                HistoricoConviventeDB.criado_em.desc(),
            )
            .offset(deslocamento_seguro)
            .limit(limite_seguro)
        )
    ).all()

    return {
        "registros": [
            {
                **{coluna.name: getattr(registro, coluna.name) for coluna in registro.__table__.columns},
                "usuario_nome": usuario_nome,
            }
            for registro, usuario_nome in registros
        ],
        "total": int(total or 0),
        "limite": limite_seguro,
        "deslocamento": deslocamento_seguro,
    }


@router.post("/conviventes/{convivente_id}/historicos", response_model=HistoricoConviventeResponse)
async def criar_historico_convivente(
    convivente_id: str,
    payload: HistoricoConviventeCreate,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)

    if not usuario_pode_criar_historico_convivente(usuario_atual):
        raise HTTPException(
            status_code=403,
            detail="Apenas Gestores e Técnicos podem inserir histórico no convivente."
        )

    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    exigir_convivente_ativo_para_registro(convivente)

    historico_legado_id = payload.historico_legado_id or None
    if historico_legado_id:
        legado = (
            await db.execute(
                select(HistoricoLegadoSIATDB).where(
                    HistoricoLegadoSIATDB.id == historico_legado_id,
                    HistoricoLegadoSIATDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                )
            )
        ).scalar_one_or_none()
        if not legado:
            raise HTTPException(status_code=404, detail="Registro legado não encontrado neste projeto.")

    registro = HistoricoConviventeDB(
        instituicao_id=obter_instituicao_escopo(usuario_atual),
        convivente_id=convivente_id,
        usuario_id=usuario_atual["sub"],
        historico_legado_id=historico_legado_id,
        origem_informacao=payload.origem_informacao.strip(),
        data_origem=payload.data_origem,
        titulo=(payload.titulo or "").strip() or None,
        descricao=payload.descricao.strip(),
        criado_em=agora_sao_paulo(),
    )

    db.add(registro)
    await db.commit()
    await db.refresh(registro)

    return {
        **{coluna.name: getattr(registro, coluna.name) for coluna in registro.__table__.columns},
        "usuario_nome": usuario_atual.get("nome"),
    }


@router.put("/conviventes/{convivente_id}/historicos/{historico_id}", response_model=HistoricoConviventeResponse)
async def atualizar_historico_convivente(
    convivente_id: str,
    historico_id: str,
    payload: HistoricoConviventeUpdate,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)

    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    if not usuario_pode_editar_historico_convivente(usuario_atual, convivente):
        raise HTTPException(
            status_code=403,
            detail="Apenas Gestores ou o Técnico responsável podem editar este histórico."
        )

    registro = (
        await db.execute(
            select(HistoricoConviventeDB).where(
                HistoricoConviventeDB.id == historico_id,
                HistoricoConviventeDB.convivente_id == convivente_id,
                HistoricoConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not registro:
        raise HTTPException(status_code=404, detail="Histórico não encontrado.")

    if registro.origem_informacao == ORIGEM_HISTORICO_ROTINA_OPERACIONAL:
        raise HTTPException(
            status_code=403,
            detail="Registros de rotina operacional são somente leitura. Consulte a aba Fluxo Diário.",
        )

    registro.origem_informacao = payload.origem_informacao.strip()
    registro.data_origem = payload.data_origem
    registro.titulo = (payload.titulo or "").strip() or None
    registro.descricao = payload.descricao.strip()

    await db.commit()
    await db.refresh(registro)

    return {
        **{coluna.name: getattr(registro, coluna.name) for coluna in registro.__table__.columns},
        "usuario_nome": usuario_atual.get("nome"),
    }


@router.delete("/conviventes/{convivente_id}/historicos/{historico_id}")
async def excluir_historico_convivente(
    convivente_id: str,
    historico_id: str,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)

    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    if not usuario_pode_editar_historico_convivente(usuario_atual, convivente):
        raise HTTPException(
            status_code=403,
            detail="Apenas Gestores ou o Técnico responsável podem excluir este histórico."
        )

    registro = (
        await db.execute(
            select(HistoricoConviventeDB).where(
                HistoricoConviventeDB.id == historico_id,
                HistoricoConviventeDB.convivente_id == convivente_id,
                HistoricoConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not registro:
        raise HTTPException(status_code=404, detail="Histórico não encontrado.")

    if registro.origem_informacao == ORIGEM_HISTORICO_ROTINA_OPERACIONAL:
        raise HTTPException(
            status_code=403,
            detail="Registros de rotina operacional são somente leitura. Consulte a aba Fluxo Diário.",
        )

    await db.delete(registro)
    await db.commit()

    return {"status": "sucesso"}

@router.put("/conviventes/{convivente_id}", response_model=ConviventeResponse)
async def atualizar_convivente(convivente_id: str, dados_atualizacao: ConviventeUpdate, db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    bloquear_usuario_global_puro(usuario_atual)
    convivente = (await db.execute(select(ConviventeDB).where(ConviventeDB.id == convivente_id, ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual)))).scalar_one_or_none()
    if not convivente: raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    leito_id_antigo = convivente.leito_id
    status_antigo = convivente.status
    obs_status = getattr(dados_atualizacao, "observacao_status", None)
    dados = dados_atualizacao.model_dump(exclude_unset=True, exclude={"observacao_status"})

    if (
        "leito_id" in dados
        and str(dados.get("leito_id") or "") != str(leito_id_antigo or "")
        and usuario_tem_perfil(usuario_atual, {PERFIL_ORIENTADOR})
    ):
        raise HTTPException(
            status_code=403,
            detail="Orientadores devem alterar a cama do convivente pelo módulo Acomodações.",
        )

    if status_antigo != dados.get("status", status_antigo):
        if not usuario_pode_alterar_status_convivente(usuario_atual, convivente):
            raise HTTPException(
                status_code=403,
                detail=(
                    "Operação negada. "
                    "Apenas Gestores, o Técnico responsável ou um Técnico quando "
                    "o convivente estiver sem técnico atrelado podem alterar o status."
                ),
            )

    if dados.get("status") in ["Inativado", "Bloqueado", "Saída qualificada"]:
        dados["leito_id"] = None

    novo_status = dados.get("status", status_antigo)
    if novo_status == "Ativo":
        dados["inativado_em"] = None
        dados["ausencia_justificada_desde"] = None
    elif novo_status == "Ausência justificada":
        dados["inativado_em"] = None
        if status_antigo != "Ausência justificada" or not convivente.ausencia_justificada_desde:
            dados["ausencia_justificada_desde"] = agora_sao_paulo().date()
    elif novo_status != status_antigo and novo_status in ["Inativado", "Bloqueado", "Saída qualificada"]:
        dados["inativado_em"] = agora_sao_paulo()
        dados["ausencia_justificada_desde"] = None

    await validar_referencias_convivente_do_projeto(
        db,
        dados,
        obter_instituicao_escopo(usuario_atual),
    )

    aplicar_credenciais_convivente_salvar(dados)

    for key, value in dados.items(): setattr(convivente, key, value)

    try:
        if leito_id_antigo != convivente.leito_id:
            if leito_id_antigo:
                l_antigo = (await db.execute(select(LeitoDB).where(LeitoDB.id == leito_id_antigo))).scalar_one_or_none()
                if l_antigo: l_antigo.status = "Livre"
            if convivente.leito_id:
                l_novo = (await db.execute(select(LeitoDB).where(LeitoDB.id == convivente.leito_id))).scalar_one_or_none()
                if l_novo: l_novo.status = "Ocupado"

        if status_antigo != convivente.status and obs_status:
            db.add(OcorrenciaConviventeDB(
                instituicao_id=obter_instituicao_escopo(usuario_atual), convivente_id=convivente.id,
                usuario_criador_id=usuario_atual["sub"], tecnico_responsavel_id=convivente.tecnico_id,
                tipo_ocorrencia="Mudança de Status Institucional",
                motivo=f"Alterado de {status_antigo} para {convivente.status}",
                descricao=obs_status, requer_acao_tecnica=False, status_resolucao="Resolvido",
                parecer_tecnico="Ação e parecer registrados diretamente pelo Técnico Responsável/Gerência."
            ))

        await db.commit()
        await db.refresh(convivente)
        registrar_evento_auditoria(
            "convivente_editado",
            usuario_atual=usuario_atual,
            convivente_id=convivente.id,
            status_anterior=status_antigo,
            status_atual=convivente.status,
            alterou_status=status_antigo != convivente.status,
            alterou_leito=leito_id_antigo != convivente.leito_id,
        )
        return convivente_para_response(convivente, usuario_atual)
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Não foi possível atualizar o convivente.",
        ) from e


@router.delete("/conviventes/{convivente_id}")
async def excluir_convivente_sem_vinculos(
    convivente_id: str,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    bloquear_usuario_global_puro(usuario_atual)

    if not usuario_pode_excluir_convivente_sem_vinculos(usuario_atual):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Apenas Gestores ou Manutenção podem excluir cadastro sem vínculos.",
        )

    instituicao_id = obter_instituicao_escopo(usuario_atual)
    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == instituicao_id,
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Convivente não encontrado.")

    bloqueios = await listar_bloqueios_exclusao_convivente(db, convivente_id)
    if bloqueios:
        bloqueios_texto = ", ".join(bloqueios[:8])
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Este cadastro já possui vínculos operacionais e não pode ser excluído. "
                f"Vínculos encontrados: {bloqueios_texto}. "
                "Use inativação/alteração de status para preservar histórico assistencial."
            ),
        )

    leito_id = convivente.leito_id
    try:
        if leito_id:
            leito = (
                await db.execute(select(LeitoDB).where(LeitoDB.id == leito_id))
            ).scalar_one_or_none()
            if leito:
                leito.status = "Livre"

        await db.delete(convivente)
        await db.commit()
        registrar_evento_auditoria(
            "convivente_excluido_sem_vinculos",
            usuario_atual=usuario_atual,
            convivente_id=convivente_id,
            leito_liberado=bool(leito_id),
        )
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Não foi possível excluir o convivente.",
        ) from e

    return {"status": "sucesso", "mensagem": "Cadastro sem vínculos excluído com sucesso."}

# =====================================================================
# ROTAS DO SISTEMA DE TICKETS / OCORRÊNCIAS
# =====================================================================

@router.get("/ocorrencias")
async def listar_todas_ocorrencias(
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
    limit: int = Query(REGISTROS_POR_PAGINA_PADRAO, ge=1, le=200),
    offset: int = Query(0, ge=0),
    prioridade: Optional[str] = Query(None),
    status_filtro: Optional[str] = Query(None, alias="status"),
    tecnico_id: Optional[str] = Query(None),
    data_inicio: Optional[date] = Query(None),
    data_fim: Optional[date] = Query(None),
    busca: Optional[str] = Query(None),
    status_convivente: Optional[str] = Query(None),
):
    perfil = usuario_atual.get("perfil_acesso")
    user_id = usuario_atual.get("sub")
    inst_id = obter_instituicao_escopo(usuario_atual)

    query = select(OcorrenciaConviventeDB).where(OcorrenciaConviventeDB.instituicao_id == inst_id)

    if perfil == "Orientador":
        subq = select(ObservadorOcorrenciaDB.ocorrencia_id).where(ObservadorOcorrenciaDB.usuario_id == user_id)
        query = query.where(
            or_(
                OcorrenciaConviventeDB.usuario_criador_id == user_id,
                OcorrenciaConviventeDB.id.in_(subq)
            )
        )

    if prioridade and prioridade != "Todas":
        query = query.where(OcorrenciaConviventeDB.prioridade == normalizar_prioridade_ocorrencia(prioridade))

    if status_filtro and status_filtro != "Todos":
        if status_filtro == "Resolvido":
            query = query.where(OcorrenciaConviventeDB.status_resolucao == "Resolvido")
        elif status_filtro == "Pendente":
            query = query.where(OcorrenciaConviventeDB.status_resolucao != "Resolvido")

    if tecnico_id:
        query = query.where(OcorrenciaConviventeDB.tecnico_responsavel_id == tecnico_id)

    if status_convivente and status_convivente != "Todos":
        subq_conviventes_status = select(ConviventeDB.id).where(
            ConviventeDB.instituicao_id == inst_id,
            ConviventeDB.status == status_convivente,
        )
        query = query.where(OcorrenciaConviventeDB.convivente_id.in_(subq_conviventes_status))

    if data_inicio:
        query = query.where(OcorrenciaConviventeDB.data_ocorrencia >= data_inicio)

    if data_fim:
        query = query.where(OcorrenciaConviventeDB.data_ocorrencia < data_fim + timedelta(days=1))

    if busca and busca.strip():
        termo_busca = f"%{busca.strip()}%"
        subq_conviventes_busca = select(ConviventeDB.id).where(
            ConviventeDB.instituicao_id == inst_id,
            or_(
                ConviventeDB.nome_completo.ilike(termo_busca),
                ConviventeDB.nome_social.ilike(termo_busca),
                ConviventeDB.cpf.ilike(termo_busca),
                cast(ConviventeDB.numero_institucional, String).ilike(termo_busca),
            ),
        )
        query = query.where(
            or_(
                OcorrenciaConviventeDB.tipo_ocorrencia.ilike(termo_busca),
                OcorrenciaConviventeDB.motivo.ilike(termo_busca),
                OcorrenciaConviventeDB.descricao.ilike(termo_busca),
                OcorrenciaConviventeDB.parecer_tecnico.ilike(termo_busca),
                OcorrenciaConviventeDB.convivente_id.in_(subq_conviventes_busca),
            )
        )

    total = (
        await db.execute(
            select(func.count()).select_from(query.order_by(None).subquery())
        )
    ).scalar_one()

    def contar_total(expressao):
        return select(func.count()).select_from(
            query.order_by(None).where(expressao).subquery()
        )

    pendentes = (await db.execute(contar_total(OcorrenciaConviventeDB.status_resolucao != "Resolvido"))).scalar_one()
    resolvidas = (await db.execute(contar_total(OcorrenciaConviventeDB.status_resolucao == "Resolvido"))).scalar_one()
    alta_critica_pendentes = (
        await db.execute(
            contar_total(
                and_(
                    OcorrenciaConviventeDB.status_resolucao != "Resolvido",
                    OcorrenciaConviventeDB.prioridade.in_(["Alta", "Crítica"]),
                )
            )
        )
    ).scalar_one()

    por_prioridade = {}
    for prioridade_nome in PRIORIDADES_OCORRENCIA:
        total_prioridade = (
            await db.execute(contar_total(OcorrenciaConviventeDB.prioridade == prioridade_nome))
        ).scalar_one()
        pendentes_prioridade = (
            await db.execute(
                contar_total(
                    and_(
                        OcorrenciaConviventeDB.prioridade == prioridade_nome,
                        OcorrenciaConviventeDB.status_resolucao != "Resolvido",
                    )
                )
            )
        ).scalar_one()
        por_prioridade[prioridade_nome] = {
            "total": total_prioridade,
            "pendentes": pendentes_prioridade,
        }

    query = query.order_by(
        case((OcorrenciaConviventeDB.status_resolucao == "Resolvido", 1), else_=0).asc(),
        case((OcorrenciaConviventeDB.tecnico_responsavel_id == user_id, 0), else_=1).asc(),
        case(
            (OcorrenciaConviventeDB.prioridade == "Crítica", 4),
            (OcorrenciaConviventeDB.prioridade == "Alta", 3),
            (OcorrenciaConviventeDB.prioridade == "Média", 2),
            else_=1,
        ).desc(),
        OcorrenciaConviventeDB.data_ocorrencia.desc(),
    )

    if limit is not None:
        query = query.offset(offset).limit(limit)

    ocorrencias_db = (await db.execute(query)).scalars().all()

    ocorrencia_ids = [oc.id for oc in ocorrencias_db]
    convivente_ids = [oc.convivente_id for oc in ocorrencias_db if oc.convivente_id]
    interacoes_por_ocorrencia = {oc_id: [] for oc_id in ocorrencia_ids}
    observadores_por_ocorrencia = {oc_id: [] for oc_id in ocorrencia_ids}
    conviventes_por_id = {}

    if convivente_ids:
        conviventes_db = (
            await db.execute(
                select(ConviventeDB).where(ConviventeDB.id.in_(convivente_ids))
            )
        ).scalars().all()
        conviventes_por_id = {conv.id: conv for conv in conviventes_db}

    if ocorrencia_ids:
        interacoes_db = (
            await db.execute(
                select(InteracaoOcorrenciaDB)
                .where(InteracaoOcorrenciaDB.ocorrencia_id.in_(ocorrencia_ids))
                .order_by(InteracaoOcorrenciaDB.data_interacao.asc())
            )
        ).scalars().all()
        for interacao in interacoes_db:
            interacoes_por_ocorrencia.setdefault(interacao.ocorrencia_id, []).append(interacao)

        observadores_db = (
            await db.execute(
                select(ObservadorOcorrenciaDB)
                .where(ObservadorOcorrenciaDB.ocorrencia_id.in_(ocorrencia_ids))
            )
        ).scalars().all()
        for observador in observadores_db:
            observadores_por_ocorrencia.setdefault(observador.ocorrencia_id, []).append(observador)

    resultado = []
    for oc in ocorrencias_db:
        oc_dict = {c.name: getattr(oc, c.name) for c in oc.__table__.columns}
        oc_dict["prioridade"] = normalizar_prioridade_ocorrencia(oc_dict.get("prioridade"))
        oc_dict["interacoes"] = [
            {c.name: getattr(interacao, c.name) for c in interacao.__table__.columns}
            for interacao in interacoes_por_ocorrencia.get(oc.id, [])
        ]
        oc_dict["observadores"] = [
            {c.name: getattr(observador, c.name) for c in observador.__table__.columns}
            for observador in observadores_por_ocorrencia.get(oc.id, [])
        ]
        convivente = conviventes_por_id.get(oc.convivente_id)
        if convivente:
            oc_dict["convivente_nome"] = convivente.nome_social or convivente.nome_completo
            oc_dict["convivente_nome_completo"] = convivente.nome_completo
            oc_dict["convivente_nome_social"] = convivente.nome_social
            oc_dict["convivente_numero_institucional"] = convivente.numero_institucional
            oc_dict["convivente_cpf"] = convivente.cpf
        resultado.append(oc_dict)

    return {
        "items": resultado,
        "total": total,
        "limit": limit,
        "offset": offset,
        "has_more": offset + len(resultado) < total,
        "resumo": {
            "total": total,
            "pendentes": pendentes,
            "resolvidas": resolvidas,
            "altaCriticaPendentes": alta_critica_pendentes,
            "porPrioridade": por_prioridade,
        },
    }


@router.get("/ocorrencias/relatorio-prioridades", response_model=OcorrenciaRelatorioPrioridades)
async def relatorio_prioridades_ocorrencias(db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    inst_id = obter_instituicao_escopo(usuario_atual)
    base = select(OcorrenciaConviventeDB).where(OcorrenciaConviventeDB.instituicao_id == inst_id)

    def contar(expressao=None):
        consulta = base
        if expressao is not None:
            consulta = consulta.where(expressao)
        return select(func.count()).select_from(consulta.order_by(None).subquery())

    total = int((await db.execute(contar())).scalar_one() or 0)
    pendentes = int((await db.execute(
        contar(OcorrenciaConviventeDB.status_resolucao != "Resolvido")
    )).scalar_one() or 0)
    resolvidas = int((await db.execute(
        contar(OcorrenciaConviventeDB.status_resolucao == "Resolvido")
    )).scalar_one() or 0)

    resumo = {}
    for prioridade in PRIORIDADES_OCORRENCIA:
        total_prioridade = int((await db.execute(
            contar(OcorrenciaConviventeDB.prioridade == prioridade)
        )).scalar_one() or 0)
        pendentes_prioridade = int((await db.execute(
            contar(
                and_(
                    OcorrenciaConviventeDB.prioridade == prioridade,
                    OcorrenciaConviventeDB.status_resolucao != "Resolvido",
                )
            )
        )).scalar_one() or 0)
        resolvidas_prioridade = int((await db.execute(
            contar(
                and_(
                    OcorrenciaConviventeDB.prioridade == prioridade,
                    OcorrenciaConviventeDB.status_resolucao == "Resolvido",
                )
            )
        )).scalar_one() or 0)
        resumo[prioridade] = {
            "prioridade": prioridade,
            "total": total_prioridade,
            "pendentes": pendentes_prioridade,
            "resolvidas": resolvidas_prioridade,
        }

    return {
        "total": total,
        "pendentes": pendentes,
        "resolvidas": resolvidas,
        "por_prioridade": [resumo[p] for p in PRIORIDADES_OCORRENCIA]
    }


@router.get("/ocorrencias/pendencias-tecnicas", response_model=List[OcorrenciaResponse])
async def listar_pendencias_tecnicas_priorizadas(
    limit: int = Query(REGISTROS_POR_PAGINA_PADRAO, ge=1, le=100),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    inst_id = obter_instituicao_escopo(usuario_atual)

    query = select(OcorrenciaConviventeDB).where(
        OcorrenciaConviventeDB.instituicao_id == inst_id,
        OcorrenciaConviventeDB.status_resolucao != "Resolvido",
        or_(
            OcorrenciaConviventeDB.requer_acao_tecnica == True,
            OcorrenciaConviventeDB.prioridade.in_(["Alta", "Crítica"])
        )
    ).order_by(OcorrenciaConviventeDB.data_ocorrencia.desc()).offset(offset).limit(limit)

    ocorrencias_db = (await db.execute(query)).scalars().all()

    resultado = []
    for oc in ocorrencias_db:
        q_int = select(InteracaoOcorrenciaDB).where(InteracaoOcorrenciaDB.ocorrencia_id == oc.id).order_by(InteracaoOcorrenciaDB.data_interacao.asc())
        interacoes = (await db.execute(q_int)).scalars().all()

        q_obs = select(ObservadorOcorrenciaDB).where(ObservadorOcorrenciaDB.ocorrencia_id == oc.id)
        observadores = (await db.execute(q_obs)).scalars().all()

        oc_dict = {c.name: getattr(oc, c.name) for c in oc.__table__.columns}
        oc_dict["interacoes"] = interacoes
        oc_dict["observadores"] = observadores
        resultado.append(oc_dict)

    resultado.sort(key=lambda oc: (PESO_PRIORIDADE.get(oc.get("prioridade", "Média"), 2), oc.get("data_ocorrencia")), reverse=True)
    return resultado

@router.post("/ocorrencias")
async def criar_ocorrencia_manual(payload: OcorrenciaCreate, db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    bloquear_usuario_global_puro(usuario_atual)
    prioridade = normalizar_prioridade_ocorrencia(payload.prioridade)
    requer_acao_tecnica = bool(payload.requer_acao_tecnica or prioridade in ["Alta", "Crítica"])
    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == payload.convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()
    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    exigir_convivente_ativo_para_registro(convivente)

    tecnico_responsavel_id = await validar_tecnico_responsavel_ocorrencia(
        db,
        convivente,
        payload.tecnico_responsavel_id,
        obter_instituicao_escopo(usuario_atual),
    )
    await validar_usuarios_do_projeto(
        db,
        payload.observadores_ids,
        obter_instituicao_escopo(usuario_atual),
        detail="Um ou mais observadores não pertencem ao projeto atual.",
    )

    assinatura_validada_em = None
    if payload.convivente_autor_ocorrencia:
        if not payload.funcionario_envolvido_id:
            raise HTTPException(
                status_code=400,
                detail="Informe o funcionário citado quando o convivente for o autor da ocorrência.",
            )

        funcionario = (
            await db.execute(
                select(UsuarioDB).where(
                    UsuarioDB.id == payload.funcionario_envolvido_id,
                    UsuarioDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                )
            )
        ).scalar_one_or_none()
        if not funcionario:
            raise HTTPException(status_code=404, detail="Funcionário citado não encontrado.")
        if not usuario_visivel_como_equipe(funcionario):
            raise HTTPException(
                status_code=400,
                detail="Funcionário citado não pode ser usuário Global ou Manutenção.",
            )

        def normalizar_codigo_assinatura(valor):
            return "".join(ch for ch in str(valor or "").strip().lower() if ch.isalnum())

        codigo_lido = normalizar_codigo_assinatura(payload.assinatura_convivente_codigo)
        cpf_limpo = "".join(ch for ch in str(convivente.cpf or "") if ch.isdigit())
        codigos_validos = {
            normalizar_codigo_assinatura(convivente.id),
            normalizar_codigo_assinatura(convivente.numero_institucional),
            normalizar_codigo_assinatura(cpf_limpo),
            normalizar_codigo_assinatura(str(convivente.id)[:8]),
        }
        if not codigo_lido or codigo_lido not in {codigo for codigo in codigos_validos if codigo}:
            raise HTTPException(
                status_code=400,
                detail="A assinatura digital não confere com a carteirinha do convivente selecionado.",
            )

        assinatura_validada_em = agora_sao_paulo()

    nova_oc = OcorrenciaConviventeDB(
        instituicao_id=obter_instituicao_escopo(usuario_atual),
        convivente_id=payload.convivente_id,
        usuario_criador_id=usuario_atual["sub"],
        tecnico_responsavel_id=tecnico_responsavel_id,
        convivente_autor_ocorrencia=payload.convivente_autor_ocorrencia,
        funcionario_envolvido_id=payload.funcionario_envolvido_id if payload.convivente_autor_ocorrencia else None,
        assinatura_convivente_metodo=payload.assinatura_convivente_metodo if payload.convivente_autor_ocorrencia else None,
        assinatura_convivente_codigo=str(payload.assinatura_convivente_codigo or "").strip() if payload.convivente_autor_ocorrencia else None,
        assinatura_convivente_validada_em=assinatura_validada_em,
        tipo_ocorrencia=payload.tipo_ocorrencia,
        motivo=payload.motivo,
        descricao=payload.descricao,
        requer_acao_tecnica=requer_acao_tecnica,
        prioridade=prioridade,
        status_resolucao=status_inicial_ocorrencia_manual()
    )
    db.add(nova_oc)
    await db.flush()

    for obs_id in payload.observadores_ids:
        obs = ObservadorOcorrenciaDB(ocorrencia_id=nova_oc.id, usuario_id=obs_id)
        db.add(obs)

    await db.commit()
    registrar_evento_auditoria(
        "ocorrencia_criada",
        usuario_atual=usuario_atual,
        ocorrencia_id=nova_oc.id,
        convivente_id=nova_oc.convivente_id,
        prioridade=prioridade,
        requer_acao_tecnica=requer_acao_tecnica,
        convivente_autor_ocorrencia=bool(payload.convivente_autor_ocorrencia),
    )
    return {"status": "sucesso", "id": nova_oc.id, "prioridade": prioridade}

@router.post("/ocorrencias/{ocorrencia_id}/interacoes")
async def adicionar_interacao(ocorrencia_id: str, payload: InteracaoOcorrenciaCreate, db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    bloquear_usuario_global_puro(usuario_atual)
    query = select(OcorrenciaConviventeDB).where(OcorrenciaConviventeDB.id == ocorrencia_id, OcorrenciaConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual))
    oc = (await db.execute(query)).scalar_one_or_none()
    
    if not oc: 
        raise HTTPException(status_code=404, detail="Ocorrência não encontrada.")

    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == oc.convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if convivente:
        exigir_convivente_ativo_para_registro(convivente)
    
    if payload.tipo_interacao == "Parecer Técnico":
        if not usuario_pode_resolver_ocorrencia(usuario_atual, oc):
            raise HTTPException(
                status_code=403,
                detail=(
                    "Apenas um Gestor ou o Técnico responsável podem "
                    "registrar parecer técnico e encerrar a ocorrência."
                ),
            )

        oc.status_resolucao = "Resolvido"
        oc.data_resolucao = agora_sao_paulo()
        oc.usuario_resolutor_id = usuario_atual["sub"]
        oc.parecer_tecnico = payload.mensagem 
        
    nova_int = InteracaoOcorrenciaDB(
        ocorrencia_id=oc.id,
        usuario_id=usuario_atual["sub"],
        mensagem=payload.mensagem,
        tipo_interacao=payload.tipo_interacao
    )
    db.add(nova_int)
    await db.commit()
    registrar_evento_auditoria(
        "ocorrencia_interacao_registrada",
        usuario_atual=usuario_atual,
        ocorrencia_id=oc.id,
        tipo_interacao=payload.tipo_interacao,
        status_resolucao=oc.status_resolucao,
    )
    
    return {"status": "sucesso"}

async def _serializar_ocorrencias_com_vinculos(
    db: AsyncSession,
    ocorrencias_db: list[OcorrenciaConviventeDB],
) -> list[dict]:
    if not ocorrencias_db:
        return []

    ocorrencia_ids = [oc.id for oc in ocorrencias_db]
    convivente_ids = [oc.convivente_id for oc in ocorrencias_db if oc.convivente_id]
    interacoes_por_ocorrencia = {oc_id: [] for oc_id in ocorrencia_ids}
    observadores_por_ocorrencia = {oc_id: [] for oc_id in ocorrencia_ids}
    conviventes_por_id = {}

    if convivente_ids:
        conviventes_db = (
            await db.execute(
                select(ConviventeDB).where(ConviventeDB.id.in_(convivente_ids))
            )
        ).scalars().all()
        conviventes_por_id = {conv.id: conv for conv in conviventes_db}

    interacoes_db = (
        await db.execute(
            select(InteracaoOcorrenciaDB)
            .where(InteracaoOcorrenciaDB.ocorrencia_id.in_(ocorrencia_ids))
            .order_by(InteracaoOcorrenciaDB.data_interacao.asc())
        )
    ).scalars().all()
    for interacao in interacoes_db:
        interacoes_por_ocorrencia.setdefault(interacao.ocorrencia_id, []).append(interacao)

    observadores_db = (
        await db.execute(
            select(ObservadorOcorrenciaDB)
            .where(ObservadorOcorrenciaDB.ocorrencia_id.in_(ocorrencia_ids))
        )
    ).scalars().all()
    for observador in observadores_db:
        observadores_por_ocorrencia.setdefault(observador.ocorrencia_id, []).append(observador)

    resultado = []
    for oc in ocorrencias_db:
        oc_dict = {c.name: getattr(oc, c.name) for c in oc.__table__.columns}
        oc_dict["prioridade"] = normalizar_prioridade_ocorrencia(oc_dict.get("prioridade"))
        oc_dict["interacoes"] = [
            {c.name: getattr(interacao, c.name) for c in interacao.__table__.columns}
            for interacao in interacoes_por_ocorrencia.get(oc.id, [])
        ]
        oc_dict["observadores"] = [
            {c.name: getattr(observador, c.name) for c in observador.__table__.columns}
            for observador in observadores_por_ocorrencia.get(oc.id, [])
        ]
        convivente = conviventes_por_id.get(oc.convivente_id)
        if convivente:
            oc_dict["convivente_nome"] = convivente.nome_social or convivente.nome_completo
            oc_dict["convivente_nome_completo"] = convivente.nome_completo
            oc_dict["convivente_nome_social"] = convivente.nome_social
            oc_dict["convivente_numero_institucional"] = convivente.numero_institucional
            oc_dict["convivente_cpf"] = convivente.cpf
        resultado.append(oc_dict)

    return resultado


@router.get("/conviventes/{convivente_id}/ocorrencias", response_model=OcorrenciaConviventeListaResponse)
async def listar_ocorrencias_convivente(
    convivente_id: str,
    data_inicio: Optional[date] = None,
    data_fim: Optional[date] = None,
    status_filtro: Optional[str] = Query(None, alias="status"),
    prioridade: Optional[str] = None,
    busca: Optional[str] = None,
    limite: int = 30,
    deslocamento: int = 0,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    inst_id = obter_instituicao_escopo(usuario_atual)
    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == inst_id,
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    condicoes = [
        OcorrenciaConviventeDB.convivente_id == convivente_id,
        OcorrenciaConviventeDB.instituicao_id == inst_id,
    ]

    if data_inicio:
        condicoes.append(OcorrenciaConviventeDB.data_ocorrencia >= data_inicio)

    if data_fim:
        condicoes.append(OcorrenciaConviventeDB.data_ocorrencia < data_fim + timedelta(days=1))

    if prioridade and prioridade != "Todas":
        condicoes.append(
            OcorrenciaConviventeDB.prioridade == normalizar_prioridade_ocorrencia(prioridade)
        )

    if status_filtro and status_filtro != "Todos":
        if status_filtro == "Resolvido":
            condicoes.append(OcorrenciaConviventeDB.status_resolucao == "Resolvido")
        elif status_filtro == "Pendente":
            condicoes.append(OcorrenciaConviventeDB.status_resolucao != "Resolvido")

    if busca and busca.strip():
        termo = f"%{busca.strip()}%"
        condicoes.append(
            or_(
                OcorrenciaConviventeDB.tipo_ocorrencia.ilike(termo),
                OcorrenciaConviventeDB.motivo.ilike(termo),
                OcorrenciaConviventeDB.descricao.ilike(termo),
                OcorrenciaConviventeDB.parecer_tecnico.ilike(termo),
            )
        )

    limite_seguro = min(max(int(limite or 30), 1), 100)
    deslocamento_seguro = max(int(deslocamento or 0), 0)

    total = (
        await db.execute(
            select(func.count(OcorrenciaConviventeDB.id)).where(*condicoes)
        )
    ).scalar_one()

    ocorrencias_db = (
        await db.execute(
            select(OcorrenciaConviventeDB)
            .where(*condicoes)
            .order_by(OcorrenciaConviventeDB.data_ocorrencia.desc())
            .offset(deslocamento_seguro)
            .limit(limite_seguro)
        )
    ).scalars().all()

    registros = await _serializar_ocorrencias_com_vinculos(db, ocorrencias_db)

    return {
        "registros": registros,
        "total": int(total or 0),
        "limite": limite_seguro,
        "deslocamento": deslocamento_seguro,
    }

# =====================================================================
# ROTAS DE UPLOAD DE DOCUMENTOS (GED)
# =====================================================================
@router.post("/conviventes/{convivente_id}/documentos", response_model=DocumentoResponse)
async def upload_documento(
    convivente_id: str,
    tipo_documento: str = Form(...),
    sensivel: bool = Form(False),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")
    exigir_convivente_ativo_para_registro(convivente)
    if sensivel and not (
        usuario_eh_gestor(usuario_atual)
        or usuario_tem_perfil(usuario_atual, {PERFIL_TECNICO})
    ):
        raise HTTPException(
            status_code=403,
            detail="Apenas Gestores/Gerentes/Master e Técnicos podem enviar documentos sensíveis.",
        )

    nome_original = validar_upload_documento(file)
    conteudo = await file.read()

    if len(conteudo) > TAMANHO_MAXIMO_DOCUMENTO_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Arquivo muito grande. O limite é 10 MB.",
        )

    extensao_final = os.path.splitext(nome_original)[1].lower()
    nome_arquivo_gravado = nome_original

    if eh_arquivo_imagem(nome_original, file.content_type):
        try:
            conteudo, extensao_final = padronizar_upload_imagem(
                conteudo,
                tipo_documento=tipo_documento,
            )
            nome_base = os.path.splitext(nome_original)[0] or "arquivo"
            nome_arquivo_gravado = f"{nome_base}{extensao_final}"
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Não foi possível processar a imagem enviada.",
            ) from exc

    nome_unico = f"{uuid.uuid4().hex}{extensao_final}"
    caminho_completo = os.path.join(UPLOAD_DIR, nome_unico)

    with open(caminho_completo, "wb") as buffer:
        buffer.write(conteudo)

    caminho_relativo = f"/uploads/documentos/{nome_unico}"

    if tipo_documento == "Foto de Perfil":
        await remover_documentos_foto_perfil(db, convivente_id)
        convivente.foto_url = caminho_relativo

    novo_doc = DocumentoConviventeDB(
        convivente_id=convivente_id,
        nome_arquivo=nome_arquivo_gravado,
        caminho_arquivo=caminho_relativo,
        tipo_documento=tipo_documento,
        sensivel=bool(sensivel),
    )
    db.add(novo_doc)
    await db.commit()
    await db.refresh(novo_doc)
    registrar_evento_auditoria(
        "convivente_documento_enviado",
        usuario_atual=usuario_atual,
        convivente_id=convivente_id,
        documento_id=novo_doc.id,
        tipo_documento=tipo_documento,
        sensivel=bool(sensivel),
        extensao=extensao_final,
    )
    return novo_doc

@router.get("/conviventes/{convivente_id}/documentos", response_model=DocumentoListaResponse)
async def listar_documentos(
    convivente_id: str,
    limite: int = Query(REGISTROS_POR_PAGINA_PADRAO, ge=1, le=200),
    deslocamento: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    filtros = [
        DocumentoConviventeDB.convivente_id == convivente_id,
        ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
    ]
    if not (
        usuario_eh_gestor(usuario_atual)
        or usuario_tem_perfil(usuario_atual, {PERFIL_TECNICO})
    ):
        filtros.append(
            or_(
                DocumentoConviventeDB.sensivel == False,
                DocumentoConviventeDB.sensivel.is_(None),
            )
        )

    base = (
        select(DocumentoConviventeDB)
        .join(ConviventeDB, ConviventeDB.id == DocumentoConviventeDB.convivente_id)
        .where(*filtros)
    )
    total = int((await db.execute(
        select(func.count()).select_from(base.order_by(None).subquery())
    )).scalar_one() or 0)

    documentos = (
        await db.execute(
            base.order_by(DocumentoConviventeDB.data_upload.desc())
            .offset(deslocamento)
            .limit(limite)
        )
    ).scalars().all()

    return {
        "registros": documentos,
        "total": total,
        "limite": limite,
        "deslocamento": deslocamento,
        "has_more": deslocamento + len(documentos) < total,
    }

@router.delete("/documentos/{documento_id}")
async def excluir_documento(documento_id: str, db: AsyncSession = Depends(get_db), usuario_atual: dict = Depends(get_usuario_logado)):
    bloquear_usuario_global_puro(usuario_atual)
    documento = (
        await db.execute(
            select(DocumentoConviventeDB)
            .join(ConviventeDB, ConviventeDB.id == DocumentoConviventeDB.convivente_id)
            .where(
                DocumentoConviventeDB.id == documento_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()
    if not documento: raise HTTPException(status_code=404, detail="Não encontrado.")
    if documento.sensivel and not usuario_eh_gestor(usuario_atual):
        raise HTTPException(status_code=403, detail="Apenas Gestores/Gerentes/Master podem excluir documentos sensíveis.")
    caminho_documento = caminho_absoluto_documento(documento.caminho_arquivo)
    if os.path.commonpath([UPLOAD_DIR_ABSOLUTO, caminho_documento]) == UPLOAD_DIR_ABSOLUTO and os.path.exists(caminho_documento):
        os.remove(caminho_documento)
    convivente_id = documento.convivente_id
    tipo_documento = documento.tipo_documento
    sensivel = bool(documento.sensivel)
    await db.delete(documento)
    await db.commit()
    registrar_evento_auditoria(
        "convivente_documento_excluido",
        usuario_atual=usuario_atual,
        convivente_id=convivente_id,
        documento_id=documento_id,
        tipo_documento=tipo_documento,
        sensivel=sensivel,
    )
    return {"status": "sucesso"}


@router.delete("/conviventes/{convivente_id}/foto-perfil")
async def remover_foto_perfil_convivente(
    convivente_id: str,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual)
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(status_code=404, detail="Convivente não encontrado.")

    await remover_documentos_foto_perfil(db, convivente_id)

    convivente.foto_url = None
    await db.commit()
    registrar_evento_auditoria(
        "convivente_foto_removida",
        usuario_atual=usuario_atual,
        convivente_id=convivente_id,
    )

    return {"status": "sucesso", "mensagem": "Foto de perfil removida com sucesso."}

# =====================================================================
# NOVAS ROTAS DE ROTINA DIÁRIA E FLUXO (PORTARIA)
# =====================================================================

@router.post("/rotina", response_model=RegistroRotinaResponse)
async def registar_rotina(
    payload: RegistroRotinaCreate,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    if payload.tipo_registro not in TIPOS_ROTINA_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail="Tipo de registro inválido."
        )

    observacao = (payload.observacao or "").strip()

    if payload.tipo_registro in TIPOS_ROTINA_COM_OBSERVACAO_OBRIGATORIA and not observacao:
        raise HTTPException(
            status_code=400,
            detail="Este tipo de interação exige relato/especificação.",
        )

    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == payload.convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(
            status_code=404,
            detail="Convivente não encontrado."
        )

    exigir_convivente_ativo_para_registro(convivente)

    hoje = agora_sao_paulo().date()

    inicio_dia = datetime.combine(
        hoje,
        datetime.min.time()
    )

    registros_hoje = (
        await db.execute(
            select(RegistroRotinaDB).where(
                RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                RegistroRotinaDB.convivente_id == payload.convivente_id,
                RegistroRotinaDB.cancelado != True,
                RegistroRotinaDB.data_registro >= inicio_dia
            ).order_by(
                RegistroRotinaDB.data_registro.asc()
            )
        )
    ).scalars().all()

    refeicoes_registradas = {}
    for registro in registros_hoje:
        if registro.tipo_registro in TIPOS_ROTINA_REFEICOES:
            refeicoes_registradas.setdefault(registro.tipo_registro, []).append(registro)

    # Entrada/Saída não podem resetar na virada do dia.
    # Por isso o último movimento precisa ser histórico, não apenas de hoje.
    ultimo_movimento = (
        await db.execute(
            select(RegistroRotinaDB).where(
                RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                RegistroRotinaDB.convivente_id == payload.convivente_id,
                RegistroRotinaDB.cancelado != True,
                RegistroRotinaDB.tipo_registro.in_(["Entrada", "Saída"])
            ).order_by(
                RegistroRotinaDB.data_registro.desc()
            )
        )
    ).scalars().first()

    esta_fora = (
        ultimo_movimento and
        ultimo_movimento.tipo_registro == "Saída"
    )

    # ============================================================
    # BLOQUEIOS OPERACIONAIS
    # ============================================================

    if payload.tipo_registro == "Entrada" and not esta_fora:

        raise HTTPException(
            status_code=400,
            detail="O convivente já consta como presente."
        )

    if payload.tipo_registro == "Saída" and esta_fora:

        raise HTTPException(
            status_code=400,
            detail="O convivente já consta como fora."
        )

    if esta_fora and payload.tipo_registro != "Entrada":
        raise HTTPException(
            status_code=400,
            detail="Convivente está fora da unidade. Registre uma entrada antes de qualquer interação."
        )

    if payload.tipo_registro in TIPOS_ROTINA_REFEICOES:

        if payload.tipo_registro in refeicoes_registradas:
            registros_refeicao = refeicoes_registradas[payload.tipo_registro]

            if not payload.confirmar_refeicao_extra:
                registro_anterior = registros_refeicao[-1]
                horario = registro_anterior.data_registro.strftime("%H:%M")
                rotulo_refeicao = ROTULOS_REFEICOES_EXTRAS.get(
                    payload.tipo_registro,
                    payload.tipo_registro.lower(),
                )

                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=(
                        f"Este convivente já teve {rotulo_refeicao} registrado hoje às {horario}. "
                        f"Confirme para registrar mais um(a) {rotulo_refeicao} como refeição extra."
                    ),
                )

    if payload.tipo_registro in TIPOS_ROTINA_PARES:
        if "Cobertor" in payload.tipo_registro:
            tipos_do_grupo = ["Retirada de Cobertor", "Entrega de Cobertor"]
            nome_grupo = "cobertor"
        else:
            tipos_do_grupo = ["Retirada de Toalha", "Entrega de Toalha"]
            nome_grupo = "toalha"

        ultima_interacao_grupo = (
            await db.execute(
                select(RegistroRotinaDB).where(
                    RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                    RegistroRotinaDB.convivente_id == payload.convivente_id,
                    RegistroRotinaDB.cancelado != True,
                    RegistroRotinaDB.tipo_registro.in_(tipos_do_grupo),
                ).order_by(
                    RegistroRotinaDB.data_registro.desc()
                )
            )
        ).scalars().first()

        if ultima_interacao_grupo and ultima_interacao_grupo.tipo_registro == payload.tipo_registro:
            proxima_acao = tipos_do_grupo[1] if payload.tipo_registro == tipos_do_grupo[0] else tipos_do_grupo[0]
            horario = ultima_interacao_grupo.data_registro.strftime("%d/%m/%Y às %H:%M")
            raise HTTPException(
                status_code=400,
                detail=(
                    f"A última movimentação de {nome_grupo} foi {payload.tipo_registro} em {horario}. "
                    f"A próxima ação esperada é {proxima_acao}."
                ),
            )

        if not ultima_interacao_grupo and payload.tipo_registro.startswith("Entrega"):
            raise HTTPException(
                status_code=400,
                detail=f"Não há retirada anterior de {nome_grupo} para registrar entrega.",
            )

    # ============================================================
    # MOVIMENTO RÁPIDO (<10 MIN)
    # ============================================================

    retorno_rapido = False

    justificativa_retorno_rapido = None

    if (
        payload.tipo_registro in {"Entrada", "Saída"}
        and ultimo_movimento
        and ultimo_movimento.tipo_registro in {"Entrada", "Saída"}
        and ultimo_movimento.tipo_registro != payload.tipo_registro
    ):

        diferenca = (
            agora_sao_paulo() -
            ultimo_movimento.data_registro
        )

        if diferenca <= timedelta(minutes=10):

            retorno_rapido = True

            if (
                not payload.justificativa_retorno_rapido
                or not payload.justificativa_retorno_rapido.strip()
            ):

                raise HTTPException(
                    status_code=400,
                    detail="Movimento em menos de 10 minutos exige justificativa."
                )

            justificativa_retorno_rapido = (
                payload.justificativa_retorno_rapido.strip()
            )

    await verificar_mes_fechado(
        db,
        usuario_atual,
        agora_sao_paulo(),
        acao="registrar movimento de rotina"
    )

    novo_registro = RegistroRotinaDB(
        instituicao_id=obter_instituicao_escopo(usuario_atual),
        convivente_id=payload.convivente_id,
        usuario_id=usuario_atual["sub"],

        tipo_registro=payload.tipo_registro,
        observacao=observacao or None,
        data_registro=agora_sao_paulo(),

        retorno_rapido=retorno_rapido,
        justificativa_retorno_rapido=justificativa_retorno_rapido
    )

    db.add(novo_registro)

    await db.commit()

    await db.refresh(novo_registro)

    return novo_registro


@router.get("/rotina/hoje")
async def resumo_rotina_hoje(
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):

    # ============================================================
    # REGISTROS DE HOJE (apenas almoço e histórico visual)
    # ============================================================

    hoje = agora_sao_paulo().date()

    inicio_dia = datetime.combine(
        hoje,
        datetime.min.time()
    )

    registros_hoje = (
        await db.execute(
            select(RegistroRotinaDB).where(
                RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                RegistroRotinaDB.cancelado != True,
                RegistroRotinaDB.data_registro >= inicio_dia
            ).order_by(
                RegistroRotinaDB.data_registro.asc()
            )
        )
    ).scalars().all()

    # ============================================================
    # ÚLTIMO MOVIMENTO HISTÓRICO REAL
    # (não pode resetar na virada do dia)
    # ============================================================

    ultimo_movimento_subq = (
        select(
            RegistroRotinaDB.convivente_id.label("convivente_id"),
            func.max(RegistroRotinaDB.data_registro).label("ultima_data")
        )
        .where(
            RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            RegistroRotinaDB.cancelado != True,
            RegistroRotinaDB.tipo_registro.in_(["Entrada", "Saída"])
        )
        .group_by(RegistroRotinaDB.convivente_id)
        .subquery()
    )

    ultimos_movimentos = (
        await db.execute(
            select(RegistroRotinaDB)
            .join(
                ultimo_movimento_subq,
                and_(
                    RegistroRotinaDB.convivente_id == ultimo_movimento_subq.c.convivente_id,
                    RegistroRotinaDB.data_registro == ultimo_movimento_subq.c.ultima_data
                )
            )
            .where(RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual))
        )
    ).scalars().all()

    resumo = {}

    # ============================================================
    # HISTÓRICO DE HOJE
    # ============================================================

    for r in registros_hoje:

        if r.convivente_id not in resumo:

            resumo[r.convivente_id] = {
                "presencas": [],
                "ultimo_movimento": None,
                "ultimo_movimento_id": None,
                "ultimo_movimento_data": None,
                "almocou": False,
                "refeicoes": {},
                "ultimas_interacoes": {},
            }

        resumo[r.convivente_id]["presencas"].append({
            "id": r.id,
            "tipo_registro": r.tipo_registro,
            "observacao": r.observacao,
            "data_registro": r.data_registro.isoformat(),
            "usuario_id": r.usuario_id,
            "retorno_rapido": bool(r.retorno_rapido),
            "justificativa_retorno_rapido": r.justificativa_retorno_rapido
        })

        if r.tipo_registro == "Almoço":
            resumo[r.convivente_id]["almocou"] = True

        if r.tipo_registro in TIPOS_ROTINA_REFEICOES:
            refeicao = resumo[r.convivente_id]["refeicoes"].setdefault(
                r.tipo_registro,
                {
                    "quantidade": 0,
                    "registros": [],
                    "primeiro_registro": None,
                    "ultimo_registro": None,
                },
            )
            item_refeicao = {
                "id": r.id,
                "data_registro": r.data_registro.isoformat(),
            }
            refeicao["quantidade"] += 1
            refeicao["registros"].append(item_refeicao)
            refeicao["primeiro_registro"] = refeicao["primeiro_registro"] or item_refeicao
            refeicao["ultimo_registro"] = item_refeicao

        if r.tipo_registro in TIPOS_ROTINA_PARES:
            grupo_interacao = "Cobertor" if "Cobertor" in r.tipo_registro else "Toalha"
            resumo[r.convivente_id]["ultimas_interacoes"][grupo_interacao] = {
                "tipo_registro": r.tipo_registro,
                "data_registro": r.data_registro.isoformat(),
            }

    # ============================================================
    # ESTADO REAL (último movimento histórico)
    # ============================================================

    for r in ultimos_movimentos:
        if r.convivente_id not in resumo:

            resumo[r.convivente_id] = {
                "presencas": [],
                "ultimo_movimento": None,
                "ultimo_movimento_id": None,
                "ultimo_movimento_data": None,
                "almocou": False,
                "refeicoes": {},
                "ultimas_interacoes": {},
            }

        if r.tipo_registro in [
            "Entrada",
            "Saída"
        ]:

            resumo[r.convivente_id]["ultimo_movimento"] = (
                r.tipo_registro
            )

            resumo[r.convivente_id]["ultimo_movimento_id"] = r.id

            resumo[r.convivente_id]["ultimo_movimento_data"] = (
                r.data_registro.isoformat()
            )

    return resumo


@router.get("/rotina/sync-status")
async def status_sincronizacao_rotina(
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    hoje = agora_sao_paulo().date()
    inicio_dia = datetime.combine(hoje, datetime.min.time())

    resultado = (
        await db.execute(
            select(
                func.count(RegistroRotinaDB.id),
                func.max(RegistroRotinaDB.data_registro),
                func.max(RegistroRotinaDB.cancelado_em),
                func.max(RegistroRotinaDB.editado_em),
            ).where(
                RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                RegistroRotinaDB.data_registro >= inicio_dia,
            )
        )
    ).one()

    total_registros, ultima_data, ultimo_cancelamento, ultima_edicao = resultado
    marcos = [
        marco
        for marco in (ultima_data, ultimo_cancelamento, ultima_edicao)
        if marco is not None
    ]
    ultimo_evento = max(marcos) if marcos else None

    return {
        "total_registros_hoje": int(total_registros or 0),
        "ultimo_evento": ultimo_evento.isoformat() if ultimo_evento else None,
        "verificado_em": agora_sao_paulo().isoformat(),
    }

# =====================================================================
# DASHBOARD OPERACIONAL DA ROTINA
# =====================================================================

@router.get("/rotina/dashboard-operacional")
async def dashboard_operacional_rotina(
    limite_listas: int = 120,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    """
    Dashboard operacional da rotina.

    Regra importante:
    - "Presentes agora" e "Fora agora" representam o estado real atual
      dos conviventes ativos da instituição, considerando o último movimento
      histórico válido de cada convivente.
    - Os indicadores "Entradas hoje", "Saídas hoje", "Almoços hoje" etc.
      representam apenas a movimentação do dia.
    """
    hoje = agora_sao_paulo().date()
    inicio_dia = datetime.combine(hoje, datetime.min.time())
    fim_dia = datetime.combine(hoje, datetime.max.time())

    inst_id = obter_instituicao_escopo(usuario_atual)

    # Base operacional: conviventes ativos da instituição.
    # Eles compõem o universo para "presentes agora" e "fora agora".
    conviventes_ativos = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.instituicao_id == inst_id,
                ConviventeDB.status == "Ativo"
            ).order_by(ConviventeDB.nome_completo.asc())
        )
    ).scalars().all()

    # Movimentação do dia: usada somente para métricas diárias e últimos registros.
    registros_hoje = (
        await db.execute(
            select(
                RegistroRotinaDB,
                ConviventeDB.nome_completo,
                ConviventeDB.nome_social,
                ConviventeDB.numero_institucional,
                UsuarioDB.nome.label("usuario_nome")
            )
            .join(ConviventeDB, RegistroRotinaDB.convivente_id == ConviventeDB.id)
            .join(UsuarioDB, RegistroRotinaDB.usuario_id == UsuarioDB.id)
            .where(
                RegistroRotinaDB.instituicao_id == inst_id,
                RegistroRotinaDB.data_registro >= inicio_dia,
                RegistroRotinaDB.data_registro <= fim_dia
            )
            .order_by(RegistroRotinaDB.data_registro.desc())
        )
    ).all()

    # Estado atual real: busca apenas o último movimento de fluxo por convivente.
    ultimo_movimento_subq = (
        select(
            RegistroRotinaDB.convivente_id.label("convivente_id"),
            func.max(RegistroRotinaDB.data_registro).label("ultima_data")
        )
        .where(
            RegistroRotinaDB.instituicao_id == inst_id,
            RegistroRotinaDB.cancelado != True,
            RegistroRotinaDB.tipo_registro.in_(["Entrada", "Saída"])
        )
        .group_by(RegistroRotinaDB.convivente_id)
        .subquery()
    )

    movimentos_historicos = (
        await db.execute(
            select(
                RegistroRotinaDB,
                ConviventeDB.nome_completo,
                ConviventeDB.nome_social,
                ConviventeDB.numero_institucional
            )
            .join(ConviventeDB, RegistroRotinaDB.convivente_id == ConviventeDB.id)
            .join(
                ultimo_movimento_subq,
                and_(
                    RegistroRotinaDB.convivente_id == ultimo_movimento_subq.c.convivente_id,
                    RegistroRotinaDB.data_registro == ultimo_movimento_subq.c.ultima_data
                )
            )
            .where(RegistroRotinaDB.instituicao_id == inst_id)
        )
    ).all()

    ultimo_movimento_por_convivente = {}

    for registro, nome_completo, nome_social, numero_institucional in movimentos_historicos:
        ultimo_movimento_por_convivente[registro.convivente_id] = {
            "id": registro.id,
            "convivente_id": registro.convivente_id,
            "convivente_nome": nome_social or nome_completo,
            "numero_institucional": numero_institucional,
            "tipo_registro": registro.tipo_registro,
            "data_registro": registro.data_registro,
            "origem_estado": "ultimo_movimento_historico"
        }

    ultimo_registro_subq = (
        select(
            RegistroRotinaDB.convivente_id.label("convivente_id"),
            func.max(RegistroRotinaDB.data_registro).label("ultima_data")
        )
        .where(
            RegistroRotinaDB.instituicao_id == inst_id,
            RegistroRotinaDB.cancelado != True,
        )
        .group_by(RegistroRotinaDB.convivente_id)
        .subquery()
    )

    ultimos_registros_historicos = (
        await db.execute(
            select(RegistroRotinaDB)
            .join(
                ultimo_registro_subq,
                and_(
                    RegistroRotinaDB.convivente_id == ultimo_registro_subq.c.convivente_id,
                    RegistroRotinaDB.data_registro == ultimo_registro_subq.c.ultima_data,
                )
            )
            .where(RegistroRotinaDB.instituicao_id == inst_id)
        )
    ).scalars().all()

    ultimo_registro_por_convivente = {
        registro.convivente_id: registro
        for registro in ultimos_registros_historicos
    }

    registros_validos_hoje = [
        registro for registro, *_ in registros_hoje
        if not registro.cancelado
    ]
    convivente_ids_com_registro_hoje = {
        registro.convivente_id
        for registro in registros_validos_hoje
    }

    presentes = []
    fora_por_saida = []
    sem_movimento = []

    for convivente in conviventes_ativos:
        ultimo = ultimo_movimento_por_convivente.get(convivente.id)
        ultimo_registro = ultimo_registro_por_convivente.get(convivente.id)

        if ultimo and ultimo["tipo_registro"] == "Entrada":
            presentes.append(ultimo)
        elif ultimo and ultimo["tipo_registro"] == "Saída":
            item_fora = {
                **ultimo,
                "origem_estado": "saida_registrada"
            }
            fora_por_saida.append(item_fora)
        else:
            presentes.append({
                "id": None,
                "convivente_id": convivente.id,
                "convivente_nome": convivente.nome_social or convivente.nome_completo,
                "numero_institucional": convivente.numero_institucional,
                "tipo_registro": None,
                "data_registro": None,
                "origem_estado": "cadastro_ativo_sem_saida",
            })

        if convivente.id not in convivente_ids_com_registro_hoje:
            item_sem_movimento = {
                "id": getattr(ultimo_registro, "id", None),
                "convivente_id": convivente.id,
                "convivente_nome": convivente.nome_social or convivente.nome_completo,
                "numero_institucional": convivente.numero_institucional,
                "tipo_registro": getattr(ultimo_registro, "tipo_registro", None),
                "data_registro": getattr(ultimo_registro, "data_registro", None),
                "origem_estado": "sem_movimento_hoje"
            }
            sem_movimento.append(item_sem_movimento)

    entradas_hoje = sum(
        1 for registro in registros_validos_hoje
        if registro.tipo_registro == "Entrada"
    )

    saidas_hoje = sum(
        1 for registro in registros_validos_hoje
        if registro.tipo_registro == "Saída"
    )

    almocos_hoje = sum(
        1 for registro in registros_validos_hoje
        if registro.tipo_registro == "Almoço"
    )

    cafes_hoje = sum(
        1 for registro in registros_validos_hoje
        if registro.tipo_registro == "Café da manhã"
    )

    jantares_hoje = sum(
        1 for registro in registros_validos_hoje
        if registro.tipo_registro == "Jantar"
    )

    lanches_noturnos_hoje = sum(
        1 for registro in registros_validos_hoje
        if registro.tipo_registro == "Lanche noturno"
    )

    retornos_rapidos_hoje = sum(
        1 for registro in registros_validos_hoje
        if registro.retorno_rapido
    )

    cancelados_hoje = sum(
        1 for registro, *_ in registros_hoje
        if registro.cancelado
    )

    editados_hoje = sum(
        1 for registro, *_ in registros_hoje
        if registro.foi_editado
    )

    ultimos_registros = []

    for registro, nome_completo, nome_social, numero_institucional, usuario_nome in registros_hoje[:12]:
        ultimos_registros.append({
            "id": registro.id,
            "convivente_id": registro.convivente_id,
            "convivente_nome": nome_social or nome_completo,
            "numero_institucional": numero_institucional,
            "tipo_registro": registro.tipo_registro,
            "data_registro": registro.data_registro,
            "usuario_nome": usuario_nome,
            "cancelado": bool(registro.cancelado),
            "foi_editado": bool(registro.foi_editado),
            "retorno_rapido": bool(registro.retorno_rapido),
            "justificativa_retorno_rapido": registro.justificativa_retorno_rapido
        })

    alertas = []

    if retornos_rapidos_hoje:
        alertas.append({
            "tipo": "retorno_rapido",
            "titulo": "Retornos rápidos hoje",
            "descricao": f"{retornos_rapidos_hoje} retorno(s) em menos de 10 minutos exigiram justificativa."
        })

    if cancelados_hoje:
        alertas.append({
            "tipo": "cancelamentos",
            "titulo": "Registros cancelados hoje",
            "descricao": f"{cancelados_hoje} registro(s) foram cancelados e precisam permanecer auditáveis."
        })

    if sem_movimento:
        alertas.append({
            "tipo": "sem_movimento",
            "titulo": "Conviventes ativos sem movimentação hoje",
            "descricao": f"{len(sem_movimento)} convivente(s) ativo(s) ainda não possuem registro de rotina hoje."
        })

    capacidade_operacional = len(conviventes_ativos)
    percentual_presentes = (
        round((len(presentes) / capacidade_operacional) * 100, 1)
        if capacidade_operacional
        else 0
    )
    limite_listas_seguro = min(max(int(limite_listas or 120), 20), 500)

    return {
        "data_referencia": hoje.isoformat(),
        "atualizado_em": agora_sao_paulo().isoformat(),
        "resumo": {
            "conviventes_ativos": capacidade_operacional,

            # Estado real atual.
            "presentes_agora": len(presentes),
            "fora_agora": len(fora_por_saida),
            "dentro_projeto": len(presentes),
            "fora_projeto": len(fora_por_saida),
            "fora_com_saida": len(fora_por_saida),
            "sem_movimento": len(sem_movimento),
            "percentual_presentes": percentual_presentes,

            # Movimento diário.
            "entradas_hoje": entradas_hoje,
            "saidas_hoje": saidas_hoje,
            "cafes_hoje": cafes_hoje,
            "almocos_hoje": almocos_hoje,
            "jantares_hoje": jantares_hoje,
            "lanches_noturnos_hoje": lanches_noturnos_hoje,
            "retornos_rapidos_hoje": retornos_rapidos_hoje,
            "cancelados_hoje": cancelados_hoje,
            "editados_hoje": editados_hoje,
            "total_registros_hoje": len(registros_validos_hoje)
        },
        "listas_totais": {
            "presentes": len(presentes),
            "fora": len(fora_por_saida),
            "fora_com_saida": len(fora_por_saida),
            "sem_movimento": len(sem_movimento)
        },
        "limite_listas": limite_listas_seguro,
        "presentes": presentes[:limite_listas_seguro],
        "fora": fora_por_saida[:limite_listas_seguro],
        "fora_com_saida": fora_por_saida[:limite_listas_seguro],
        "sem_movimento": sem_movimento[:limite_listas_seguro],
        "ultimos_registros": ultimos_registros,
        "alertas": alertas
    }


# =====================================================================
# HISTÓRICO GERAL DA ROTINA / PORTARIA
# =====================================================================


def _aplicar_filtros_historico_rotina(
    query,
    data_inicio: str = None,
    data_fim: str = None,
    tipo_registro: str = None,
    convivente_id: str = None,
    busca: str = None,
    status_registro: str = None,
    apenas_editados: bool = False,
    apenas_cancelados: bool = False,
    apenas_retorno_rapido: bool = False
):
    if data_inicio:
        try:
            inicio = datetime.fromisoformat(data_inicio)
            query = query.where(RegistroRotinaDB.data_registro >= inicio)
        except ValueError:
            raise HTTPException(status_code=400, detail="Data inicial inválida.")

    if data_fim:
        try:
            fim = datetime.fromisoformat(data_fim)
            if len(data_fim) == 10:
                fim = fim.replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.where(RegistroRotinaDB.data_registro <= fim)
        except ValueError:
            raise HTTPException(status_code=400, detail="Data final inválida.")

    if tipo_registro:
        grupos_tipo_registro = {
            "Cobertor": ["Retirada de Cobertor", "Entrega de Cobertor"],
            "Toalha": ["Retirada de Toalha", "Entrega de Toalha"],
            "Documentos": ["Bipar documentos guardados", "Bipar documentos retirados"],
            "Bagageiro": ["Movimentação de Bagageiro"],
            "Lavanderia": [
                "Lavanderia - Entrega",
                "Lavanderia - Retirada",
                "Lavanderia - Cancelamento",
            ],
            "Pertences recolhidos": ["Pertences recolhidos - Retirada"],
        }
        if tipo_registro in grupos_tipo_registro:
            query = query.where(
                RegistroRotinaDB.tipo_registro.in_(grupos_tipo_registro[tipo_registro])
            )
        else:
            query = query.where(RegistroRotinaDB.tipo_registro == tipo_registro)

    if convivente_id:
        query = query.where(RegistroRotinaDB.convivente_id == convivente_id)

    if busca:
        termo = f"%{busca.strip()}%"
        query = query.where(
            or_(
                ConviventeDB.nome_completo.ilike(termo),
                ConviventeDB.nome_social.ilike(termo),
                func.cast(ConviventeDB.numero_institucional, String).ilike(termo)
            )
        )

    if status_registro == "ativos":
        query = query.where(RegistroRotinaDB.cancelado == False)

    if status_registro == "cancelados":
        query = query.where(RegistroRotinaDB.cancelado == True)

    if apenas_editados:
        query = query.where(RegistroRotinaDB.foi_editado == True)

    if apenas_cancelados:
        query = query.where(RegistroRotinaDB.cancelado == True)

    if apenas_retorno_rapido:
        query = query.where(RegistroRotinaDB.retorno_rapido == True)

    return query


def _query_base_historico_rotina(usuario_atual: dict):
    return (
        select(
            RegistroRotinaDB,
            ConviventeDB.nome_completo,
            ConviventeDB.nome_social,
            ConviventeDB.numero_institucional,
            UsuarioDB.nome.label("usuario_nome"),
            UsuarioDB.perfil_acesso.label("usuario_perfil")
        )
        .join(ConviventeDB, RegistroRotinaDB.convivente_id == ConviventeDB.id)
        .join(UsuarioDB, RegistroRotinaDB.usuario_id == UsuarioDB.id)
        .where(
            RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual)
        )
        .order_by(RegistroRotinaDB.data_registro.desc())
    )


def _linha_historico_para_dict(
    registro,
    nome_completo,
    nome_social,
    numero_institucional,
    usuario_nome,
    usuario_perfil
):
    return {
        "id": registro.id,
        "instituicao_id": registro.instituicao_id,
        "convivente_id": registro.convivente_id,
        "convivente_nome": nome_social or nome_completo,
        "convivente_nome_completo": nome_completo,
        "numero_institucional": numero_institucional,

        "tipo_registro": registro.tipo_registro,
        "observacao": registro.observacao,
        "data_registro": registro.data_registro,

        "usuario_id": registro.usuario_id,
        "usuario_nome": usuario_nome,
        "usuario_perfil": usuario_perfil,

        "retorno_rapido": bool(registro.retorno_rapido),
        "justificativa_retorno_rapido": registro.justificativa_retorno_rapido,

        "foi_editado": bool(registro.foi_editado),
        "editado_por_id": registro.editado_por_id,
        "editado_em": registro.editado_em,
        "motivo_edicao": registro.motivo_edicao,
        "tipo_registro_original": registro.tipo_registro_original,
        "data_registro_original": registro.data_registro_original,

        "cancelado": bool(registro.cancelado),
        "cancelado_por_id": registro.cancelado_por_id,
        "cancelado_em": registro.cancelado_em,
        "motivo_cancelamento": registro.motivo_cancelamento
    }


def _evento_fluxo_operacional_dict(
    *,
    evento_id: str,
    instituicao_id: str,
    convivente_id: str,
    convivente_nome: str | None,
    numero_institucional,
    tipo_registro: str,
    observacao: str | None,
    data_registro,
    usuario_id: str,
    usuario_nome: str | None,
    usuario_perfil: str | None,
    cancelado: bool = False,
    motivo_cancelamento: str | None = None,
) -> dict:
    return {
        "id": evento_id,
        "instituicao_id": instituicao_id,
        "convivente_id": convivente_id,
        "convivente_nome": convivente_nome,
        "convivente_nome_completo": convivente_nome,
        "numero_institucional": numero_institucional,
        "tipo_registro": tipo_registro,
        "observacao": observacao,
        "data_registro": data_registro,
        "usuario_id": usuario_id,
        "usuario_nome": usuario_nome,
        "usuario_perfil": usuario_perfil,
        "retorno_rapido": False,
        "justificativa_retorno_rapido": None,
        "foi_editado": False,
        "editado_por_id": None,
        "editado_em": None,
        "motivo_edicao": None,
        "tipo_registro_original": None,
        "data_registro_original": None,
        "cancelado": cancelado,
        "cancelado_por_id": None,
        "cancelado_em": None,
        "motivo_cancelamento": motivo_cancelamento,
        "origem_operacional": True,
    }


async def _mapear_usuarios_fluxo_operacional(
    db: AsyncSession,
    usuario_ids: set[str],
) -> dict[str, tuple[str | None, str | None]]:
    ids_limpos = {item for item in usuario_ids if item}
    if not ids_limpos:
        return {}

    linhas = (
        await db.execute(
            select(UsuarioDB.id, UsuarioDB.nome, UsuarioDB.perfil_acesso).where(
                UsuarioDB.id.in_(ids_limpos)
            )
        )
    ).all()

    return {
        usuario_id: (nome, perfil)
        for usuario_id, nome, perfil in linhas
    }


def _data_evento_no_periodo(data_registro, inicio_dt, fim_dt) -> bool:
    if not data_registro:
        return False
    if inicio_dt and data_registro < inicio_dt:
        return False
    if fim_dt and data_registro > fim_dt:
        return False
    return True


async def _coletar_eventos_fluxo_operacional_convivente(
    db: AsyncSession,
    instituicao_id: str,
    convivente_id: str,
    convivente_nome: str | None,
    numero_institucional,
    data_inicio: str = None,
    data_fim: str = None,
) -> list[dict]:
    eventos: list[dict] = []
    usuario_ids: set[str] = set()

    inicio_dt = None
    fim_dt = None
    if data_inicio:
        try:
            inicio_dt = datetime.fromisoformat(data_inicio)
        except ValueError:
            raise HTTPException(status_code=400, detail="Data inicial inválida.")
    if data_fim:
        try:
            fim_dt = datetime.fromisoformat(data_fim)
            if len(data_fim) == 10:
                fim_dt = fim_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        except ValueError:
            raise HTTPException(status_code=400, detail="Data final inválida.")

    condicoes_lavanderia = [
        LavanderiaRegistroDB.instituicao_id == instituicao_id,
        LavanderiaRegistroDB.convivente_id == convivente_id,
    ]
    if inicio_dt or fim_dt:
        janelas_lavanderia = []
        if inicio_dt and fim_dt:
            janelas_lavanderia.extend([
                and_(LavanderiaRegistroDB.entregue_em >= inicio_dt, LavanderiaRegistroDB.entregue_em <= fim_dt),
                and_(LavanderiaRegistroDB.retirado_em.isnot(None), LavanderiaRegistroDB.retirado_em >= inicio_dt, LavanderiaRegistroDB.retirado_em <= fim_dt),
                and_(LavanderiaRegistroDB.cancelado_em.isnot(None), LavanderiaRegistroDB.cancelado_em >= inicio_dt, LavanderiaRegistroDB.cancelado_em <= fim_dt),
            ])
        elif inicio_dt:
            janelas_lavanderia.extend([
                LavanderiaRegistroDB.entregue_em >= inicio_dt,
                and_(LavanderiaRegistroDB.retirado_em.isnot(None), LavanderiaRegistroDB.retirado_em >= inicio_dt),
                and_(LavanderiaRegistroDB.cancelado_em.isnot(None), LavanderiaRegistroDB.cancelado_em >= inicio_dt),
            ])
        else:
            janelas_lavanderia.extend([
                LavanderiaRegistroDB.entregue_em <= fim_dt,
                and_(LavanderiaRegistroDB.retirado_em.isnot(None), LavanderiaRegistroDB.retirado_em <= fim_dt),
                and_(LavanderiaRegistroDB.cancelado_em.isnot(None), LavanderiaRegistroDB.cancelado_em <= fim_dt),
            ])
        condicoes_lavanderia.append(or_(*janelas_lavanderia))

    lavanderias = (
        await db.execute(
            select(LavanderiaRegistroDB).where(*condicoes_lavanderia)
        )
    ).scalars().all()

    for registro in lavanderias:
        usuario_ids.add(registro.usuario_entrega_id)
        if registro.usuario_retirada_id:
            usuario_ids.add(registro.usuario_retirada_id)
        if registro.cancelado_por_id:
            usuario_ids.add(registro.cancelado_por_id)

    condicoes_baixas = [
        PertenceRecolhidoBaixaDB.instituicao_id == instituicao_id,
        PertenceRecolhidoBaixaDB.convivente_id == convivente_id,
    ]
    if inicio_dt:
        condicoes_baixas.append(PertenceRecolhidoBaixaDB.baixado_em >= inicio_dt)
    if fim_dt:
        condicoes_baixas.append(PertenceRecolhidoBaixaDB.baixado_em <= fim_dt)

    baixas_pertences = (
        await db.execute(
            select(PertenceRecolhidoBaixaDB, QuartoDB.nome.label("quarto_nome"))
            .join(
                PertenceRecolhidoDB,
                PertenceRecolhidoDB.id == PertenceRecolhidoBaixaDB.pertence_recolhido_id,
            )
            .join(QuartoDB, QuartoDB.id == PertenceRecolhidoDB.quarto_id)
            .where(*condicoes_baixas)
        )
    ).all()

    for baixa, _quarto_nome in baixas_pertences:
        usuario_ids.add(baixa.usuario_id)

    usuarios = await _mapear_usuarios_fluxo_operacional(db, usuario_ids)

    for registro in lavanderias:
        nome_entrega, perfil_entrega = usuarios.get(
            registro.usuario_entrega_id, (None, None)
        )
        prazo = registro.prazo_retirada_em.strftime("%d/%m/%Y %H:%M") if registro.prazo_retirada_em else "—"
        obs_entrega = (
            f"Entrega de {registro.quantidade_entregue} peça(s). Prazo de retirada: {prazo}."
        )
        if registro.observacao_entrega:
            obs_entrega += f" Observação: {registro.observacao_entrega}"

        if _data_evento_no_periodo(registro.entregue_em, inicio_dt, fim_dt):
            eventos.append(
                _evento_fluxo_operacional_dict(
                    evento_id=f"lavanderia-entrega-{registro.id}",
                    instituicao_id=instituicao_id,
                    convivente_id=convivente_id,
                    convivente_nome=convivente_nome,
                    numero_institucional=numero_institucional,
                    tipo_registro="Lavanderia - Entrega",
                    observacao=obs_entrega,
                    data_registro=registro.entregue_em,
                    usuario_id=registro.usuario_entrega_id,
                    usuario_nome=nome_entrega,
                    usuario_perfil=perfil_entrega,
                )
            )

        if registro.quantidade_retirada and registro.retirado_em:
            nome_retirada, perfil_retirada = usuarios.get(
                registro.usuario_retirada_id or registro.usuario_entrega_id,
                (nome_entrega, perfil_entrega),
            )
            saldo = max(
                int(registro.quantidade_entregue or 0) - int(registro.quantidade_retirada or 0),
                0,
            )
            obs_retirada = (
                f"Retirada acumulada de {registro.quantidade_retirada} peça(s) "
                f"de {registro.quantidade_entregue}. Saldo pendente: {saldo}. "
                f"Status: {registro.status}."
            )
            if registro.observacao_retirada:
                obs_retirada += f"\n{registro.observacao_retirada}"

            if _data_evento_no_periodo(registro.retirado_em, inicio_dt, fim_dt):
                eventos.append(
                    _evento_fluxo_operacional_dict(
                        evento_id=f"lavanderia-retirada-{registro.id}",
                    instituicao_id=instituicao_id,
                    convivente_id=convivente_id,
                    convivente_nome=convivente_nome,
                    numero_institucional=numero_institucional,
                    tipo_registro="Lavanderia - Retirada",
                    observacao=obs_retirada,
                    data_registro=registro.retirado_em,
                    usuario_id=registro.usuario_retirada_id or registro.usuario_entrega_id,
                    usuario_nome=nome_retirada,
                    usuario_perfil=perfil_retirada,
                    )
                )

        if registro.cancelado_em and _data_evento_no_periodo(registro.cancelado_em, inicio_dt, fim_dt):
            nome_cancelamento, perfil_cancelamento = usuarios.get(
                registro.cancelado_por_id or registro.usuario_entrega_id,
                (nome_entrega, perfil_entrega),
            )
            eventos.append(
                _evento_fluxo_operacional_dict(
                    evento_id=f"lavanderia-cancelamento-{registro.id}",
                    instituicao_id=instituicao_id,
                    convivente_id=convivente_id,
                    convivente_nome=convivente_nome,
                    numero_institucional=numero_institucional,
                    tipo_registro="Lavanderia - Cancelamento",
                    observacao=f"Registro de lavanderia cancelado. Motivo: {registro.motivo_cancelamento or '—'}",
                    data_registro=registro.cancelado_em,
                    usuario_id=registro.cancelado_por_id or registro.usuario_entrega_id,
                    usuario_nome=nome_cancelamento,
                    usuario_perfil=perfil_cancelamento,
                    cancelado=True,
                    motivo_cancelamento=registro.motivo_cancelamento,
                )
            )

    for baixa, quarto_nome in baixas_pertences:
        nome_usuario, perfil_usuario = usuarios.get(baixa.usuario_id, (None, None))
        obs = (
            f"Retirada de {baixa.quantidade} item(ns) recolhido(s) do quarto {quarto_nome or '—'}."
        )
        if baixa.justificativa:
            obs += f" Observação: {baixa.justificativa}"

        eventos.append(
            _evento_fluxo_operacional_dict(
                evento_id=f"pertences-retirada-{baixa.id}",
                instituicao_id=instituicao_id,
                convivente_id=convivente_id,
                convivente_nome=convivente_nome,
                numero_institucional=numero_institucional,
                tipo_registro="Pertences recolhidos - Retirada",
                observacao=obs,
                data_registro=baixa.baixado_em,
                usuario_id=baixa.usuario_id,
                usuario_nome=nome_usuario,
                usuario_perfil=perfil_usuario,
            )
        )

    return eventos


def _filtrar_eventos_fluxo_merged(
    eventos: list[dict],
    *,
    data_inicio: str = None,
    data_fim: str = None,
    tipo_registro: str = None,
    busca: str = None,
    status_registro: str = None,
    apenas_editados: bool = False,
    apenas_cancelados: bool = False,
    apenas_retorno_rapido: bool = False,
) -> list[dict]:
    filtrados = eventos

    if data_inicio:
        try:
            inicio = datetime.fromisoformat(data_inicio)
            filtrados = [
                evento for evento in filtrados
                if evento.get("data_registro") and evento["data_registro"] >= inicio
            ]
        except ValueError:
            raise HTTPException(status_code=400, detail="Data inicial inválida.")

    if data_fim:
        try:
            fim = datetime.fromisoformat(data_fim)
            if len(data_fim) == 10:
                fim = fim.replace(hour=23, minute=59, second=59, microsecond=999999)
            filtrados = [
                evento for evento in filtrados
                if evento.get("data_registro") and evento["data_registro"] <= fim
            ]
        except ValueError:
            raise HTTPException(status_code=400, detail="Data final inválida.")

    if tipo_registro:
        grupos_tipo_registro = {
            "Cobertor": ["Retirada de Cobertor", "Entrega de Cobertor"],
            "Toalha": ["Retirada de Toalha", "Entrega de Toalha"],
            "Documentos": ["Bipar documentos guardados", "Bipar documentos retirados"],
            "Bagageiro": ["Movimentação de Bagageiro"],
            "Lavanderia": [
                "Lavanderia - Entrega",
                "Lavanderia - Retirada",
                "Lavanderia - Cancelamento",
            ],
            "Pertences recolhidos": ["Pertences recolhidos - Retirada"],
        }
        if tipo_registro in grupos_tipo_registro:
            tipos_validos = set(grupos_tipo_registro[tipo_registro])
            filtrados = [
                evento for evento in filtrados
                if evento.get("tipo_registro") in tipos_validos
            ]
        else:
            filtrados = [
                evento for evento in filtrados
                if evento.get("tipo_registro") == tipo_registro
            ]

    if busca:
        termo = busca.strip().casefold()
        filtrados = [
            evento for evento in filtrados
            if termo in (evento.get("convivente_nome") or "").casefold()
            or termo in (evento.get("convivente_nome_completo") or "").casefold()
            or termo in str(evento.get("numero_institucional") or "").casefold()
            or termo in (evento.get("observacao") or "").casefold()
            or termo in (evento.get("tipo_registro") or "").casefold()
        ]

    if status_registro == "ativos":
        filtrados = [evento for evento in filtrados if not evento.get("cancelado")]

    if status_registro == "cancelados":
        filtrados = [evento for evento in filtrados if evento.get("cancelado")]

    if apenas_editados:
        filtrados = [evento for evento in filtrados if evento.get("foi_editado")]

    if apenas_cancelados:
        filtrados = [evento for evento in filtrados if evento.get("cancelado")]

    if apenas_retorno_rapido:
        filtrados = [evento for evento in filtrados if evento.get("retorno_rapido")]

    return filtrados


@router.get("/rotina/historico/resumo-evolucao")
async def resumo_evolucao_historico_rotina(
    data_inicio: str = None,
    data_fim: str = None,
    tecnico_id: str = None,
    status_convivente: str = None,
    busca: str = None,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    data_dia = func.date(RegistroRotinaDB.data_registro).label("data")
    query = (
        select(
            data_dia,
            func.count(RegistroRotinaDB.id).label("atendimentos"),
            func.sum(case((RegistroRotinaDB.tipo_registro == "Entrada", 1), else_=0)).label("entradas"),
            func.sum(case((RegistroRotinaDB.tipo_registro == "Saída", 1), else_=0)).label("saidas"),
            func.sum(case((RegistroRotinaDB.tipo_registro.in_(list(TIPOS_ROTINA_REFEICOES)), 1), else_=0)).label("almocos"),
        )
        .join(ConviventeDB, RegistroRotinaDB.convivente_id == ConviventeDB.id)
        .where(RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual))
    )

    query = _aplicar_filtros_historico_rotina(
        query=query,
        data_inicio=data_inicio,
        data_fim=data_fim,
        busca=busca,
    )

    if tecnico_id:
        query = query.where(ConviventeDB.tecnico_id == tecnico_id)

    if status_convivente and status_convivente != "Todos":
        query = query.where(ConviventeDB.status == status_convivente)

    query = query.group_by(data_dia).order_by(data_dia.asc())
    resultado = await db.execute(query)

    return [
        {
            "data": data,
            "atendimentos": int(atendimentos or 0),
            "entradas": int(entradas or 0),
            "saidas": int(saidas or 0),
            "almocos": int(almocos or 0),
        }
        for data, atendimentos, entradas, saidas, almocos in resultado.all()
    ]


@router.get("/rotina/historico", response_model=RotinaHistoricoListaResponse)
async def listar_historico_rotina(
    data_inicio: str = None,
    data_fim: str = None,
    tipo_registro: str = None,
    convivente_id: str = None,
    busca: str = None,
    status_registro: str = None,
    apenas_editados: bool = False,
    apenas_cancelados: bool = False,
    apenas_retorno_rapido: bool = False,
    limite: int = Query(REGISTROS_POR_PAGINA_PADRAO, ge=1, le=EXPORT_ROTINA_HISTORICO_LIMITE_MAX),
    deslocamento: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    data_inicio, data_fim = _normalizar_periodo_listagem_operacional(data_inicio, data_fim)
    limite_seguro = min(max(int(limite or REGISTROS_POR_PAGINA_PADRAO), 1), EXPORT_ROTINA_HISTORICO_LIMITE_MAX)
    deslocamento_seguro = max(int(deslocamento or 0), 0)

    query = _query_base_historico_rotina(usuario_atual)
    query = _aplicar_filtros_historico_rotina(
        query=query,
        data_inicio=data_inicio,
        data_fim=data_fim,
        tipo_registro=tipo_registro,
        convivente_id=convivente_id,
        busca=busca,
        status_registro=status_registro,
        apenas_editados=apenas_editados,
        apenas_cancelados=apenas_cancelados,
        apenas_retorno_rapido=apenas_retorno_rapido,
    )

    if convivente_id:
        resultado = await db.execute(query)
        linhas = resultado.all()

        convivente_nome = None
        numero_institucional = None
        if linhas:
            _, nome_completo, nome_social, numero_institucional, _, _ = linhas[0]
            convivente_nome = nome_social or nome_completo
        else:
            convivente = (
                await db.execute(
                    select(ConviventeDB).where(
                        ConviventeDB.id == convivente_id,
                        ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                    )
                )
            ).scalar_one_or_none()
            if convivente:
                convivente_nome = convivente.nome_social or convivente.nome_completo
                numero_institucional = convivente.numero_institucional

        itens = [
            _linha_historico_para_dict(
                registro,
                nome_completo,
                nome_social,
                numero_institucional_linha,
                usuario_nome,
                usuario_perfil,
            )
            for registro, nome_completo, nome_social, numero_institucional_linha, usuario_nome, usuario_perfil in linhas
        ]

        eventos_operacionais = await _coletar_eventos_fluxo_operacional_convivente(
            db,
            obter_instituicao_escopo(usuario_atual),
            convivente_id,
            convivente_nome,
            numero_institucional,
            data_inicio=data_inicio,
            data_fim=data_fim,
        )
        itens.extend(eventos_operacionais)
        itens = _filtrar_eventos_fluxo_merged(
            itens,
            data_inicio=data_inicio,
            data_fim=data_fim,
            tipo_registro=tipo_registro,
            busca=busca,
            status_registro=status_registro,
            apenas_editados=apenas_editados,
            apenas_cancelados=apenas_cancelados,
            apenas_retorno_rapido=apenas_retorno_rapido,
        )
        itens.sort(
            key=lambda item: item.get("data_registro") or datetime.min,
            reverse=True,
        )
        total = len(itens)
        pagina = itens[deslocamento_seguro:deslocamento_seguro + limite_seguro]
        return {
            "registros": pagina,
            "total": total,
            "limite": limite_seguro,
            "deslocamento": deslocamento_seguro,
            "has_more": deslocamento_seguro + len(pagina) < total,
            "resumo_periodo": None,
        }

    total = int((
        await db.execute(
            select(func.count()).select_from(query.order_by(None).subquery())
        )
    ).scalar_one() or 0)

    def contar_resumo(expressao):
        return select(func.count()).select_from(
            query.order_by(None).where(expressao).subquery()
        )

    resumo_periodo = {
        "total": total,
        "entradas": int((await db.execute(contar_resumo(RegistroRotinaDB.tipo_registro == "Entrada"))).scalar_one() or 0),
        "saidas": int((await db.execute(contar_resumo(RegistroRotinaDB.tipo_registro == "Saída"))).scalar_one() or 0),
        "editados": int((await db.execute(contar_resumo(RegistroRotinaDB.foi_editado == True))).scalar_one() or 0),
        "cancelados": int((await db.execute(contar_resumo(RegistroRotinaDB.cancelado == True))).scalar_one() or 0),
        "retornos_rapidos": int((await db.execute(contar_resumo(RegistroRotinaDB.retorno_rapido == True))).scalar_one() or 0),
    }

    query = query.offset(deslocamento_seguro).limit(limite_seguro)
    resultado = await db.execute(query)
    linhas = resultado.all()

    registros = [
        _linha_historico_para_dict(
            registro,
            nome_completo,
            nome_social,
            numero_institucional,
            usuario_nome,
            usuario_perfil,
        )
        for registro, nome_completo, nome_social, numero_institucional, usuario_nome, usuario_perfil in linhas
    ]

    return {
        "registros": registros,
        "total": total,
        "limite": limite_seguro,
        "deslocamento": deslocamento_seguro,
        "has_more": deslocamento_seguro + len(registros) < total,
        "resumo_periodo": resumo_periodo,
    }


@router.get("/rotina/historico/exportar-xlsx")
async def exportar_historico_rotina_xlsx(
    data_inicio: str = None,
    data_fim: str = None,
    tipo_registro: str = None,
    convivente_id: str = None,
    busca: str = None,
    status_registro: str = None,
    apenas_editados: bool = False,
    apenas_cancelados: bool = False,
    apenas_retorno_rapido: bool = False,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    query = _query_base_historico_rotina(usuario_atual)
    data_inicio, data_fim = _normalizar_periodo_listagem_operacional(data_inicio, data_fim)

    query = _aplicar_filtros_historico_rotina(
        query=query,
        data_inicio=data_inicio,
        data_fim=data_fim,
        tipo_registro=tipo_registro,
        convivente_id=convivente_id,
        busca=busca,
        status_registro=status_registro,
        apenas_editados=apenas_editados,
        apenas_cancelados=apenas_cancelados,
        apenas_retorno_rapido=apenas_retorno_rapido
    )

    resultado = await db.execute(
        query.limit(EXPORT_ROTINA_HISTORICO_LIMITE_MAX)
    )
    linhas = resultado.all()

    historico = [
        _linha_historico_para_dict(
            registro,
            nome_completo,
            nome_social,
            numero_institucional,
            usuario_nome,
            usuario_perfil
        )
        for registro, nome_completo, nome_social, numero_institucional, usuario_nome, usuario_perfil in linhas
    ]

    arquivo = gerar_xlsx_historico(historico)

    nome_arquivo = f"historico_rotina_{agora_sao_paulo().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        BytesIO(arquivo),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"'
        }
    )



# =====================================================================
# CONVÊNIO / SISA / FECHAMENTO MENSAL
# =====================================================================

def _parse_data_simples(valor: str, nome_campo: str) -> datetime:
    try:
        return datetime.fromisoformat(valor)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail=f"{nome_campo} inválida. Use o formato AAAA-MM-DD."
        )


def _inicio_fim_dia(data: str):
    base_data = _parse_data_simples(data, "Data").date()
    inicio = datetime.combine(base_data, datetime.min.time())
    fim = datetime.combine(base_data, datetime.max.time())
    return inicio, fim


def _inicio_fim_mes(ano: int, mes: int):
    if mes < 1 or mes > 12:
        raise HTTPException(status_code=400, detail="Mês inválido.")

    inicio = datetime(ano, mes, 1)

    if mes == 12:
        fim = datetime(ano + 1, 1, 1) - timedelta(microseconds=1)
    else:
        fim = datetime(ano, mes + 1, 1) - timedelta(microseconds=1)

    return inicio, fim


async def _conviventes_ativos_instituicao(
    db: AsyncSession,
    instituicao_id: str
):
    resultado = await db.execute(
        select(ConviventeDB).where(
            ConviventeDB.instituicao_id == instituicao_id,
            ConviventeDB.status == "Ativo"
        ).order_by(
            ConviventeDB.nome_completo.asc()
        )
    )

    return resultado.scalars().all()


async def _registros_rotina_periodo(
    db: AsyncSession,
    instituicao_id: str,
    inicio: datetime,
    fim: datetime
):
    resultado = await db.execute(
        select(RegistroRotinaDB).where(
            RegistroRotinaDB.instituicao_id == instituicao_id,
            RegistroRotinaDB.cancelado != True,
            RegistroRotinaDB.data_registro >= inicio,
            RegistroRotinaDB.data_registro <= fim
        ).order_by(
            RegistroRotinaDB.data_registro.asc()
        )
    )

    return resultado.scalars().all()


def _nome_convivente_relatorio(convivente: ConviventeDB):
    return convivente.nome_social or convivente.nome_completo


def _item_atende_filtro_sisa(item: dict, tipo_atendimento: str | None) -> bool:
    filtro = (tipo_atendimento or "todos").strip()

    if filtro == "todos":
        return True

    tem_entrada = bool(item.get("entrada")) or int(item.get("entradas") or 0) > 0
    tem_saida = bool(item.get("saida")) or int(item.get("saidas") or 0) > 0
    tem_cafe = item.get("cafe") == "Sim" or int(item.get("cafes") or 0) > 0
    tem_almoco = item.get("almoco") == "Sim" or int(item.get("almocos") or 0) > 0
    tem_jantar = item.get("jantar") == "Sim" or int(item.get("jantares") or 0) > 0
    tem_lanche = item.get("lanche") == "Sim" or int(item.get("lanches") or 0) > 0
    tem_banho = item.get("banho") == "Sim" or int(item.get("banhos") or 0) > 0
    tem_alimentacao = tem_cafe or tem_almoco or tem_jantar or tem_lanche
    tem_movimento = int(item.get("total_movimentos") or 0) > 0
    tem_presenca = (
        item.get("presenca") == "Sim"
        or int(item.get("dias_presentes") or item.get("total_atendimentos") or 0) > 0
    )
    tem_atendimento = tem_movimento or tem_entrada or tem_saida or tem_alimentacao or tem_banho

    if filtro == "com_presenca":
        return tem_presenca
    if filtro == "sem_presenca":
        return not tem_presenca
    if filtro == "com_atendimento":
        return tem_atendimento
    if filtro == "entrada":
        return tem_entrada
    if filtro == "cafe":
        return tem_cafe
    if filtro == "almoco":
        return tem_almoco
    if filtro == "jantar":
        return tem_jantar
    if filtro == "lanche":
        return tem_lanche
    if filtro == "banho":
        return tem_banho
    if filtro == "alimentacao":
        return tem_alimentacao
    if filtro == "saida":
        return tem_saida
    if filtro == "entrada_almoco":
        return tem_entrada and tem_almoco
    if filtro == "sem_atendimento":
        return not tem_atendimento

    return True


def _calcular_dias_presenca_operacional(
    movimentos: list[dict],
    data_inicio: date,
    data_fim: date,
    data_entrada: date | None = None,
) -> list[str]:
    movimentos_ordenados = sorted(
        [
            movimento for movimento in movimentos
            if movimento.get("tipo_registro") in {"Entrada", "Saída"}
        ],
        key=lambda movimento: movimento["data_registro"],
    )

    dias_presentes = []
    dia = data_inicio

    while dia <= data_fim:
        if data_entrada and dia < data_entrada:
            dia += timedelta(days=1)
            continue

        inicio_dia = datetime.combine(dia, datetime.min.time())
        fim_dia = datetime.combine(dia, datetime.max.time())

        ultimo_antes_do_dia = None
        ultimo_ate_fim_do_dia = None
        entrou_no_dia = False

        for movimento in movimentos_ordenados:
            data_movimento = movimento["data_registro"]

            if data_movimento < inicio_dia:
                ultimo_antes_do_dia = movimento
                ultimo_ate_fim_do_dia = movimento
                continue

            if data_movimento > fim_dia:
                break

            ultimo_ate_fim_do_dia = movimento
            if movimento["tipo_registro"] == "Entrada":
                entrou_no_dia = True

        amanheceu_dentro = (
            ultimo_antes_do_dia is None
            or ultimo_antes_do_dia["tipo_registro"] == "Entrada"
        )
        dentro_no_fechamento = (
            ultimo_ate_fim_do_dia is None
            or ultimo_ate_fim_do_dia["tipo_registro"] == "Entrada"
        )

        if dentro_no_fechamento or amanheceu_dentro or entrou_no_dia:
            dias_presentes.append(dia.isoformat())

        dia += timedelta(days=1)

    return dias_presentes


TIPOS_SISA_PRESENCA_CARECORE = {
    "Entrada",
    "Saída",
    "Café da manhã",
    "Almoço",
    "Jantar",
    "Lanche noturno",
}


def _texto_planilha(valor):
    if valor is None:
        return None

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    texto = str(valor).strip()
    return texto or None


def _quantidade_extra_refeicao(registros: list) -> int:
    return max(len(registros) - 1, 0)


def _data_br_para_date(valor):
    texto = _texto_planilha(valor)
    if not texto or texto == "---":
        return None

    try:
        return datetime.strptime(texto, "%d/%m/%Y").date()
    except ValueError:
        return None


def _extrair_primeira_data(texto):
    if not texto:
        return None

    match = re.search(r"(\d{2}/\d{2}/\d{4})", str(texto))
    if not match:
        return None

    return _data_br_para_date(match.group(1))


def _extrair_periodo_referencia_sisa(texto):
    datas = [
        _data_br_para_date(data)
        for data in re.findall(r"\d{2}/\d{2}/\d{4}", str(texto or ""))
    ]
    datas = [data for data in datas if data]

    if len(datas) >= 2:
        return datas[0], datas[1]
    if datas:
        return datas[0], datas[0]

    return None, None


def _normalizar_numero_sisa(valor):
    texto = _texto_planilha(valor)
    if not texto:
        return None

    return re.sub(r"\D", "", texto) or texto


def _mapear_sexo_sisa(valor):
    texto = _texto_planilha(valor)
    if texto == "M":
        return "Masculino"
    if texto == "F":
        return "Feminino"
    return texto


def _linha_planilha_para_registro_sisa(sheet, indice_linha):
    numero_sisa = _normalizar_numero_sisa(sheet.cell_value(indice_linha, 3))
    nome = _texto_planilha(sheet.cell_value(indice_linha, 7))

    if not numero_sisa or not nome:
        return None

    dias_valor = sheet.cell_value(indice_linha, 15)
    try:
        dias_permanencia = int(dias_valor or 0)
    except (TypeError, ValueError):
        dias_permanencia = 0

    return {
        "numero_sisa": numero_sisa,
        "nome_planilha": nome,
        "nome_social_planilha": _texto_planilha(sheet.cell_value(indice_linha, 8)),
        "data_nascimento": _data_br_para_date(sheet.cell_value(indice_linha, 11)),
        "sexo": _mapear_sexo_sisa(sheet.cell_value(indice_linha, 12)),
        "data_vinculacao": _data_br_para_date(sheet.cell_value(indice_linha, 13)),
        "data_desligamento": _data_br_para_date(sheet.cell_value(indice_linha, 14)),
        "dias_permanencia": dias_permanencia,
    }


def _dias_sisa_no_recorte(
    *,
    data_inicio_recorte: date,
    data_fim_recorte: date,
    data_vinculacao: date | None,
    data_desligamento: date | None,
    dias_permanencia_acumulado: int,
) -> int:
    inicio = max(data_inicio_recorte, data_vinculacao or data_inicio_recorte)
    fim = min(data_fim_recorte, data_desligamento or data_fim_recorte)

    if fim < inicio:
        return 0

    dias_no_recorte = (fim - inicio).days + 1
    return max(0, min(dias_no_recorte, int(dias_permanencia_acumulado or 0)))


def _desligamento_valido_no_recorte(data_desligamento: date | None, data_referencia: date) -> bool:
    return bool(data_desligamento and data_desligamento <= data_referencia)


def _chave_ultimo_movimento_sisa(linha: dict, data_referencia: date):
    movimentos = []
    if linha.get("data_vinculacao") and linha["data_vinculacao"] <= data_referencia:
        movimentos.append((linha["data_vinculacao"], 1))
    if linha.get("data_desligamento") and linha["data_desligamento"] <= data_referencia:
        movimentos.append((linha["data_desligamento"], 2))

    if not movimentos:
        return date.min, 0

    return max(movimentos)


def _consolidar_linhas_sisa_por_ultimo_movimento(linhas: list[dict], data_referencia: date) -> list[dict]:
    linhas_por_numero = {}

    for linha in linhas:
        numero_sisa = linha["numero_sisa"]
        chave_movimento = _chave_ultimo_movimento_sisa(linha, data_referencia)
        if numero_sisa not in linhas_por_numero or chave_movimento > linhas_por_numero[numero_sisa][0]:
            linhas_por_numero[numero_sisa] = (chave_movimento, linha)

    return [linha for _, linha in linhas_por_numero.values()]


def _linha_sisa_ativa_no_recorte(linha: dict, data_referencia: date) -> bool:
    chave_movimento = _chave_ultimo_movimento_sisa(linha, data_referencia)
    return chave_movimento[1] != 2


def _ler_planilha_sisa(conteudo: bytes, nome_arquivo: str):
    extensao = os.path.splitext(nome_arquivo.lower())[1]

    if extensao == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="Leitor de planilha .xls não instalado. Rode pip install -r requirements.txt e reinicie o backend."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=conteudo)
        sheet = workbook.sheet_by_index(0)

        linhas = []
        data_inicio_referencia = None
        data_referencia = None
        servico = None

        for indice in range(min(sheet.nrows, 12)):
            valores = [
                _texto_planilha(sheet.cell_value(indice, coluna))
                for coluna in range(sheet.ncols)
            ]
            texto_linha = " ".join(valor for valor in valores if valor)
            inicio_linha, fim_linha = _extrair_periodo_referencia_sisa(texto_linha)
            if inicio_linha and fim_linha and not data_referencia:
                data_inicio_referencia = inicio_linha
                data_referencia = fim_linha
            if "Tipo de Servi" in texto_linha or "SIAT" in texto_linha:
                servico = texto_linha

        for indice in range(10, sheet.nrows):
            registro = _linha_planilha_para_registro_sisa(sheet, indice)
            if registro:
                linhas.append(registro)

        return {
            "data_inicio_referencia": data_inicio_referencia or data_referencia,
            "data_referencia": data_referencia,
            "servico": servico,
            "linhas": linhas,
        }

    if extensao == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="Leitor de planilha .xlsx não instalado. Rode pip install -r requirements.txt e reinicie o backend."
            ) from exc

        workbook = load_workbook(BytesIO(conteudo), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        data_inicio_referencia = None
        data_referencia = None
        servico = None
        linhas = []

        for row in rows[:12]:
            texto_linha = " ".join(_texto_planilha(valor) or "" for valor in row)
            inicio_linha, fim_linha = _extrair_periodo_referencia_sisa(texto_linha)
            if inicio_linha and fim_linha and not data_referencia:
                data_inicio_referencia = inicio_linha
                data_referencia = fim_linha
            if "Tipo de Servi" in texto_linha or "SIAT" in texto_linha:
                servico = texto_linha.strip()

        class SheetAdapter:
            ncols = 16

            def cell_value(self, row_index, col_index):
                try:
                    return rows[row_index][col_index]
                except IndexError:
                    return None

        adapter = SheetAdapter()
        for indice in range(10, len(rows)):
            registro = _linha_planilha_para_registro_sisa(adapter, indice)
            if registro:
                linhas.append(registro)

        return {
            "data_inicio_referencia": data_inicio_referencia or data_referencia,
            "data_referencia": data_referencia,
            "servico": servico,
            "linhas": linhas,
        }

    raise HTTPException(
        status_code=400,
        detail="Envie uma planilha .xls ou .xlsx exportada do SISA."
    )


def _normalizar_chave_nome_sisa(nome: str | None):
    texto = unicodedata.normalize("NFKD", str(nome or ""))
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^a-zA-Z0-9]+", " ", texto).strip().lower()
    return re.sub(r"\s+", " ", texto)


def _linha_previa_sisa(linha: dict, **extras):
    return {
        "numero_sisa": linha["numero_sisa"],
        "nome": linha["nome_planilha"],
        "nome_social": linha["nome_social_planilha"],
        "data_nascimento": linha["data_nascimento"],
        "data_vinculacao": linha["data_vinculacao"],
        "data_desligamento": linha["data_desligamento"],
        "dias_permanencia": linha["dias_permanencia"],
        **extras,
    }


async def _montar_contexto_previa_sisa(
    db: AsyncSession,
    *,
    instituicao_id: str,
    linhas: list[dict],
    data_referencia: date,
):
    conviventes = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.instituicao_id == instituicao_id,
            )
        )
    ).scalars().all()
    conviventes_por_sisa = {}
    for convivente in conviventes:
        if convivente.numero_sisa:
            conviventes_por_sisa.setdefault(str(convivente.numero_sisa), []).append(convivente)
    conviventes_por_nome_data = {}
    for convivente in conviventes:
        chave = (_normalizar_chave_nome_sisa(convivente.nome_completo), convivente.data_nascimento)
        if chave[0] and chave[1]:
            conviventes_por_nome_data.setdefault(chave, []).append(convivente)

    numeros_sisa_planilha = {linha["numero_sisa"] for linha in linhas}
    criar_ativos = []
    criar_inativos = []
    possiveis_duplicidades = []
    reativar_existentes = []
    inativar_existentes = []
    vinculados = 0

    for linha in linhas:
        conviventes_sisa = conviventes_por_sisa.get(linha["numero_sisa"], [])
        desligamento_no_recorte = _desligamento_valido_no_recorte(
            linha["data_desligamento"],
            data_referencia,
        )
        if conviventes_sisa:
            vinculados += 1
            for convivente in conviventes_sisa:
                if not desligamento_no_recorte and convivente.status != "Ativo":
                    reativar_existentes.append(_linha_previa_sisa(
                        linha,
                        status_sugerido="Ativo",
                        convivente_id=convivente.id,
                        convivente_nome=convivente.nome_social or convivente.nome_completo,
                        motivo="Aparece no SISA como ativo neste recorte, mas está inativo no CareCore+.",
                    ))
                elif desligamento_no_recorte and convivente.status == "Ativo":
                    inativar_existentes.append(_linha_previa_sisa(
                        linha,
                        status_sugerido="Inativado",
                        convivente_id=convivente.id,
                        convivente_nome=convivente.nome_social or convivente.nome_completo,
                        motivo="Aparece no SISA com desligamento já vigente neste recorte.",
                    ))
            continue

        chave = (_normalizar_chave_nome_sisa(linha["nome_planilha"]), linha["data_nascimento"])
        duplicidades = conviventes_por_nome_data.get(chave, [])
        if duplicidades:
            nomes = ", ".join(
                duplicidade.nome_social or duplicidade.nome_completo
                for duplicidade in duplicidades[:3]
            )
            duplicidade_unica = duplicidades[0] if len(duplicidades) == 1 else None
            possiveis_duplicidades.append(_linha_previa_sisa(
                linha,
                status_sugerido="Mesclar" if duplicidade_unica else "Revisar",
                convivente_id=duplicidade_unica.id if duplicidade_unica else None,
                convivente_nome=(
                    duplicidade_unica.nome_social or duplicidade_unica.nome_completo
                    if duplicidade_unica else None
                ),
                motivo=(
                    f"Nome e nascimento parecidos com cadastro existente: {nomes}. "
                    "Ao mesclar, o número SISA da planilha substituirá o número cadastrado no CareCore+."
                    if duplicidade_unica
                    else f"Nome e nascimento parecidos com mais de um cadastro existente: {nomes}."
                ),
            ))
            continue

        if desligamento_no_recorte:
            criar_inativos.append(_linha_previa_sisa(
                linha,
                status_sugerido="Inativado",
                motivo="Não existe no CareCore+ e já possui desligamento vigente neste recorte SISA.",
            ))
        else:
            criar_ativos.append(_linha_previa_sisa(
                linha,
                status_sugerido="Ativo",
                motivo="Não existe no CareCore+ e está ativo no recorte SISA.",
            ))

    for convivente in conviventes:
        if (
            (
                not convivente.numero_sisa
                or str(convivente.numero_sisa) not in numeros_sisa_planilha
            )
            and convivente.status == "Ativo"
        ):
            inativar_existentes.append({
                "numero_sisa": str(convivente.numero_sisa or f"SEM-SISA-{convivente.id}"),
                "nome": convivente.nome_social or convivente.nome_completo,
                "nome_social": convivente.nome_social,
                "data_nascimento": convivente.data_nascimento,
                "data_vinculacao": None,
                "data_desligamento": None,
                "dias_permanencia": 0,
                "status_sugerido": "Inativado",
                "convivente_id": convivente.id,
                "convivente_nome": convivente.nome_social or convivente.nome_completo,
                "motivo": (
                    "Convivente ativo no CareCore+ sem número SISA. Confirme se deve permanecer ativo "
                    "ou se deve ser inativado por não ter vínculo na planilha importada."
                    if not convivente.numero_sisa
                    else "Convivente ativo no CareCore+ não apareceu na planilha SISA importada."
                ),
            })

    return {
        "conviventes_por_sisa": conviventes_por_sisa,
        "criar_ativos": criar_ativos,
        "criar_inativos": criar_inativos,
        "possiveis_duplicidades": possiveis_duplicidades,
        "inativar_existentes": inativar_existentes,
        "reativar_existentes": reativar_existentes,
        "vinculados": vinculados,
    }


def _parse_acoes_importacao_sisa(acoes_json: str | None) -> SisaAcoesImportacao:
    if not acoes_json:
        return SisaAcoesImportacao()

    try:
        dados = json.loads(acoes_json)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Ações da importação SISA inválidas.") from exc

    return SisaAcoesImportacao(**dados)


async def _proximo_numero_institucional(db: AsyncSession, instituicao_id: str) -> int:
    maior_numero = (
        await db.execute(
            select(func.max(ConviventeDB.numero_institucional)).where(
                ConviventeDB.instituicao_id == instituicao_id
            )
        )
    ).scalar()
    return (maior_numero or 0) + 1


async def _registrar_historico_importacao_sisa(
    db: AsyncSession,
    *,
    instituicao_id: str,
    convivente_id: str,
    usuario_id: str,
    data_referencia: date,
    titulo: str,
    descricao: str,
):
    db.add(HistoricoConviventeDB(
        instituicao_id=instituicao_id,
        convivente_id=convivente_id,
        usuario_id=usuario_id,
        origem_informacao="Importação SISA",
        data_origem=data_referencia,
        titulo=titulo,
        descricao=descricao,
    ))


async def _aplicar_acoes_cadastros_sisa(
    db: AsyncSession,
    *,
    instituicao_id: str,
    usuario_id: str,
    linhas: list[dict],
    data_inicio_referencia: date,
    data_referencia: date,
    acoes: SisaAcoesImportacao,
):
    linhas_por_numero = {linha["numero_sisa"]: linha for linha in linhas}
    criar_ativos = set(acoes.criar_ativos or [])
    criar_inativos = set(acoes.criar_inativos or [])

    for chave_mesclagem in set(acoes.mesclar_duplicidades or []):
        numero_sisa, _, convivente_id = str(chave_mesclagem).partition(":")
        if not numero_sisa or not convivente_id:
            continue

        linha = linhas_por_numero.get(numero_sisa)
        if not linha:
            continue

        convivente = (
            await db.execute(
                select(ConviventeDB).where(
                    ConviventeDB.instituicao_id == instituicao_id,
                    ConviventeDB.id == convivente_id,
                )
            )
        ).scalar_one_or_none()
        if not convivente or convivente.numero_sisa:
            if convivente and convivente.numero_sisa and str(convivente.numero_sisa) != numero_sisa:
                sisa_anterior = convivente.numero_sisa
                convivente.numero_sisa = numero_sisa
                reativou = False
                if _linha_sisa_ativa_no_recorte(linha, data_referencia) and convivente.status != "Ativo":
                    convivente.status = "Ativo"
                    convivente.inativado_em = None
                    reativou = True
                if not convivente.data_nascimento:
                    convivente.data_nascimento = linha["data_nascimento"]
                if not convivente.data_entrada:
                    convivente.data_entrada = linha["data_vinculacao"]
                if not convivente.nome_social:
                    convivente.nome_social = linha["nome_social_planilha"]

                await _registrar_historico_importacao_sisa(
                    db,
                    instituicao_id=instituicao_id,
                    convivente_id=convivente.id,
                    usuario_id=usuario_id,
                    data_referencia=data_referencia,
                    titulo="Número SISA corrigido pela importação",
                    descricao=(
                        f"Número SISA anterior {sisa_anterior} substituído por {numero_sisa} "
                        "após confirmação manual na prévia da importação. A planilha SISA "
                        "foi considerada a fonte oficial do número."
                        + (" Cadastro também reativado por estar ativo no recorte SISA." if reativou else "")
                    ),
                )
            continue

        convivente.numero_sisa = numero_sisa
        if _linha_sisa_ativa_no_recorte(linha, data_referencia) and convivente.status != "Ativo":
            convivente.status = "Ativo"
            convivente.inativado_em = None
        if not convivente.data_nascimento:
            convivente.data_nascimento = linha["data_nascimento"]
        if not convivente.data_entrada:
            convivente.data_entrada = linha["data_vinculacao"]
        if not convivente.nome_social:
            convivente.nome_social = linha["nome_social_planilha"]

        await _registrar_historico_importacao_sisa(
            db,
            instituicao_id=instituicao_id,
            convivente_id=convivente.id,
            usuario_id=usuario_id,
            data_referencia=data_referencia,
            titulo="Cadastro mesclado pela importação SISA",
            descricao=(
                f"Número SISA {numero_sisa} vinculado ao cadastro existente após confirmação manual "
                f"na prévia da importação do recorte {data_inicio_referencia.strftime('%d/%m/%Y')} "
                f"a {data_referencia.strftime('%d/%m/%Y')}."
            ),
        )

    for numero_sisa in list(criar_ativos | criar_inativos):
        linha = linhas_por_numero.get(numero_sisa)
        if not linha:
            continue

        existente = (
            await db.execute(
                select(ConviventeDB.id).where(
                    ConviventeDB.instituicao_id == instituicao_id,
                    ConviventeDB.numero_sisa == numero_sisa,
                )
            )
        ).scalar_one_or_none()
        if existente:
            continue

        status = "Inativado" if numero_sisa in criar_inativos else "Ativo"
        novo_convivente = ConviventeDB(
            instituicao_id=instituicao_id,
            numero_institucional=await _proximo_numero_institucional(db, instituicao_id),
            status=status,
            inativado_em=agora_sao_paulo() if status == "Inativado" else None,
            data_entrada=linha["data_vinculacao"],
            nome_completo=linha["nome_planilha"],
            nome_social=linha["nome_social_planilha"],
            data_nascimento=linha["data_nascimento"],
            identidade_genero=linha["sexo"],
            numero_sisa=linha["numero_sisa"],
        )
        db.add(novo_convivente)
        await db.flush()

        await _registrar_historico_importacao_sisa(
            db,
            instituicao_id=instituicao_id,
            convivente_id=novo_convivente.id,
            usuario_id=usuario_id,
            data_referencia=data_referencia,
            titulo="Cadastro criado pela importação SISA",
            descricao=(
                f"Cadastro criado automaticamente a partir da planilha SISA "
                f"do recorte {data_inicio_referencia.strftime('%d/%m/%Y')} a "
                f"{data_referencia.strftime('%d/%m/%Y')}. Status inicial: {status}."
            ),
        )

    if acoes.inativar_existentes:
        conviventes_para_inativar = (
            await db.execute(
                select(ConviventeDB).where(
                    ConviventeDB.instituicao_id == instituicao_id,
                    ConviventeDB.id.in_(acoes.inativar_existentes),
                    ConviventeDB.status == "Ativo",
                )
            )
        ).scalars().all()

        for convivente in conviventes_para_inativar:
            convivente.status = "Inativado"
            convivente.inativado_em = agora_sao_paulo()
            await _registrar_historico_importacao_sisa(
                db,
                instituicao_id=instituicao_id,
                convivente_id=convivente.id,
                usuario_id=usuario_id,
                data_referencia=data_referencia,
                titulo="Inativação sugerida pela importação SISA",
                descricao=(
                    "Convivente inativado após confirmação manual porque não apareceu "
                    f"na planilha SISA do recorte {data_inicio_referencia.strftime('%d/%m/%Y')} "
                    f"a {data_referencia.strftime('%d/%m/%Y')}."
                ),
            )

    if acoes.reativar_existentes:
        conviventes_para_reativar = (
            await db.execute(
                select(ConviventeDB).where(
                    ConviventeDB.instituicao_id == instituicao_id,
                    ConviventeDB.id.in_(acoes.reativar_existentes),
                    ConviventeDB.status != "Ativo",
                )
            )
        ).scalars().all()

        for convivente in conviventes_para_reativar:
            convivente.status = "Ativo"
            convivente.inativado_em = None
            await _registrar_historico_importacao_sisa(
                db,
                instituicao_id=instituicao_id,
                convivente_id=convivente.id,
                usuario_id=usuario_id,
                data_referencia=data_referencia,
                titulo="Reativação sugerida pela importação SISA",
                descricao=(
                    "Convivente reativado após confirmação manual porque apareceu "
                    f"sem desligamento na planilha SISA do recorte {data_inicio_referencia.strftime('%d/%m/%Y')} "
                    f"a {data_referencia.strftime('%d/%m/%Y')}."
                ),
            )


def _resumo_registros_carecore(registros):
    resumo = {
        "entradas": 0,
        "saidas": 0,
        "cafes": 0,
        "almocos": 0,
        "jantares": 0,
        "lanches": 0,
        "total_interacoes": len(registros),
    }

    for registro in registros:
        if registro.tipo_registro == "Entrada":
            resumo["entradas"] += 1
        elif registro.tipo_registro == "Saída":
            resumo["saidas"] += 1
        elif registro.tipo_registro == "Café da manhã":
            resumo["cafes"] += 1
        elif registro.tipo_registro == "Almoço":
            resumo["almocos"] += 1
        elif registro.tipo_registro == "Jantar":
            resumo["jantares"] += 1
        elif registro.tipo_registro == "Lanche noturno":
            resumo["lanches"] += 1

    return resumo


STATUS_TRATATIVA_SISA = {"Pendente", "Em análise", "Justificado", "Resolvido"}


def _serializar_divergencia_sisa(divergencia, status_convivente=None, tem_desligamento=None):
    dados = {
        coluna.name: getattr(divergencia, coluna.name)
        for coluna in divergencia.__table__.columns
    }
    dados["status_convivente"] = status_convivente
    dados["tem_desligamento"] = tem_desligamento
    return dados


def _ordenacao_divergencias_sisa():
    return (
        case(
            (SisaDivergenciaDB.prioridade == "Crítica", 0),
            (SisaDivergenciaDB.prioridade == "Alta", 1),
            (SisaDivergenciaDB.prioridade == "Média", 2),
            else_=3,
        ),
        SisaDivergenciaDB.nome_convivente.asc(),
    )


def _aplicar_filtros_divergencias_sisa(
    query,
    *,
    busca: str | None = None,
    tipo: str = "todos",
    prioridade: str = "todas",
    status_tratativa: str = "todos",
    status_convivente: str = "todos",
    desligamento: str = "todos",
    somente_diferenca: bool = False,
    dif_minima: int = 0,
    importacao_id: str,
    instituicao_id: str,
):
    if tipo and tipo != "todos":
        query = query.where(SisaDivergenciaDB.tipo == tipo)

    if prioridade and prioridade != "todas":
        query = query.where(SisaDivergenciaDB.prioridade == prioridade)

    if status_tratativa and status_tratativa != "todos":
        query = query.where(SisaDivergenciaDB.status == status_tratativa)

    if status_convivente == "sem_cadastro":
        query = query.where(SisaDivergenciaDB.convivente_id.is_(None))
    elif status_convivente and status_convivente != "todos":
        query = query.join(
            ConviventeDB,
            and_(
                ConviventeDB.id == SisaDivergenciaDB.convivente_id,
                ConviventeDB.instituicao_id == instituicao_id,
            ),
        ).where(ConviventeDB.status == status_convivente)

    if desligamento in {"com", "sem"}:
        presenca_com_desligamento = exists(
            select(SisaPresencaImportadaDB.id).where(
                SisaPresencaImportadaDB.importacao_id == importacao_id,
                SisaPresencaImportadaDB.instituicao_id == instituicao_id,
                SisaPresencaImportadaDB.numero_sisa == SisaDivergenciaDB.numero_sisa,
                SisaPresencaImportadaDB.data_desligamento.is_not(None),
            )
        )
        if desligamento == "com":
            query = query.where(presenca_com_desligamento)
        else:
            query = query.where(~presenca_com_desligamento)

    if somente_diferenca:
        query = query.where(SisaDivergenciaDB.diferenca.is_not(None), SisaDivergenciaDB.diferenca != 0)

    if dif_minima and dif_minima > 0:
        query = query.where(func.abs(SisaDivergenciaDB.diferenca) >= dif_minima)

    busca_limpa = (busca or "").strip()
    if busca_limpa:
        termo = f"%{busca_limpa.lower()}%"
        query = query.where(
            or_(
                func.lower(SisaDivergenciaDB.nome_convivente).like(termo),
                func.lower(SisaDivergenciaDB.numero_sisa).like(termo),
            )
        )

    return query


async def _serializar_divergencias_sisa(
    db: AsyncSession,
    *,
    instituicao_id: str,
    importacao_id: str,
    divergencias,
):
    convivente_ids = {d.convivente_id for d in divergencias if d.convivente_id}
    status_por_convivente = {}
    if convivente_ids:
        conviventes = (
            await db.execute(
                select(ConviventeDB.id, ConviventeDB.status).where(
                    ConviventeDB.instituicao_id == instituicao_id,
                    ConviventeDB.id.in_(convivente_ids),
                )
            )
        ).all()
        status_por_convivente = {cid: status for cid, status in conviventes}

    numeros_sisa = {d.numero_sisa for d in divergencias if d.numero_sisa}
    desligamento_por_numero = {}
    if numeros_sisa:
        presencas = (
            await db.execute(
                select(
                    SisaPresencaImportadaDB.numero_sisa,
                    SisaPresencaImportadaDB.data_desligamento,
                ).where(
                    SisaPresencaImportadaDB.importacao_id == importacao_id,
                    SisaPresencaImportadaDB.instituicao_id == instituicao_id,
                    SisaPresencaImportadaDB.numero_sisa.in_(numeros_sisa),
                )
            )
        ).all()
        desligamento_por_numero = {
            numero: bool(desligamento) for numero, desligamento in presencas
        }

    return [
        _serializar_divergencia_sisa(
            divergencia,
            status_convivente=status_por_convivente.get(divergencia.convivente_id),
            tem_desligamento=desligamento_por_numero.get(divergencia.numero_sisa, False),
        )
        for divergencia in divergencias
    ]


async def _resumo_divergencias_sisa_filtradas(db: AsyncSession, query_filtrada):
    total = int((await db.execute(
        select(func.count()).select_from(query_filtrada.subquery())
    )).scalar_one() or 0)

    sub = query_filtrada.subquery()
    pendencias = int((await db.execute(
        select(func.count()).select_from(sub).where(sub.c.tipo.notin_(["OK", "SEM_BASE_ANTERIOR"]))
    )).scalar_one() or 0)

    alertas_criticos = int((await db.execute(
        select(func.count()).select_from(sub).where(
            or_(sub.c.tipo == "SISA_MENOR", sub.c.prioridade == "Crítica")
        )
    )).scalar_one() or 0)

    dias_perdidos_filtrados = int((await db.execute(
        select(
            func.coalesce(
                func.sum(
                    case(
                        (sub.c.diferenca < 0, func.abs(sub.c.diferenca)),
                        else_=0,
                    )
                ),
                0,
            )
        ).select_from(sub)
    )).scalar_one() or 0)

    return {
        "total": total,
        "pendencias": pendencias,
        "alertas_criticos": alertas_criticos,
        "dias_perdidos_filtrados": dias_perdidos_filtrados,
    }


async def _carregar_importacao_sisa(db: AsyncSession, importacao_id: str, instituicao_id: str):
    return (
        await db.execute(
            select(SisaImportacaoDB).where(
                SisaImportacaoDB.id == importacao_id,
                SisaImportacaoDB.instituicao_id == instituicao_id,
            )
        )
    ).scalar_one_or_none()


async def _criar_divergencia_sisa(
    db: AsyncSession,
    *,
    importacao: SisaImportacaoDB,
    presenca: SisaPresencaImportadaDB,
    convivente: ConviventeDB | None,
    data_inicio_referencia: date,
    data_referencia: date,
):
    dias_sisa_periodo = _dias_sisa_no_recorte(
        data_inicio_recorte=data_inicio_referencia,
        data_fim_recorte=data_referencia,
        data_vinculacao=presenca.data_vinculacao,
        data_desligamento=presenca.data_desligamento,
        dias_permanencia_acumulado=presenca.dias_permanencia,
    )

    if not convivente:
        return SisaDivergenciaDB(
            importacao_id=importacao.id,
            instituicao_id=importacao.instituicao_id,
            convivente_id=None,
            numero_sisa=presenca.numero_sisa,
            nome_convivente=presenca.nome_planilha,
            tipo="CONVIVENTE_NAO_ENCONTRADO",
            prioridade="Alta",
            data_inicio=data_inicio_referencia,
            data_fim=data_referencia,
            dias_sisa_atual=presenca.dias_permanencia,
            dias_sisa_delta=dias_sisa_periodo,
            dias_carecore=0,
            mensagem="Código SISA não encontrado no cadastro de conviventes.",
        )

    anterior = (
        await db.execute(
            select(SisaPresencaImportadaDB)
            .where(
                SisaPresencaImportadaDB.instituicao_id == importacao.instituicao_id,
                SisaPresencaImportadaDB.convivente_id == convivente.id,
                SisaPresencaImportadaDB.data_referencia < data_referencia,
            )
            .order_by(SisaPresencaImportadaDB.data_referencia.desc())
        )
    ).scalars().first()

    inicio_periodo = datetime.combine(data_inicio_referencia, datetime.min.time())
    fim_periodo = datetime.combine(data_referencia, datetime.max.time())

    registros = (
        await db.execute(
            select(RegistroRotinaDB)
            .where(
                RegistroRotinaDB.instituicao_id == importacao.instituicao_id,
                RegistroRotinaDB.convivente_id == convivente.id,
                RegistroRotinaDB.cancelado != True,
                RegistroRotinaDB.data_registro >= inicio_periodo,
                RegistroRotinaDB.data_registro <= fim_periodo,
                RegistroRotinaDB.tipo_registro.in_(TIPOS_SISA_PRESENCA_CARECORE),
            )
            .order_by(RegistroRotinaDB.data_registro.asc())
        )
    ).scalars().all()

    dias_carecore_lista = sorted({
        registro.data_registro.date().isoformat()
        for registro in registros
    })
    dias_carecore = len(dias_carecore_lista)
    dias_sisa_delta = dias_sisa_periodo
    diferenca = dias_sisa_delta - dias_carecore

    if diferenca == 0:
        tipo = "OK"
        prioridade = "Baixa"
        mensagem = "CareCore+ e SISA estão compatíveis no período."
    elif diferenca < 0:
        tipo = "SISA_MENOR"
        prioridade = "Crítica"
        mensagem = (
            "CareCore+ tem mais dias com entrada/saída/alimentação do que o SISA reconheceu. "
            "Conferir possível perda de repasse."
        )
    else:
        tipo = "SISA_MAIOR"
        prioridade = "Média"
        mensagem = "SISA reconheceu mais dias do que os registros encontrados no CareCore+."

    if not anterior:
        mensagem = (
            f"{mensagem} Primeira importação deste convivente; "
            "dias do SISA calculados pelo recorte do relatório."
        )

    return SisaDivergenciaDB(
        importacao_id=importacao.id,
        instituicao_id=importacao.instituicao_id,
        convivente_id=convivente.id,
        numero_sisa=presenca.numero_sisa,
        nome_convivente=_nome_convivente_relatorio(convivente),
        tipo=tipo,
        prioridade=prioridade,
        data_inicio=data_inicio_referencia,
        data_fim=data_referencia,
        dias_sisa_anterior=anterior.dias_permanencia if anterior else None,
        dias_sisa_atual=presenca.dias_permanencia,
        dias_sisa_delta=dias_sisa_delta,
        dias_carecore=dias_carecore,
        diferenca=diferenca,
        dias_carecore_lista=json.dumps(dias_carecore_lista, ensure_ascii=False),
        resumo_carecore_json=json.dumps(_resumo_registros_carecore(registros), ensure_ascii=False),
        mensagem=mensagem,
    )


@router.get("/relatorios/pia", response_model=RelatorioPiaListaResponse)
async def listar_registros_pia_relatorios(
    limite: int = Query(200, ge=1, le=EXPORT_ROTINA_HISTORICO_LIMITE_MAX),
    deslocamento: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    instituicao_id = obter_instituicao_escopo(usuario_atual)
    base = (
        select(RegistroPIADB)
        .join(ConviventeDB, ConviventeDB.id == RegistroPIADB.convivente_id)
        .where(
            RegistroPIADB.instituicao_id == instituicao_id,
            ConviventeDB.instituicao_id == instituicao_id,
        )
    )
    total = int((await db.execute(
        select(func.count()).select_from(base.subquery())
    )).scalar_one() or 0)

    registros = (
        await db.execute(
            select(
                RegistroPIADB,
                ConviventeDB.nome_completo.label("convivente_nome_completo"),
                ConviventeDB.nome_social.label("convivente_nome_social"),
                ConviventeDB.numero_institucional.label("convivente_numero_institucional"),
                ConviventeDB.status.label("convivente_status"),
                ConviventeDB.tecnico_id.label("convivente_tecnico_id"),
                UsuarioDB.nome.label("usuario_nome"),
            )
            .join(ConviventeDB, ConviventeDB.id == RegistroPIADB.convivente_id)
            .join(UsuarioDB, UsuarioDB.id == RegistroPIADB.usuario_id)
            .where(
                RegistroPIADB.instituicao_id == instituicao_id,
                ConviventeDB.instituicao_id == instituicao_id,
            )
            .order_by(RegistroPIADB.data_registro.desc())
            .offset(deslocamento)
            .limit(limite)
        )
    ).all()

    items = [
        {
            **{coluna.name: getattr(registro, coluna.name) for coluna in registro.__table__.columns},
            "convivente_nome_completo": convivente_nome_completo,
            "convivente_nome_social": convivente_nome_social,
            "convivente_numero_institucional": convivente_numero_institucional,
            "convivente_status": convivente_status,
            "convivente_tecnico_id": convivente_tecnico_id,
            "usuario_nome": usuario_nome,
        }
        for (
            registro,
            convivente_nome_completo,
            convivente_nome_social,
            convivente_numero_institucional,
            convivente_status,
            convivente_tecnico_id,
            usuario_nome,
        ) in registros
    ]

    return {
        "registros": items,
        "total": total,
        "limite": limite,
        "deslocamento": deslocamento,
        "has_more": deslocamento + len(items) < total,
    }


@router.get("/convenio-sisa/importacoes", response_model=SisaImportacaoListaResponse)
async def listar_importacoes_sisa(
    limit: int = Query(REGISTROS_POR_PAGINA_PADRAO, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    bloquear_usuario_global_puro(usuario_atual)
    instituicao_id = obter_instituicao_escopo(usuario_atual)
    base = select(SisaImportacaoDB).where(
        SisaImportacaoDB.instituicao_id == instituicao_id
    )
    total = int((await db.execute(
        select(func.count()).select_from(base.subquery())
    )).scalar_one() or 0)

    items = (
        await db.execute(
            base.order_by(SisaImportacaoDB.data_referencia.desc(), SisaImportacaoDB.importado_em.desc())
            .offset(offset)
            .limit(limit)
        )
    ).scalars().all()

    return {
        "items": items,
        "total": total,
        "limit": limit,
        "offset": offset,
        "has_more": offset + len(items) < total,
    }


@router.delete("/convenio-sisa/importacoes/{importacao_id}")
async def excluir_importacao_sisa(
    importacao_id: str,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    bloquear_usuario_global_puro(usuario_atual)

    if not usuario_pode_excluir_importacao_sisa(usuario_atual):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Apenas Gestores ou Manutenção podem excluir importações SISA.",
        )

    instituicao_id = obter_instituicao_escopo(usuario_atual)
    importacao = (
        await db.execute(
            select(SisaImportacaoDB).where(
                SisaImportacaoDB.id == importacao_id,
                SisaImportacaoDB.instituicao_id == instituicao_id,
            )
        )
    ).scalar_one_or_none()

    if not importacao:
        raise HTTPException(status_code=404, detail="Importação SISA não encontrada.")

    resumo_auditoria = {
        "importacao_id": importacao.id,
        "nome_arquivo": importacao.nome_arquivo,
        "data_referencia": importacao.data_referencia.isoformat(),
        "total_linhas": importacao.total_linhas,
    }

    await db.execute(
        delete(SisaDivergenciaDB).where(
            SisaDivergenciaDB.importacao_id == importacao_id,
            SisaDivergenciaDB.instituicao_id == instituicao_id,
        )
    )
    await db.execute(
        delete(SisaPresencaImportadaDB).where(
            SisaPresencaImportadaDB.importacao_id == importacao_id,
            SisaPresencaImportadaDB.instituicao_id == instituicao_id,
        )
    )
    await db.delete(importacao)
    await db.commit()

    registrar_evento_auditoria(
        "sisa_importacao_excluida",
        usuario_atual=usuario_atual,
        **resumo_auditoria,
    )

    return {"status": "sucesso", "mensagem": "Importação SISA excluída com sucesso."}


@router.get("/convenio-sisa/importacoes/{importacao_id}", response_model=SisaImportacaoDetalheResponse)
async def detalhar_importacao_sisa(
    importacao_id: str,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    instituicao_id = obter_instituicao_escopo(usuario_atual)
    importacao = await _carregar_importacao_sisa(db, importacao_id, instituicao_id)

    if not importacao:
        raise HTTPException(status_code=404, detail="Importação SISA não encontrada.")

    return {
        **{coluna.name: getattr(importacao, coluna.name) for coluna in importacao.__table__.columns},
        "divergencias": [],
    }


@router.get(
    "/convenio-sisa/importacoes/{importacao_id}/divergencias",
    response_model=SisaDivergenciaListaResponse,
)
async def listar_divergencias_importacao_sisa(
    importacao_id: str,
    limit: int = Query(REGISTROS_POR_PAGINA_PADRAO, ge=1, le=200),
    offset: int = Query(0, ge=0),
    busca: str | None = Query(None),
    tipo: str = Query("todos"),
    prioridade: str = Query("todas"),
    status_tratativa: str = Query("todos"),
    status_convivente: str = Query("todos"),
    desligamento: str = Query("todos"),
    somente_diferenca: bool = Query(False),
    dif_minima: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    bloquear_usuario_global_puro(usuario_atual)
    instituicao_id = obter_instituicao_escopo(usuario_atual)
    importacao = await _carregar_importacao_sisa(db, importacao_id, instituicao_id)

    if not importacao:
        raise HTTPException(status_code=404, detail="Importação SISA não encontrada.")

    query = select(SisaDivergenciaDB).where(
        SisaDivergenciaDB.importacao_id == importacao_id,
        SisaDivergenciaDB.instituicao_id == instituicao_id,
    )
    query = _aplicar_filtros_divergencias_sisa(
        query,
        busca=busca,
        tipo=tipo,
        prioridade=prioridade,
        status_tratativa=status_tratativa,
        status_convivente=status_convivente,
        desligamento=desligamento,
        somente_diferenca=somente_diferenca,
        dif_minima=dif_minima,
        importacao_id=importacao_id,
        instituicao_id=instituicao_id,
    )

    resumo = await _resumo_divergencias_sisa_filtradas(db, query)
    total = resumo["total"]

    divergencias = (
        await db.execute(
            query.order_by(*_ordenacao_divergencias_sisa()).offset(offset).limit(limit)
        )
    ).scalars().all()

    divergencias_serializadas = await _serializar_divergencias_sisa(
        db,
        instituicao_id=instituicao_id,
        importacao_id=importacao_id,
        divergencias=divergencias,
    )

    return {
        "items": divergencias_serializadas,
        "total": total,
        "limit": limit,
        "offset": offset,
        "has_more": offset + len(divergencias_serializadas) < total,
        "resumo": {
            "pendencias": resumo["pendencias"],
            "alertas_criticos": resumo["alertas_criticos"],
            "dias_perdidos_filtrados": resumo["dias_perdidos_filtrados"],
        },
    }


@router.patch(
    "/convenio-sisa/divergencias/{divergencia_id}",
    response_model=SisaDivergenciaResponse,
)
async def atualizar_status_divergencia_sisa(
    divergencia_id: str,
    payload: SisaDivergenciaStatusUpdate,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    bloquear_usuario_global_puro(usuario_atual)
    if payload.status not in STATUS_TRATATIVA_SISA:
        raise HTTPException(
            status_code=400,
            detail=f"Status de tratativa inválido. Use um destes: {', '.join(sorted(STATUS_TRATATIVA_SISA))}.",
        )

    divergencia = (
        await db.execute(
            select(SisaDivergenciaDB).where(
                SisaDivergenciaDB.id == divergencia_id,
                SisaDivergenciaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if not divergencia:
        raise HTTPException(status_code=404, detail="Divergência SISA não encontrada.")

    divergencia.status = payload.status
    await db.commit()
    await db.refresh(divergencia)

    status_convivente = None
    if divergencia.convivente_id:
        status_convivente = (
            await db.execute(
                select(ConviventeDB.status).where(
                    ConviventeDB.id == divergencia.convivente_id,
                    ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                )
            )
        ).scalar_one_or_none()

    tem_desligamento = (
        await db.execute(
            select(SisaPresencaImportadaDB.data_desligamento).where(
                SisaPresencaImportadaDB.importacao_id == divergencia.importacao_id,
                SisaPresencaImportadaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                SisaPresencaImportadaDB.numero_sisa == divergencia.numero_sisa,
            )
        )
    ).scalars().first()

    return _serializar_divergencia_sisa(
        divergencia,
        status_convivente=status_convivente,
        tem_desligamento=bool(tem_desligamento),
    )


@router.post("/convenio-sisa/importacoes/previsualizar", response_model=SisaPreviaImportacaoResponse)
async def previsualizar_importacao_sisa(
    arquivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    bloquear_usuario_global_puro(usuario_atual)
    conteudo = await arquivo.read()
    dados = _ler_planilha_sisa(conteudo, arquivo.filename or "")
    data_inicio_referencia = dados["data_inicio_referencia"]
    data_referencia = dados["data_referencia"]
    linhas = dados["linhas"]

    if not data_inicio_referencia or not data_referencia:
        raise HTTPException(status_code=400, detail="Não foi possível identificar o recorte de referência da planilha SISA.")

    if not linhas:
        raise HTTPException(status_code=400, detail="Nenhum cidadão foi encontrado na planilha SISA.")

    linhas_cadastros = _consolidar_linhas_sisa_por_ultimo_movimento(linhas, data_referencia)
    contexto = await _montar_contexto_previa_sisa(
        db,
        instituicao_id=obter_instituicao_escopo(usuario_atual),
        linhas=linhas_cadastros,
        data_referencia=data_referencia,
    )

    return {
        "nome_arquivo": arquivo.filename or "planilha_sisa.xls",
        "servico": dados["servico"],
        "data_inicio_referencia": data_inicio_referencia,
        "data_referencia": data_referencia,
        "total_linhas": len(linhas),
        "vinculados": contexto["vinculados"],
        "criar_ativos": contexto["criar_ativos"],
        "criar_inativos": contexto["criar_inativos"],
        "possiveis_duplicidades": contexto["possiveis_duplicidades"],
        "inativar_existentes": contexto["inativar_existentes"],
        "reativar_existentes": contexto["reativar_existentes"],
    }


@router.post("/convenio-sisa/importacoes", response_model=SisaImportacaoDetalheResponse)
async def importar_planilha_sisa(
    arquivo: UploadFile = File(...),
    acoes_json: str | None = Form(None),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    conteudo = await arquivo.read()
    dados = _ler_planilha_sisa(conteudo, arquivo.filename or "")
    data_inicio_referencia = dados["data_inicio_referencia"]
    data_referencia = dados["data_referencia"]
    linhas = dados["linhas"]

    if not data_inicio_referencia or not data_referencia:
        raise HTTPException(status_code=400, detail="Não foi possível identificar o recorte de referência da planilha SISA.")

    if not linhas:
        raise HTTPException(status_code=400, detail="Nenhum cidadão foi encontrado na planilha SISA.")

    instituicao_id = obter_instituicao_escopo(usuario_atual)
    acoes = _parse_acoes_importacao_sisa(acoes_json)
    await _aplicar_acoes_cadastros_sisa(
        db,
        instituicao_id=instituicao_id,
        usuario_id=usuario_atual["sub"],
        linhas=_consolidar_linhas_sisa_por_ultimo_movimento(linhas, data_referencia),
        data_inicio_referencia=data_inicio_referencia,
        data_referencia=data_referencia,
        acoes=acoes,
    )
    await db.flush()

    conviventes_resultado = await db.execute(
        select(ConviventeDB).where(
            ConviventeDB.instituicao_id == instituicao_id,
            ConviventeDB.numero_sisa.is_not(None),
        )
    )
    conviventes_por_sisa = {
        str(convivente.numero_sisa): convivente
        for convivente in conviventes_resultado.scalars().all()
    }

    importacao = SisaImportacaoDB(
        instituicao_id=instituicao_id,
        usuario_id=usuario_atual["sub"],
        nome_arquivo=arquivo.filename or "planilha_sisa.xls",
        servico=dados["servico"],
        data_referencia=data_referencia,
        importado_em=agora_sao_paulo(),
        total_linhas=len(linhas),
    )
    db.add(importacao)
    await db.flush()

    divergencias = []
    status_por_numero = {}
    desligamento_por_numero = {}
    total_vinculados = 0
    total_nao_encontrados = 0

    for linha in linhas:
        convivente = conviventes_por_sisa.get(linha["numero_sisa"])
        status_vinculo = "Vinculado" if convivente else "Não encontrado"
        status_por_numero[linha["numero_sisa"]] = convivente.status if convivente else None
        desligamento_por_numero[linha["numero_sisa"]] = bool(linha["data_desligamento"])

        if convivente:
            total_vinculados += 1
        else:
            total_nao_encontrados += 1

        presenca = SisaPresencaImportadaDB(
            importacao_id=importacao.id,
            instituicao_id=instituicao_id,
            convivente_id=convivente.id if convivente else None,
            numero_sisa=linha["numero_sisa"],
            nome_planilha=linha["nome_planilha"],
            nome_social_planilha=linha["nome_social_planilha"],
            data_referencia=data_referencia,
            data_nascimento=linha["data_nascimento"],
            sexo=linha["sexo"],
            data_vinculacao=linha["data_vinculacao"],
            data_desligamento=linha["data_desligamento"],
            dias_permanencia=linha["dias_permanencia"],
            status_vinculo=status_vinculo,
        )
        db.add(presenca)
        await db.flush()

        divergencia = await _criar_divergencia_sisa(
            db,
            importacao=importacao,
            presenca=presenca,
            convivente=convivente,
            data_inicio_referencia=data_inicio_referencia,
            data_referencia=data_referencia,
        )
        db.add(divergencia)
        divergencias.append(divergencia)

    importacao.total_vinculados = total_vinculados
    importacao.total_nao_encontrados = total_nao_encontrados
    importacao.total_divergencias = sum(
        1 for divergencia in divergencias
        if divergencia.tipo not in {"OK", "SEM_BASE_ANTERIOR"}
    )
    importacao.total_alertas_criticos = sum(
        1 for divergencia in divergencias
        if divergencia.tipo == "SISA_MENOR" or divergencia.prioridade == "Crítica"
    )

    await db.commit()
    await db.refresh(importacao)

    return {
        **{coluna.name: getattr(importacao, coluna.name) for coluna in importacao.__table__.columns},
        "divergencias": [],
    }


@router.get("/convenio-sisa/diario")
async def relatorio_sisa_diario(
    data: str = None,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    if not data:
        data = agora_sao_paulo().date().isoformat()

    inicio, fim = _inicio_fim_dia(data)

    data_referencia = inicio.date()
    conviventes = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                ConviventeDB.status.in_(["Ativo", "Ausência justificada"]),
            ).order_by(
                ConviventeDB.nome_completo.asc()
            )
        )
    ).scalars().all()

    registros_resultado = await db.execute(
        select(
            RegistroRotinaDB.convivente_id,
            RegistroRotinaDB.tipo_registro,
            RegistroRotinaDB.data_registro,
            RegistroRotinaDB.retorno_rapido
        ).where(
            RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            RegistroRotinaDB.cancelado != True,
            RegistroRotinaDB.data_registro >= inicio,
            RegistroRotinaDB.data_registro <= fim
        ).order_by(
            RegistroRotinaDB.data_registro.asc()
        )
    )

    registros_por_convivente = {}

    for convivente_id, tipo_registro, data_registro, retorno_rapido in registros_resultado.all():
        registros_por_convivente.setdefault(
            convivente_id,
            []
        ).append({
            "tipo_registro": tipo_registro,
            "data_registro": data_registro,
            "retorno_rapido": retorno_rapido
        })

    convivente_ids = [convivente.id for convivente in conviventes]
    movimentos_resultado = await db.execute(
        select(
            RegistroRotinaDB.convivente_id,
            RegistroRotinaDB.tipo_registro,
            RegistroRotinaDB.data_registro,
        ).where(
            RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            RegistroRotinaDB.cancelado != True,
            RegistroRotinaDB.convivente_id.in_(convivente_ids),
            RegistroRotinaDB.tipo_registro.in_(["Entrada", "Saída"]),
            RegistroRotinaDB.data_registro <= fim,
        ).order_by(
            RegistroRotinaDB.data_registro.asc()
        )
    )

    movimentos_por_convivente = {}
    for convivente_id, tipo_registro, data_registro in movimentos_resultado.all():
        movimentos_por_convivente.setdefault(
            convivente_id,
            []
        ).append({
            "tipo_registro": tipo_registro,
            "data_registro": data_registro,
        })

    linhas = []

    for convivente in conviventes:
        regs = registros_por_convivente.get(convivente.id, [])

        entradas = [
            r for r in regs
            if r["tipo_registro"] == "Entrada"
        ]

        saidas = [
            r for r in regs
            if r["tipo_registro"] == "Saída"
        ]

        almocos = [
            r for r in regs
            if r["tipo_registro"] == "Almoço"
        ]
        cafes = [
            r for r in regs
            if r["tipo_registro"] == "Café da manhã"
        ]
        jantares = [
            r for r in regs
            if r["tipo_registro"] == "Jantar"
        ]
        lanches = [
            r for r in regs
            if r["tipo_registro"] == "Lanche noturno"
        ]
        banhos = [
            r for r in regs
            if r["tipo_registro"] == "Banho"
        ]
        banhos = [
            r for r in regs
            if r["tipo_registro"] == "Banho"
        ]

        ausencia_justificada = (
            convivente.status == "Ausência justificada"
            and (
                convivente.ausencia_justificada_desde is None
                or convivente.ausencia_justificada_desde <= data_referencia
            )
        )
        dias_presenca_operacional = _calcular_dias_presenca_operacional(
            movimentos_por_convivente.get(convivente.id, []),
            data_referencia,
            data_referencia,
            convivente.data_entrada,
        )
        presente = bool(dias_presenca_operacional) or ausencia_justificada

        primeira_entrada = entradas[0]["data_registro"] if entradas else None
        ultima_saida = saidas[-1]["data_registro"] if saidas else None

        observacoes = []
        if ausencia_justificada:
            observacoes.append("Presença por justificativa: ausência justificada")

        if any(r["retorno_rapido"] for r in regs):
            observacoes.append("Retorno rápido")

        extras_cafe = _quantidade_extra_refeicao(cafes)
        extras_almoco = _quantidade_extra_refeicao(almocos)
        extras_jantar = _quantidade_extra_refeicao(jantares)
        extras_lanche = _quantidade_extra_refeicao(lanches)
        extras_refeicoes = (
            extras_cafe
            + extras_almoco
            + extras_jantar
            + extras_lanche
        )

        if extras_refeicoes:
            observacoes.append(
                "Refeições extras: "
                f"café {extras_cafe}, almoço {extras_almoco}, "
                f"jantar {extras_jantar}, lanche {extras_lanche}"
            )

        linhas.append({
            "convivente_id": convivente.id,
            "nome": _nome_convivente_relatorio(convivente),
            "nome_completo": convivente.nome_completo,
            "prontuario": convivente.numero_institucional,
            "numero_sisa": convivente.numero_sisa,
            "presenca": "Sim" if presente else "Não",
            "presenca_por_justificativa": "Sim" if ausencia_justificada else "Não",
            "entrada": primeira_entrada,
            "saida": ultima_saida,
            "almoco": "Sim" if almocos else "Não",
            "cafe": "Sim" if cafes else "Não",
            "jantar": "Sim" if jantares else "Não",
            "lanche": "Sim" if lanches else "Não",
            "cafes": len(cafes),
            "almocos": len(almocos),
            "jantares": len(jantares),
            "lanches": len(lanches),
            "cafes_extras": extras_cafe,
            "almocos_extras": extras_almoco,
            "jantares_extras": extras_jantar,
            "lanches_extras": extras_lanche,
            "refeicoes_extras": extras_refeicoes,
            "banho": "Sim" if banhos else "Não",
            "total_movimentos": len(regs),
            "observacoes": ", ".join(observacoes)
        })

    resumo = {
        "data": data,
        "conviventes_ativos": len(conviventes),
        "presentes": sum(1 for item in linhas if item["presenca"] == "Sim"),
        "presentes_por_justificativa": sum(1 for item in linhas if item.get("presenca_por_justificativa") == "Sim"),
        "ausentes": sum(1 for item in linhas if item["presenca"] == "Não"),
        "cafes": sum(item["cafes"] for item in linhas),
        "almocos": sum(item["almocos"] for item in linhas),
        "jantares": sum(item["jantares"] for item in linhas),
        "lanches": sum(item["lanches"] for item in linhas),
        "cafes_extras": sum(item["cafes_extras"] for item in linhas),
        "almocos_extras": sum(item["almocos_extras"] for item in linhas),
        "jantares_extras": sum(item["jantares_extras"] for item in linhas),
        "lanches_extras": sum(item["lanches_extras"] for item in linhas),
        "refeicoes_extras": sum(item["refeicoes_extras"] for item in linhas),
        "banhos": sum(1 for item in linhas if item["banho"] == "Sim"),
        "entradas": sum(len([
            r for r in regs
            if r["tipo_registro"] == "Entrada"
        ]) for regs in registros_por_convivente.values()),
        "saidas": sum(len([
            r for r in regs
            if r["tipo_registro"] == "Saída"
        ]) for regs in registros_por_convivente.values()),
        "retornos_rapidos": sum(
            1
            for regs in registros_por_convivente.values()
            for registro in regs
            if registro["retorno_rapido"]
        )
    }

    return {
        "resumo": resumo,
        "items": linhas
    }


@router.get("/convenio-sisa/mensal")
async def relatorio_sisa_mensal(
    ano: int,
    mes: int,
    tipo_atendimento: str = "todos",
    data_inicio: str = None,
    data_fim: str = None,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    inicio_mes, fim_mes = _inicio_fim_mes(ano, mes)
    inicio = inicio_mes
    fim = fim_mes

    if data_inicio:
        inicio = max(
            inicio_mes,
            datetime.combine(
                _parse_data_simples(data_inicio, "Data inicial").date(),
                datetime.min.time()
            )
        )

    if data_fim:
        fim = min(
            fim_mes,
            datetime.combine(
                _parse_data_simples(data_fim, "Data final").date(),
                datetime.max.time()
            )
        )

    conviventes = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                ConviventeDB.status.in_(["Ativo", "Ausência justificada"]),
            ).order_by(
                ConviventeDB.nome_completo.asc()
            )
        )
    ).scalars().all()

    registros_resultado = await db.execute(
        select(
            RegistroRotinaDB.convivente_id,
            RegistroRotinaDB.tipo_registro,
            RegistroRotinaDB.data_registro,
            RegistroRotinaDB.retorno_rapido
        ).where(
            RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            RegistroRotinaDB.cancelado != True,
            RegistroRotinaDB.data_registro >= inicio,
            RegistroRotinaDB.data_registro <= fim
        ).order_by(
            RegistroRotinaDB.data_registro.asc()
        )
    )

    registros_por_convivente = {}

    for convivente_id, tipo_registro, data_registro, retorno_rapido in registros_resultado.all():
        registros_por_convivente.setdefault(
            convivente_id,
            []
        ).append({
            "tipo_registro": tipo_registro,
            "data_registro": data_registro,
            "retorno_rapido": retorno_rapido,
        })

    convivente_ids = [convivente.id for convivente in conviventes]
    movimentos_resultado = await db.execute(
        select(
            RegistroRotinaDB.convivente_id,
            RegistroRotinaDB.tipo_registro,
            RegistroRotinaDB.data_registro,
        ).where(
            RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            RegistroRotinaDB.cancelado != True,
            RegistroRotinaDB.convivente_id.in_(convivente_ids),
            RegistroRotinaDB.tipo_registro.in_(["Entrada", "Saída"]),
            RegistroRotinaDB.data_registro <= fim,
        ).order_by(
            RegistroRotinaDB.data_registro.asc()
        )
    )

    movimentos_por_convivente = {}
    for convivente_id, tipo_registro, data_registro in movimentos_resultado.all():
        movimentos_por_convivente.setdefault(
            convivente_id,
            []
        ).append({
            "tipo_registro": tipo_registro,
            "data_registro": data_registro,
        })

    lancamentos_resultado = await db.execute(
        select(
            SisaLancamentoDB,
            UsuarioDB.nome.label("usuario_nome")
        )
        .join(UsuarioDB, SisaLancamentoDB.lancado_por_id == UsuarioDB.id)
        .where(
            SisaLancamentoDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            SisaLancamentoDB.ano == ano,
            SisaLancamentoDB.mes == mes
        )
    )

    lancamentos_por_convivente = {
        lancamento.convivente_id: {
            "id": lancamento.id,
            "status": lancamento.status,
            "lancado_por_id": lancamento.lancado_por_id,
            "lancado_por_nome": usuario_nome,
            "lancado_em": lancamento.lancado_em,
            "observacoes": lancamento.observacoes
        }
        for lancamento, usuario_nome in lancamentos_resultado.all()
    }

    linhas = []

    for convivente in conviventes:
        regs = registros_por_convivente.get(convivente.id, [])

        dias_presenca_operacional = set(_calcular_dias_presenca_operacional(
            movimentos_por_convivente.get(convivente.id, []),
            inicio.date(),
            fim.date(),
            convivente.data_entrada,
        ))
        dias_justificados = set()

        if convivente.status == "Ausência justificada":
            data_base_justificativa = convivente.ausencia_justificada_desde or inicio.date()
            dia = max(data_base_justificativa, inicio.date())
            fim_periodo = fim.date()

            while dia <= fim_periodo:
                dias_justificados.add(dia.isoformat())
                dia += timedelta(days=1)

        dias_presentes = sorted(dias_presenca_operacional | dias_justificados)

        entradas = [
            r for r in regs
            if r["tipo_registro"] == "Entrada"
        ]

        saidas = [
            r for r in regs
            if r["tipo_registro"] == "Saída"
        ]

        almocos = [
            r for r in regs
            if r["tipo_registro"] == "Almoço"
        ]
        cafes = [
            r for r in regs
            if r["tipo_registro"] == "Café da manhã"
        ]
        jantares = [
            r for r in regs
            if r["tipo_registro"] == "Jantar"
        ]
        lanches = [
            r for r in regs
            if r["tipo_registro"] == "Lanche noturno"
        ]
        banhos = [
            r for r in regs
            if r["tipo_registro"] == "Banho"
        ]

        lancamento_sisa = lancamentos_por_convivente.get(convivente.id)
        extras_cafe = _quantidade_extra_refeicao(cafes)
        extras_almoco = _quantidade_extra_refeicao(almocos)
        extras_jantar = _quantidade_extra_refeicao(jantares)
        extras_lanche = _quantidade_extra_refeicao(lanches)
        extras_refeicoes = (
            extras_cafe
            + extras_almoco
            + extras_jantar
            + extras_lanche
        )

        linhas.append({
            "convivente_id": convivente.id,
            "nome": _nome_convivente_relatorio(convivente),
            "nome_completo": convivente.nome_completo,
            "prontuario": convivente.numero_institucional,
            "numero_sisa": convivente.numero_sisa,
            "dias_presentes": len(dias_presentes),
            "dias_presentes_lista": dias_presentes,
            "dias_justificados": len(dias_justificados),
            "dias_justificados_lista": sorted(dias_justificados),
            "entradas": len(entradas),
            "saidas": len(saidas),
            "cafes": len(cafes),
            "almocos": len(almocos),
            "jantares": len(jantares),
            "lanches": len(lanches),
            "cafes_extras": extras_cafe,
            "almocos_extras": extras_almoco,
            "jantares_extras": extras_jantar,
            "lanches_extras": extras_lanche,
            "refeicoes_extras": extras_refeicoes,
            "banhos": len(banhos),
            "total_atendimentos": len(dias_presentes),
            "total_movimentos": len(regs),
            "retornos_rapidos": sum(1 for r in regs if r["retorno_rapido"]),
            "lancado_sisa": bool(lancamento_sisa),
            "lancamento_sisa_id": lancamento_sisa["id"] if lancamento_sisa else None,
            "lancado_por_id": lancamento_sisa["lancado_por_id"] if lancamento_sisa else None,
            "lancado_por_nome": lancamento_sisa["lancado_por_nome"] if lancamento_sisa else None,
            "lancado_em": lancamento_sisa["lancado_em"] if lancamento_sisa else None,
            "observacoes_lancamento_sisa": lancamento_sisa["observacoes"] if lancamento_sisa else None
        })

    linhas_filtradas = [
        item for item in linhas
        if _item_atende_filtro_sisa(item, tipo_atendimento)
    ]

    fechamento = (
        await db.execute(
            select(FechamentoMensalDB).where(
                FechamentoMensalDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                FechamentoMensalDB.ano == ano,
                FechamentoMensalDB.mes == mes
            )
        )
    ).scalar_one_or_none()

    mes_esta_fechado = bool(fechamento and fechamento.status == "Fechado")

    resumo = {
        "ano": ano,
        "mes": mes,
        "tipo_atendimento": tipo_atendimento,
        "data_inicio": inicio.date().isoformat(),
        "data_fim": fim.date().isoformat(),
        "conviventes_ativos": len(linhas_filtradas),
        "total_atendimentos": sum(item["total_atendimentos"] for item in linhas_filtradas),
        "total_justificativas": sum(item["dias_justificados"] for item in linhas_filtradas),
        "total_cafes": sum(item["cafes"] for item in linhas_filtradas),
        "total_almocos": sum(item["almocos"] for item in linhas_filtradas),
        "total_jantares": sum(item["jantares"] for item in linhas_filtradas),
        "total_lanches": sum(item["lanches"] for item in linhas_filtradas),
        "total_cafes_extras": sum(item["cafes_extras"] for item in linhas_filtradas),
        "total_almocos_extras": sum(item["almocos_extras"] for item in linhas_filtradas),
        "total_jantares_extras": sum(item["jantares_extras"] for item in linhas_filtradas),
        "total_lanches_extras": sum(item["lanches_extras"] for item in linhas_filtradas),
        "total_refeicoes_extras": sum(item["refeicoes_extras"] for item in linhas_filtradas),
        "total_banhos": sum(item["banhos"] for item in linhas_filtradas),
        "total_entradas": sum(item["entradas"] for item in linhas_filtradas),
        "total_saidas": sum(item["saidas"] for item in linhas_filtradas),
        "total_retornos_rapidos": sum(item["retornos_rapidos"] for item in linhas_filtradas),
        "lancados_sisa": sum(1 for item in linhas_filtradas if item.get("lancado_sisa")),
        "pendentes_sisa": sum(1 for item in linhas_filtradas if not item.get("lancado_sisa")),
        "fechado": mes_esta_fechado,
        "fechamento_id": fechamento.id if fechamento else None,
        "status_fechamento": fechamento.status if fechamento else None,
        "protocolo": fechamento.protocolo if fechamento else None,
        "fechado_em": fechamento.fechado_em if fechamento else None
    }

    return {
        "resumo": resumo,
        "items": linhas_filtradas
    }


@router.get("/convenio-sisa/fechamentos", response_model=FechamentoMensalListaResponse)
async def listar_fechamentos_mensais(
    limit: int = Query(REGISTROS_POR_PAGINA_PADRAO, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado),
):
    bloquear_usuario_global_puro(usuario_atual)
    instituicao_id = obter_instituicao_escopo(usuario_atual)
    base = select(FechamentoMensalDB).where(
        FechamentoMensalDB.instituicao_id == instituicao_id
    )
    total = int((await db.execute(
        select(func.count()).select_from(base.subquery())
    )).scalar_one() or 0)

    items = (
        await db.execute(
            base.order_by(
                FechamentoMensalDB.ano.desc(),
                FechamentoMensalDB.mes.desc(),
            )
            .offset(offset)
            .limit(limit)
        )
    ).scalars().all()

    return {
        "items": items,
        "total": total,
        "limit": limit,
        "offset": offset,
        "has_more": offset + len(items) < total,
    }


@router.post("/convenio-sisa/fechar-mes", response_model=FechamentoMensalResponse)
async def fechar_mes_convenio_sisa(
    payload: FechamentoMensalCreate,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    if not usuario_eh_gestor(usuario_atual):
        raise HTTPException(
            status_code=403,
            detail="Apenas gestão pode fechar o mês."
        )

    _inicio_fim_mes(payload.ano, payload.mes)

    fechamento_existente = (
        await db.execute(
            select(FechamentoMensalDB).where(
                FechamentoMensalDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                FechamentoMensalDB.ano == payload.ano,
                FechamentoMensalDB.mes == payload.mes
            )
        )
    ).scalar_one_or_none()

    protocolo = (
        f"SISA-{payload.ano}{str(payload.mes).zfill(2)}-"
        f"{agora_sao_paulo().strftime('%Y%m%d%H%M%S')}"
    )

    if fechamento_existente:
        if fechamento_existente.status == "Fechado":
            raise HTTPException(
                status_code=400,
                detail="Este mês já está fechado."
            )

        # Refecha um mês que havia sido reaberto, preservando auditoria.
        fechamento_existente.status = "Fechado"
        fechamento_existente.protocolo = protocolo
        fechamento_existente.fechado_por_id = usuario_atual["sub"]
        fechamento_existente.fechado_em = agora_sao_paulo()
        fechamento_existente.observacoes = payload.observacoes.strip() if payload.observacoes else None

        await db.commit()
        await db.refresh(fechamento_existente)

        return fechamento_existente

    fechamento = FechamentoMensalDB(
        instituicao_id=obter_instituicao_escopo(usuario_atual),
        ano=payload.ano,
        mes=payload.mes,
        protocolo=protocolo,
        fechado_por_id=usuario_atual["sub"],
        fechado_em=agora_sao_paulo(),
        observacoes=payload.observacoes.strip() if payload.observacoes else None
    )

    db.add(fechamento)

    await db.commit()
    await db.refresh(fechamento)

    return fechamento




@router.patch("/convenio-sisa/fechamentos/{fechamento_id}/reabrir", response_model=FechamentoMensalResponse)
async def reabrir_fechamento_convenio_sisa(
    fechamento_id: str,
    payload: FechamentoMensalReabertura,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    if not usuario_eh_gestor(usuario_atual):
        raise HTTPException(
            status_code=403,
            detail="Apenas gestão pode reabrir um mês fechado."
        )

    if not payload.motivo_reabertura or not payload.motivo_reabertura.strip():
        raise HTTPException(
            status_code=400,
            detail="Informe o motivo da reabertura."
        )

    fechamento = (
        await db.execute(
            select(FechamentoMensalDB).where(
                FechamentoMensalDB.id == fechamento_id,
                FechamentoMensalDB.instituicao_id == obter_instituicao_escopo(usuario_atual)
            )
        )
    ).scalar_one_or_none()

    if not fechamento:
        raise HTTPException(
            status_code=404,
            detail="Fechamento mensal não encontrado."
        )

    if fechamento.status != "Fechado":
        raise HTTPException(
            status_code=400,
            detail="Este mês não está fechado."
        )

    fechamento.status = "Reaberto"
    fechamento.reaberto_por_id = usuario_atual["sub"]
    fechamento.reaberto_em = agora_sao_paulo()
    fechamento.motivo_reabertura = payload.motivo_reabertura.strip()

    await db.commit()
    await db.refresh(fechamento)

    return fechamento


@router.post("/convenio-sisa/lancamentos", response_model=SisaLancamentoResponse)
async def marcar_lancamento_sisa(
    payload: SisaLancamentoCreate,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    fechamento = (
        await db.execute(
            select(FechamentoMensalDB).where(
                FechamentoMensalDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                FechamentoMensalDB.ano == payload.ano,
                FechamentoMensalDB.mes == payload.mes
            )
        )
    ).scalar_one_or_none()

    if fechamento and fechamento.status == "Fechado":
        raise HTTPException(
            status_code=400,
            detail="Mês fechado. Não é possível alterar lançamentos SISA."
        )

    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == payload.convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual)
            )
        )
    ).scalar_one_or_none()

    if not convivente:
        raise HTTPException(
            status_code=404,
            detail="Convivente não encontrado."
        )

    exigir_convivente_ativo_para_registro(convivente)

    existente = (
        await db.execute(
            select(SisaLancamentoDB).where(
                SisaLancamentoDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                SisaLancamentoDB.ano == payload.ano,
                SisaLancamentoDB.mes == payload.mes,
                SisaLancamentoDB.convivente_id == payload.convivente_id
            )
        )
    ).scalar_one_or_none()

    if existente:
        existente.status = "Lancado"
        existente.lancado_por_id = usuario_atual["sub"]
        existente.lancado_em = agora_sao_paulo()
        existente.observacoes = payload.observacoes.strip() if payload.observacoes else None

        await db.commit()
        await db.refresh(existente)

        return existente

    lancamento = SisaLancamentoDB(
        instituicao_id=obter_instituicao_escopo(usuario_atual),
        ano=payload.ano,
        mes=payload.mes,
        convivente_id=payload.convivente_id,
        status="Lancado",
        lancado_por_id=usuario_atual["sub"],
        lancado_em=agora_sao_paulo(),
        observacoes=payload.observacoes.strip() if payload.observacoes else None
    )

    db.add(lancamento)

    await db.commit()
    await db.refresh(lancamento)

    return lancamento


@router.delete("/convenio-sisa/lancamentos/{lancamento_id}")
async def desfazer_lancamento_sisa(
    lancamento_id: str,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    lancamento = (
        await db.execute(
            select(SisaLancamentoDB).where(
                SisaLancamentoDB.id == lancamento_id,
                SisaLancamentoDB.instituicao_id == obter_instituicao_escopo(usuario_atual)
            )
        )
    ).scalar_one_or_none()

    if not lancamento:
        raise HTTPException(
            status_code=404,
            detail="Lançamento SISA não encontrado."
        )

    fechamento = (
        await db.execute(
            select(FechamentoMensalDB).where(
                FechamentoMensalDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
                FechamentoMensalDB.ano == lancamento.ano,
                FechamentoMensalDB.mes == lancamento.mes
            )
        )
    ).scalar_one_or_none()

    if fechamento and fechamento.status == "Fechado":
        raise HTTPException(
            status_code=400,
            detail="Mês fechado. Não é possível desfazer lançamento SISA."
        )

    await db.delete(lancamento)
    await db.commit()

    return {"status": "sucesso"}


@router.get("/convenio-sisa/mensal/exportar-xlsx")
async def exportar_sisa_mensal_xlsx(
    ano: int,
    mes: int,
    tipo_atendimento: str = "todos",
    data_inicio: str = None,
    data_fim: str = None,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    dados = await relatorio_sisa_mensal(
        ano=ano,
        mes=mes,
        tipo_atendimento=tipo_atendimento,
        data_inicio=data_inicio,
        data_fim=data_fim,
        db=db,
        usuario_atual=usuario_atual
    )

    arquivo = gerar_xlsx_convenio_sisa_mensal(dados)

    nome_arquivo = f"relatorio_sisa_convenio_{ano}_{str(mes).zfill(2)}.xlsx"

    return StreamingResponse(
        BytesIO(arquivo),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"'
        }
    )

# =====================================================================
# EDIÇÃO E CANCELAMENTO DE REGISTROS DA ROTINA
# =====================================================================

async def verificar_permissao_edicao(
    db: AsyncSession,
    usuario_atual: dict,
    registro: RegistroRotinaDB
):

    if usuario_eh_gestor(usuario_atual):
        return

    convivente = (
        await db.execute(
            select(ConviventeDB).where(
                ConviventeDB.id == registro.convivente_id,
                ConviventeDB.instituicao_id == obter_instituicao_escopo(usuario_atual),
            )
        )
    ).scalar_one_or_none()

    if (
        convivente and
        str(convivente.tecnico_id) ==
        str(usuario_atual["sub"])
    ):
        return

    raise HTTPException(
        status_code=403,
        detail="Sem permissão para editar este registro."
    )


@router.put("/rotina/{registro_id}", response_model=RegistroRotinaResponse)
async def editar_registro_rotina(
    registro_id: str,
    payload: RegistroRotinaEdicao,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    if payload.tipo_registro not in TIPOS_ROTINA_VALIDOS:
        raise HTTPException(
            status_code=400,
            detail="Tipo de registro inválido."
        )

    if not payload.motivo_edicao or not payload.motivo_edicao.strip():
        raise HTTPException(
            status_code=400,
            detail="Informe o motivo da edição."
        )

    observacao = (payload.observacao or "").strip()
    if payload.tipo_registro in TIPOS_ROTINA_COM_OBSERVACAO_OBRIGATORIA and not observacao:
        raise HTTPException(
            status_code=400,
            detail="Este tipo de interação exige relato/especificação.",
        )

    registro = (
        await db.execute(
            select(RegistroRotinaDB).where(
                RegistroRotinaDB.id == registro_id,
                RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual)
            )
        )
    ).scalar_one_or_none()

    if not registro:
        raise HTTPException(
            status_code=404,
            detail="Registro de rotina não encontrado."
        )

    await verificar_permissao_edicao(
        db,
        usuario_atual,
        registro
    )

    await verificar_mes_fechado(
        db,
        usuario_atual,
        registro.data_registro,
        acao="editar registro de rotina"
    )

    if registro.cancelado:
        raise HTTPException(
            status_code=400,
            detail="Não é possível editar um registro cancelado."
        )

    if not registro.foi_editado:
        registro.tipo_registro_original = registro.tipo_registro
        registro.data_registro_original = registro.data_registro

    registro.tipo_registro = payload.tipo_registro
    registro.observacao = observacao or None
    registro.foi_editado = True
    registro.editado_por_id = usuario_atual["sub"]
    registro.editado_em = agora_sao_paulo()
    registro.motivo_edicao = payload.motivo_edicao.strip()

    await db.commit()
    await db.refresh(registro)

    return registro


@router.patch("/rotina/{registro_id}/cancelar", response_model=RegistroRotinaResponse)
async def cancelar_registro_rotina(
    registro_id: str,
    payload: RegistroRotinaCancelamento,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)
    if not payload.motivo_cancelamento or not payload.motivo_cancelamento.strip():
        raise HTTPException(
            status_code=400,
            detail="Informe o motivo do cancelamento."
        )

    registro = (
        await db.execute(
            select(RegistroRotinaDB).where(
                RegistroRotinaDB.id == registro_id,
                RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual)
            )
        )
    ).scalar_one_or_none()

    if not registro:
        raise HTTPException(
            status_code=404,
            detail="Registro de rotina não encontrado."
        )

    await verificar_permissao_edicao(
        db,
        usuario_atual,
        registro
    )

    await verificar_mes_fechado(
        db,
        usuario_atual,
        registro.data_registro,
        acao="cancelar registro de rotina"
    )

    if registro.cancelado:
        raise HTTPException(
            status_code=400,
            detail="Este registro já está cancelado."
        )

    registro.cancelado = True
    registro.cancelado_por_id = usuario_atual["sub"]
    registro.cancelado_em = agora_sao_paulo()
    registro.motivo_cancelamento = payload.motivo_cancelamento.strip()

    await db.commit()
    await db.refresh(registro)

    return registro

@router.patch(
    "/rotina/{registro_id}/desfazer",
    response_model=RegistroRotinaResponse
)
async def desfazer_registro_rapido(
    registro_id: str,
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    bloquear_usuario_global_puro(usuario_atual)

    registro = (
        await db.execute(
            select(RegistroRotinaDB).where(
                RegistroRotinaDB.id == registro_id,
                RegistroRotinaDB.instituicao_id == obter_instituicao_escopo(usuario_atual)
            )
        )
    ).scalar_one_or_none()

    if not registro:

        raise HTTPException(
            status_code=404,
            detail="Registro não encontrado."
        )

    if registro.cancelado:

        raise HTTPException(
            status_code=400,
            detail="Registro já cancelado."
        )

    if str(registro.usuario_id) != str(usuario_atual["sub"]):

        raise HTTPException(
            status_code=403,
            detail="Apenas o operador original pode desfazer."
        )

    diferenca = (
        agora_sao_paulo() -
        registro.data_registro
    )

    if diferenca > timedelta(minutes=1):

        raise HTTPException(
            status_code=403,
            detail="Prazo de desfazer expirado."
        )

    await verificar_mes_fechado(
        db,
        usuario_atual,
        registro.data_registro,
        acao="desfazer registro de rotina"
    )

    registro.cancelado = True

    registro.cancelado_por_id = usuario_atual["sub"]

    registro.cancelado_em = agora_sao_paulo()

    registro.motivo_cancelamento = (
        "Erro operacional - correção imediata"
    )

    await db.commit()

    await db.refresh(registro)

    return registro

# =====================================================================
# DASHBOARD — SÉRIES REAIS DE ATENDIMENTOS
# =====================================================================

def _inicio_mes_referencia(data_base: datetime, meses_atras: int = 0) -> datetime:
    ano = data_base.year
    mes = data_base.month - meses_atras

    while mes <= 0:
        mes += 12
        ano -= 1

    return datetime(ano, mes, 1)


def _chave_semana(data: datetime):
    ano, semana, _ = data.isocalendar()
    return f"{ano}-S{str(semana).zfill(2)}"


def _series_dashboard_vazias(agora: datetime):
    hoje = agora.date()

    meses = []
    for offset in range(5, -1, -1):
        inicio_mes = _inicio_mes_referencia(agora, offset)
        meses.append({
            "chave": inicio_mes.strftime("%Y-%m"),
            "rotulo": f"{str(inicio_mes.month).zfill(2)}/{inicio_mes.year}",
            "atendimentos": 0,
            "novos_conviventes": 0
        })

    dias = []
    for offset in range(6, -1, -1):
        dia = hoje - timedelta(days=offset)
        dias.append({
            "chave": dia.isoformat(),
            "rotulo": dia.strftime("%d/%m"),
            "atendimentos": 0
        })

    semanas = []
    for offset in range(5, -1, -1):
        data_ref = hoje - timedelta(weeks=offset)
        chave = _chave_semana(datetime.combine(data_ref, datetime.min.time()))
        semanas.append({
            "chave": chave,
            "rotulo": chave.replace("-", " "),
            "atendimentos": 0
        })

    return meses, dias, semanas


async def _dashboard_series_otimizada(db: AsyncSession, inst_id: str, agora: datetime):
    hoje = agora.date()
    inicio_6_meses = _inicio_mes_referencia(agora, 5)
    meses, dias, semanas = _series_dashboard_vazias(agora)

    indice_meses = {item["chave"]: item for item in meses}
    indice_dias = {item["chave"]: item for item in dias}
    indice_semanas = {item["chave"]: item for item in semanas}

    registros_por_dia = await db.execute(
        select(
            func.date(RegistroRotinaDB.data_registro).label("dia"),
            func.count(RegistroRotinaDB.id).label("total")
        ).where(
            RegistroRotinaDB.instituicao_id == inst_id,
            RegistroRotinaDB.cancelado != True,
            RegistroRotinaDB.data_registro >= inicio_6_meses
        ).group_by(func.date(RegistroRotinaDB.data_registro))
    )

    for dia_registro, total in registros_por_dia.all():
        if not dia_registro:
            continue

        if isinstance(dia_registro, str):
            data_registro = datetime.fromisoformat(dia_registro)
        else:
            data_registro = datetime.combine(dia_registro, datetime.min.time())

        chave_mes = data_registro.strftime("%Y-%m")
        if chave_mes in indice_meses:
            indice_meses[chave_mes]["atendimentos"] += total

        chave_dia = data_registro.date().isoformat()
        if chave_dia in indice_dias:
            indice_dias[chave_dia]["atendimentos"] += total

        chave_semana = _chave_semana(data_registro)
        if chave_semana in indice_semanas:
            indice_semanas[chave_semana]["atendimentos"] += total

    conviventes_periodo = await db.execute(
        select(ConviventeDB.data_entrada).where(
            ConviventeDB.instituicao_id == inst_id,
            ConviventeDB.data_entrada != None,
            ConviventeDB.data_entrada >= inicio_6_meses.date()
        )
    )

    for data_entrada in conviventes_periodo.scalars().all():
        if not data_entrada:
            continue

        chave = data_entrada.strftime("%Y-%m")
        if chave in indice_meses:
            indice_meses[chave]["novos_conviventes"] += 1

    total_hoje = indice_dias.get(hoje.isoformat(), {}).get("atendimentos", 0)
    semana_atual_chave = _chave_semana(datetime.combine(hoje, datetime.min.time()))
    total_semana = indice_semanas.get(semana_atual_chave, {}).get("atendimentos", 0)
    mes_atual_chave = agora.strftime("%Y-%m")
    total_mes = indice_meses.get(mes_atual_chave, {}).get("atendimentos", 0)

    return {
        "mensal_6_meses": meses,
        "diario_7_dias": dias,
        "semanal_6_semanas": semanas,
        "resumo": {
            "atendimentos_hoje": total_hoje,
            "atendimentos_semana": total_semana,
            "atendimentos_mes": total_mes,
            "novos_conviventes_mes": indice_meses.get(mes_atual_chave, {}).get("novos_conviventes", 0)
        }
    }


async def _resumo_ocorrencias_dashboard(
    db: AsyncSession,
    inst_id: str,
    perfil: str | None,
    user_id: str | None,
):
    base = select(OcorrenciaConviventeDB).where(OcorrenciaConviventeDB.instituicao_id == inst_id)
    if perfil == "Orientador":
        subq = select(ObservadorOcorrenciaDB.ocorrencia_id).where(
            ObservadorOcorrenciaDB.usuario_id == user_id
        )
        base = base.where(
            or_(
                OcorrenciaConviventeDB.usuario_criador_id == user_id,
                OcorrenciaConviventeDB.id.in_(subq),
            )
        )

    def contar(expressao=None):
        consulta = base
        if expressao is not None:
            consulta = consulta.where(expressao)
        return select(func.count()).select_from(consulta.order_by(None).subquery())

    ocorrencias_pendentes = int((await db.execute(
        contar(OcorrenciaConviventeDB.status_resolucao != "Resolvido")
    )).scalar_one() or 0)

    ocorrencias_criticas_altas_pendentes = int((await db.execute(
        contar(
            and_(
                OcorrenciaConviventeDB.status_resolucao != "Resolvido",
                OcorrenciaConviventeDB.prioridade.in_(["Alta", "Crítica"]),
            )
        )
    )).scalar_one() or 0)

    resumo_prioridades = {}
    for prioridade in PRIORIDADES_OCORRENCIA:
        total_prioridade = int((await db.execute(
            contar(OcorrenciaConviventeDB.prioridade == prioridade)
        )).scalar_one() or 0)
        pendentes_prioridade = int((await db.execute(
            contar(
                and_(
                    OcorrenciaConviventeDB.prioridade == prioridade,
                    OcorrenciaConviventeDB.status_resolucao != "Resolvido",
                )
            )
        )).scalar_one() or 0)
        resumo_prioridades[prioridade] = {
            "total": total_prioridade,
            "pendentes": pendentes_prioridade,
        }

    peso_sql = case(
        (OcorrenciaConviventeDB.prioridade == "Crítica", 4),
        (OcorrenciaConviventeDB.prioridade == "Alta", 3),
        (OcorrenciaConviventeDB.prioridade == "Média", 2),
        else_=1,
    )

    alertas_db = (
        await db.execute(
            base.where(OcorrenciaConviventeDB.status_resolucao != "Resolvido")
            .order_by(peso_sql.desc(), OcorrenciaConviventeDB.data_ocorrencia.desc())
            .limit(10)
        )
    ).scalars().all()

    ocorrencias_em_alerta = [
        {
            "id": ocorrencia.id,
            "tecnico_responsavel_id": ocorrencia.tecnico_responsavel_id,
            "tipo_ocorrencia": ocorrencia.tipo_ocorrencia,
            "motivo": ocorrencia.motivo,
            "descricao": ocorrencia.descricao,
            "data_ocorrencia": ocorrencia.data_ocorrencia,
            "requer_acao_tecnica": ocorrencia.requer_acao_tecnica,
            "status_resolucao": ocorrencia.status_resolucao,
            "prioridade": normalizar_prioridade_ocorrencia(ocorrencia.prioridade),
        }
        for ocorrencia in alertas_db
    ]

    tecnico_expr = func.coalesce(OcorrenciaConviventeDB.tecnico_responsavel_id, "sem_tecnico")
    filtro_pendencias = and_(
        OcorrenciaConviventeDB.status_resolucao != "Resolvido",
        or_(
            OcorrenciaConviventeDB.requer_acao_tecnica == True,
            OcorrenciaConviventeDB.tecnico_responsavel_id.isnot(None),
            OcorrenciaConviventeDB.prioridade.in_(["Alta", "Crítica"]),
        ),
    )

    query_grupo = (
        select(
            tecnico_expr.label("tecnico_id"),
            func.count().label("pendentes"),
            func.sum(case((OcorrenciaConviventeDB.prioridade == "Alta", 1), else_=0)).label("alta"),
            func.sum(case((OcorrenciaConviventeDB.prioridade == "Crítica", 1), else_=0)).label("critica"),
        )
        .select_from(OcorrenciaConviventeDB)
        .where(OcorrenciaConviventeDB.instituicao_id == inst_id)
        .where(filtro_pendencias)
    )
    if perfil == "Orientador":
        subq = select(ObservadorOcorrenciaDB.ocorrencia_id).where(
            ObservadorOcorrenciaDB.usuario_id == user_id
        )
        query_grupo = query_grupo.where(
            or_(
                OcorrenciaConviventeDB.usuario_criador_id == user_id,
                OcorrenciaConviventeDB.id.in_(subq),
            )
        )

    grupos = (await db.execute(query_grupo.group_by(tecnico_expr))).all()

    tecnico_ids = [
        grupo.tecnico_id
        for grupo in grupos
        if grupo.tecnico_id and grupo.tecnico_id != "sem_tecnico"
    ]
    tecnicos = {}
    if tecnico_ids:
        tecnicos = {
            usuario.id: usuario
            for usuario in (
                await db.execute(
                    select(UsuarioDB).where(
                        UsuarioDB.instituicao_id == inst_id,
                        UsuarioDB.id.in_(tecnico_ids),
                    )
                )
            ).scalars().all()
        }

    pendencias_tecnicos = []
    for grupo in grupos:
        tecnico_id = grupo.tecnico_id or "sem_tecnico"
        tecnico = tecnicos.get(tecnico_id)
        alta = int(grupo.alta or 0)
        critica = int(grupo.critica or 0)
        pendentes = int(grupo.pendentes or 0)
        maior_prioridade = (
            "Crítica" if critica > 0
            else "Alta" if alta > 0
            else "Baixa"
        )
        prioridade = (
            "Crítica" if critica > 0
            else "Alta" if alta > 0
            else "Crítico" if pendentes >= 5
            else "Médio" if pendentes >= 3
            else maior_prioridade
        )
        pendencias_tecnicos.append({
            "tecnico_id": tecnico_id,
            "tecnico_nome": tecnico.nome if tecnico else "Sem técnico definido",
            "tecnico_avatar_url": tecnico.avatar_url if tecnico else "",
            "perfil": tecnico.perfil_acesso if tecnico else "Técnico",
            "pendentes": pendentes,
            "alta": alta,
            "critica": critica,
            "maiorPrioridade": maior_prioridade,
            "prioridade": prioridade,
        })

    pendencias_tecnicos.sort(
        key=lambda item: (
            PESO_PRIORIDADE.get(normalizar_prioridade_ocorrencia(item["prioridade"]), 2),
            item["pendentes"],
        ),
        reverse=True,
    )
    pendencias_tecnicos = pendencias_tecnicos[:8]

    return {
        "ocorrencias_pendentes": ocorrencias_pendentes,
        "ocorrencias_criticas_altas_pendentes": ocorrencias_criticas_altas_pendentes,
        "resumo_prioridades": resumo_prioridades,
        "ocorrencias_em_alerta": ocorrencias_em_alerta,
        "pendencias_tecnicos": pendencias_tecnicos,
    }


@router.get("/dashboard/resumo")
async def dashboard_resumo(
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    inst_id = obter_instituicao_escopo(usuario_atual)
    perfil = usuario_atual.get("perfil_acesso")
    user_id = usuario_atual.get("sub")
    agora = agora_sao_paulo()

    total_conviventes = (await db.execute(
        select(func.count(ConviventeDB.id)).where(
            ConviventeDB.instituicao_id == inst_id
        )
    )).scalar_one()

    ativos = (await db.execute(
        select(func.count(ConviventeDB.id)).where(
            ConviventeDB.instituicao_id == inst_id,
            ConviventeDB.status == "Ativo"
        )
    )).scalar_one()

    total_leitos = (await db.execute(
        select(func.count(LeitoDB.id))
        .join(QuartoDB, QuartoDB.id == LeitoDB.quarto_id)
        .where(QuartoDB.instituicao_id == inst_id)
    )).scalar_one()

    leitos_ocupados = (await db.execute(
        select(func.count(LeitoDB.id))
        .join(QuartoDB, QuartoDB.id == LeitoDB.quarto_id)
        .where(
            QuartoDB.instituicao_id == inst_id,
            LeitoDB.status == "Ocupado"
        )
    )).scalar_one()

    resumo_ocorrencias = await _resumo_ocorrencias_dashboard(db, inst_id, perfil, user_id)
    series = await _dashboard_series_otimizada(db, inst_id, agora)

    return {
        "totalConviventes": total_conviventes,
        "ativos": ativos,
        "leitosOcupados": leitos_ocupados,
        "totalLeitos": total_leitos,
        "atendimentosMes": series["resumo"]["atendimentos_mes"],
        "atendimentosHoje": series["resumo"]["atendimentos_hoje"],
        "ocorrenciasPendentes": resumo_ocorrencias["ocorrencias_pendentes"],
        "ocorrenciasCriticasAltasPendentes": resumo_ocorrencias["ocorrencias_criticas_altas_pendentes"],
        "alertasOcorrencias": resumo_ocorrencias["ocorrencias_pendentes"],
        "pendenciasTecnicas": sum(item["pendentes"] for item in resumo_ocorrencias["pendencias_tecnicos"]),
        "ocorrenciasEmAlerta": resumo_ocorrencias["ocorrencias_em_alerta"],
        "pendenciasTecnicos": resumo_ocorrencias["pendencias_tecnicos"],
        "resumoPrioridades": resumo_ocorrencias["resumo_prioridades"],
        "series": series,
    }


@router.get("/dashboard/series-atendimentos")
async def dashboard_series_atendimentos(
    db: AsyncSession = Depends(get_db),
    usuario_atual: dict = Depends(get_usuario_logado)
):
    """
    Séries reais para o Dashboard.

    Fonte:
    - registros_rotina não cancelados;
    - agrupamento por dia, semana e mês;
    - atendimento = registro operacional válido de rotina.
    """
    inst_id = obter_instituicao_escopo(usuario_atual)
    agora = agora_sao_paulo()
    hoje = agora.date()

    inicio_7_dias = datetime.combine(
        hoje - timedelta(days=6),
        datetime.min.time()
    )

    inicio_6_semanas = datetime.combine(
        hoje - timedelta(weeks=5),
        datetime.min.time()
    )

    inicio_6_meses = _inicio_mes_referencia(agora, 5)

    registros = (
        await db.execute(
            select(RegistroRotinaDB).where(
                RegistroRotinaDB.instituicao_id == inst_id,
                RegistroRotinaDB.cancelado != True,
                RegistroRotinaDB.data_registro >= inicio_6_meses
            )
        )
    ).scalars().all()

    conviventes_criados_por_mes = {
        _inicio_mes_referencia(agora, offset).strftime("%Y-%m"): 0
        for offset in range(5, -1, -1)
    }

    try:
        conviventes_periodo = (
            await db.execute(
                select(ConviventeDB).where(
                    ConviventeDB.instituicao_id == inst_id,
                    ConviventeDB.data_entrada != None,
                    ConviventeDB.data_entrada >= inicio_6_meses.date()
                )
            )
        ).scalars().all()

        for convivente in conviventes_periodo:
            if convivente.data_entrada:
                chave = convivente.data_entrada.strftime("%Y-%m")
                if chave in conviventes_criados_por_mes:
                    conviventes_criados_por_mes[chave] += 1
    except Exception:
        pass

    # Últimos 6 meses
    meses = []
    for offset in range(5, -1, -1):
        inicio_mes = _inicio_mes_referencia(agora, offset)
        chave = inicio_mes.strftime("%Y-%m")
        rotulo = f"{str(inicio_mes.month).zfill(2)}/{inicio_mes.year}"
        meses.append({
            "chave": chave,
            "rotulo": rotulo,
            "atendimentos": 0,
            "novos_conviventes": conviventes_criados_por_mes.get(chave, 0)
        })

    indice_meses = {item["chave"]: item for item in meses}

    # Últimos 7 dias
    dias = []
    for offset in range(6, -1, -1):
        dia = hoje - timedelta(days=offset)
        dias.append({
            "chave": dia.isoformat(),
            "rotulo": dia.strftime("%d/%m"),
            "atendimentos": 0
        })

    indice_dias = {item["chave"]: item for item in dias}

    # Últimas 6 semanas
    semanas = []
    semanas_chaves = []
    for offset in range(5, -1, -1):
        data_ref = hoje - timedelta(weeks=offset)
        chave = _chave_semana(datetime.combine(data_ref, datetime.min.time()))
        semanas_chaves.append(chave)
        semanas.append({
            "chave": chave,
            "rotulo": chave.replace("-", " "),
            "atendimentos": 0
        })

    indice_semanas = {item["chave"]: item for item in semanas}

    for registro in registros:
        data_registro = registro.data_registro
        if not data_registro:
            continue

        chave_mes = data_registro.strftime("%Y-%m")
        if chave_mes in indice_meses:
            indice_meses[chave_mes]["atendimentos"] += 1

        chave_dia = data_registro.date().isoformat()
        if chave_dia in indice_dias:
            indice_dias[chave_dia]["atendimentos"] += 1

        chave_semana = _chave_semana(data_registro)
        if chave_semana in indice_semanas:
            indice_semanas[chave_semana]["atendimentos"] += 1

    total_hoje = indice_dias.get(hoje.isoformat(), {}).get("atendimentos", 0)

    semana_atual_chave = _chave_semana(datetime.combine(hoje, datetime.min.time()))
    total_semana = indice_semanas.get(semana_atual_chave, {}).get("atendimentos", 0)

    mes_atual_chave = agora.strftime("%Y-%m")
    total_mes = indice_meses.get(mes_atual_chave, {}).get("atendimentos", 0)

    return {
        "mensal_6_meses": meses,
        "diario_7_dias": dias,
        "semanal_6_semanas": semanas,
        "resumo": {
            "atendimentos_hoje": total_hoje,
            "atendimentos_semana": total_semana,
            "atendimentos_mes": total_mes,
            "novos_conviventes_mes": indice_meses.get(mes_atual_chave, {}).get("novos_conviventes", 0)
        }
    }